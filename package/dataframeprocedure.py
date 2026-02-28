# -*- coding: utf-8 -*-
"""
Created on Tue Feb  2 17:35:04 2021

@author: ian.ko
"""
import pickle
try:
    from keras import backend as bcd #转换为张量
    from keras.models import load_model
except:
    print('keras imported failed!!!')
from sklearn import metrics as skm
from sklearn.model_selection import GridSearchCV
from sklearn import datasets
from sklearn.svm import SVR
from sklearn.model_selection import train_test_split
from sklearn.cluster import DBSCAN
from sklearn.neighbors import KNeighborsClassifier
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import SelectPercentile
from sklearn.feature_selection import chi2
import random as rdm
import matplotlib.pyplot as plt
import numpy as np
from datetime import date
import os
import pandas as pd
from mpl_toolkits import mplot3d
from datetime import datetime as dt
import datetime as dtC
import sys
import csv
import statistics as stc
import openpyxl
import joblib
from sklearn.preprocessing import StandardScaler as Stdscr
from scipy import signal
from scipy.optimize import curve_fit
from scipy.optimize import Bounds
from scipy.optimize import minimize
from scipy.interpolate import interp1d
from scipy.fft import fft, ifft, fftfreq
from scipy import stats
from scipy.stats import pointbiserialr
from scipy.stats import shapiro
from scipy.stats import kstest
from scipy.stats import chi2 as chi2_stats
from scipy.stats import linregress
from scipy.ndimage import gaussian_filter1d as gsf1d
from copy import deepcopy as dcp
import threading
# import pytesseract as pyt
from package import LOGger
stamp_process = LOGger.stamp_process
mylist = LOGger.mylist
parse = LOGger.parse
type_string = LOGger.type_string
dtp = LOGger.dtp
m_show_detail = True
m_encap_out_type = pd.core.frame.DataFrame
showlog = 0
empty_print = 0
def printer(*logs):
    if(showlog):
        logs = logs if(empty_print) else tuple([log for log in logs if log!=''])
        print(*logs)
#%%
m_noise_values = [np.nan, None, '', 'nan']
m_uniq_thru_set=lambda tensor, **kwags:sorted(list(set(tensor)), **kwags)
m_uniq_thru_np=lambda tensor, **kwags:np.unique(tensor, **kwags)
m_allTensors = {}
m_allTensors['shape12'] = np.random.random((12,))
m_allTensors['shape12Series'] = pd.Series(dcp(m_allTensors['shape12']))
m_allTensors['shape12_1'] = m_allTensors['shape12'].reshape(-1,1)
m_allTensors['shape12List'] = m_allTensors['shape12'].tolist()
m_allTensors['listOfShape12'] = [dcp(m_allTensors['shape12'])]
m_allTensors['shape12_2'] = np.random.random((12,2))
m_allTensors['shape12_2List'] = m_allTensors['shape12_2'].tolist()
m_allTensors['shape3_4_2'] = np.random.random((3,4,2))
m_allTensors['shape3_FrameOf2'] = pd.DataFrame(np.random.random((3,1,2)).tolist())
m_allTensors['shape3_FrameOfnp2_4'] = pd.DataFrame(np.random.random((3,1,2,4)).tolist())
m_allTensors['shape3_FrameOf2_5_3'] = pd.DataFrame(np.random.random((3,1,2,5,3)).tolist())
m_allTensors['shape3_FrameOfnp2_5_3'] = pd.DataFrame([[np.array(x) for x in m_allTensors['shape3_FrameOf2_5_3'][0]]]).T
#TODO:type_transform
path_sep = os.path.join('a','b')[1]
np2pd = lambda d: pd.DataFrame(d)
pd2np = lambda d: np.array(d)
np2ls = lambda d: [v for v in d]
ls2np = lambda d: np.array(d)
idfcn = lambda d: d


def append(data1, data2, sort=False, **kwags):
    if(int(pd.__version__[0])==1):
        data = data1.append(data2, sort=sort, **kwags).copy()
    else:
        data = pd.concat([data1, data2], sort=sort, axis=0, **kwags)
    return data

def join(data1, data2, sort=False, **kwags):
    if(int(pd.__version__[0])==1):
        data = data1.join(data2, sort=sort, **kwags).copy()
    else:
        data = pd.concat([data1, data2], sort=sort, axis=1, **kwags)
    return data

def type_transforming_in_method(
        method, type_transform=(idfcn, idfcn), vrb_named=False, 
        not_trsfming_method_output_mask=[], **method_inputs):
    method_inputs_consts = method_inputs.pop('method_inputs_consts', {})
    trsfmed_method_inputs = {}
    for data_key, trsfming_data in method_inputs.items():
        trsfmed_data = type_transform[0](trsfming_data)
        trsfmed_method_inputs[data_key] = dcp(trsfmed_data)
    trsfmed_method_inputs.update(method_inputs_consts)
    if(vrb_named):
        trsfming_method_outputs = method(**trsfmed_method_inputs)
        trsfmed_method_outputs = {}
        for data_key, trsfming_method_output in trsfming_method_outputs.items():
            trsfmed_method_outputs[data_key] = type_transform[1](trsfming_method_output)
    else:
        trsfming_method_outputs = [*(method(**trsfmed_method_inputs))] if(
                not str(type(method(**trsfmed_method_inputs))).find('pandas')>-1 and 
                not str(type(method(**trsfmed_method_inputs))).find('ndarray')>-1) else [
                method(**trsfmed_method_inputs)]
        new_datas = []
        for i, method_output in enumerate(trsfming_method_outputs):
            new_datas += [type_transform[1](method_output)] if(
                    not i in not_trsfming_method_output_mask) else [method_output]
    return new_datas if(vrb_named) else (tuple(new_datas) if(len(new_datas)>1) else new_datas[0])

def ary2dict(array, keys=[]):
    dc_array = {}
    if(len(np.array(array).shape)>0):
        n_array = np.array(array).shape[0]
        dc_array = {((keys[k] if(keys[k]!=None) else k) if(
                k<len(keys)) else k):array[k] for k in range(n_array)}
    return dc_array

#TODO:astype(maxdot, d_type=lambda x:x+0, default=0)
def astype(*args, criterion=None, d_type=float, default=None, default_method=None, is_dearray=True, **kwags):
    kwags.update({k:v for k,v in locals().items() if k!='args' and k!='kwags'})
    def method(arg, **kwags):
        try:
            arg = d_type(arg)
        except:
            arg = default_method(arg) if(default_method!=None) else default if(default!='self') else arg
        return arg
    if(is_dearray):
        ret = dearray_process(*args, method=method, **kwags)
    else:
        ret = [method(*args)]
    return ret[0] if(len(ret)==1) else (default if(len(ret)==0) else ret)

#TODO:astype_or_remain
def astype_or_remain(*args, criterion=None, d_type=float, **kwags):
    kwags['default_method'] = lambda x:x
    return astype(*args, criterion=criterion, d_type=d_type, **kwags)

#TODO:astype_or_remain_strictly
def astype_or_remain_indearray(*args, criterion=None, d_type=float, **kwags):
    kwags['default_method'] = lambda x:x
    return astype(*args, criterion=criterion, d_type=d_type, is_dearray=False, **kwags)

#TODO:astype_or_remain
def astype_datetime_float_or_remain(*args, criterion=None, **kwags):
    kwags['default_method'] = lambda x:x
    def astype_to_datetime_or_float(a):
        # 修正: 先檢查 pd.Timestamp，再檢查 dtC.datetime，最後檢查 dtC.date
        if isinstance(a, pd.Timestamp):
            return a
        elif isinstance(a, dtC.datetime):
            return a
        elif isinstance(a, dtC.date) or isinstance(a, dtC.time):
            return a
        else:
            try:
                return float(a)
            except:
                return a
    return astype(*args, criterion=criterion, d_type=astype_to_datetime_or_float, **kwags)

def datetime_to_float(x, default=None, **kwags):
    # 導入 DateTimeHandler
    try:
        from package.DateTimeHandler import DateTimeHandler
        # 使用 DateTimeHandler 如果可用
        return DateTimeHandler.to_float(x, default, **kwags)
    except ImportError:
        # 如果找不到 DateTimeHandler 模塊，使用原有的實現
        try:
            # 處理 None 值
            if x is None:
                return default
                
            # 處理 datetime.time
            if isinstance(x, dtC.time):
                # 時間轉換為秒數
                return x.hour * 3600 + x.minute * 60 + x.second + x.microsecond / 1e6
            
            # 處理 pandas.Timestamp (必須在 datetime.datetime 之前檢查，因為 Timestamp 是 datetime 的子類)
            elif isinstance(x, pd.Timestamp):
                # Pandas Timestamp 轉換為 UNIX timestamp
                return x.timestamp()
            
            # 處理 datetime.datetime
            elif isinstance(x, dtC.datetime):
                # datetime 轉換為 UNIX timestamp（以秒為單位）
                return (x - dtC.datetime(1970, 1, 1)).total_seconds()
            
            # 處理 datetime.date
            elif isinstance(x, dtC.date):
                # 日期轉換為 UNIX timestamp（以日期為單位）
                return (x - dtC.date(1970, 1, 1)).days
            
            # 處理 numpy datetime64
            elif isinstance(x, np.datetime64):
                # numpy datetime64 轉換為 UNIX timestamp
                return pd.Timestamp(x).timestamp()
            
            # 處理字符串
            elif isinstance(x, str):
                try:
                    # 擴充更多常見時間格式
                    formats = [
                        # 標準格式
                        '%Y-%m-%d %H:%M:%S',     # 2024-03-19 10:30:00
                        '%Y/%m/%d %H:%M:%S',     # 2024/03/19 10:30:00
                        '%Y-%m-%d',              # 2024-03-19
                        '%Y%m%d',                # 20240319
                        '%H:%M:%S',              # 10:30:00
                        '%Y-%m-%dT%H:%M:%S',     # 2024-03-19T10:30:00 (ISO format)
                        '%Y-%m-%d %H:%M:%S.%f',  # 2024-03-19 10:30:00.123456
                        
                        # 增加更多常見格式
                        '%Y%m%d%H%M%S',          # 20240319103000
                        '%Y%m%d%H%M',            # 202403191030
                        '%Y-%m-%d %H:%M',        # 2024-03-19 10:30
                        '%Y/%m/%d %H:%M',        # 2024/03/19 10:30
                        '%m/%d/%Y %H:%M:%S',     # 03/19/2024 10:30:00
                        '%d/%m/%Y %H:%M:%S',     # 19/03/2024 10:30:00
                        '%Y.%m.%d %H:%M:%S',     # 2024.03.19 10:30:00
                        
                        # 時間格式
                        '%H:%M',                 # 10:30
                        '%H%M%S',                # 103000
                        '%H%M',                  # 1030
                        
                        # 含時區的格式
                        '%Y-%m-%d %H:%M:%S%z',   # 2024-03-19 10:30:00+0800
                        '%Y-%m-%dT%H:%M:%S%z',   # 2024-03-19T10:30:00+0800
                        '%Y-%m-%d %H:%M:%S.%f%z' # 2024-03-19 10:30:00.123456+0800
                    ]
                    
                    for fmt in formats:
                        try:
                            ts = pd.to_datetime(x, format=fmt)
                            
                            # 特殊處理純時間格式
                            if fmt in ['%H:%M:%S', '%H:%M', '%H%M%S', '%H%M']:
                                time_parts = ts.time()
                                return time_parts.hour * 3600 + time_parts.minute * 60 + \
                                    (time_parts.second if hasattr(time_parts, 'second') else 0) + \
                                    (time_parts.microsecond / 1e6 if hasattr(time_parts, 'microsecond') else 0)
                            
                            # 其他格式返回 timestamp
                            return ts.timestamp()
                        except ValueError:
                            continue
                        except Exception as e:
                            LOGger.addlog(f'格式 {fmt} 解析失敗: {str(e)}', 
                                        stamps=['str_to_datetime', str(x)], 
                                        colora=LOGger.WARNING)
                            continue
                    
                    # 在所有格式都失敗後，嘗試一些特殊處理
                    try:
                        # 嘗試移除多餘的字符
                        cleaned_x = ''.join(c for c in x if c.isdigit() or c in ':-/.T ')
                        ts = pd.to_datetime(cleaned_x)
                        return ts.timestamp()
                    except:
                        # 如果還是失敗，記錄詳細資訊
                        LOGger.addlog(f'無法解析時間字串: {x}', 
                                    stamps=['str_to_datetime'], 
                                    colora=LOGger.WARNING)
                        LOGger.addlog(f'已嘗試的格式: {formats}', 
                                    stamps=['str_to_datetime'], 
                                    colora=LOGger.WARNING)
                        return default
                    
                except Exception as e2:
                    LOGger.addlog(f'時間字串轉換過程發生錯誤: {str(e2)}', 
                                stamps=['str_to_datetime', str(x)], 
                                colora=LOGger.WARNING)
                    return default
        except Exception as e:
            LOGger.addlog(f'時間轉換失敗，直接轉成數字: {str(e)}', stamps=['datetime_to_float', str(x)], colora=LOGger.WARNING)
            
            # 如果都不是上述類型，嘗試轉換為浮點數
            try:
                return float(x)
            except:
                LOGger.addlog(f'時間轉換失敗，直接轉成數字: {str(e)}', stamps=['datetime_to_float', str(x)], colora=LOGger.FAIL)
                return default
            
    except Exception as e:
        LOGger.addlog(f'時間轉換失敗: {str(e)}', stamps=['datetime_to_float', str(x)], colora=LOGger.FAIL)
        return default
def astype_evenif_datetime(*args, default=0, **kwags):
    def default_method(x, default=default):
        if(isinstance(x, dt) or isinstance(x, dtC.date) or isinstance(x, dtC.time)):
            return datetime_to_float(x, default=default)
        try:
            arg = float(x)
        except:
            arg = default
        return arg
    return astype(*args, default_method=default_method, **kwags)

#TODO:asnumpyarr
def asnumpyarr(arr, default=None, reshape_singledim=None, **kwags):
    addlog = kwags.get('addlog', LOGger.addloger(logfile=kwags.get('logfile','')))
    try:
        arr = np.array(arr)
        if(reshape_singledim!=None and len(arr.shape)==1):
            arr = arr.reshape(*reshape_singledim)
        return arr
    except Exception as e:
        LOGger.exception_process(e, logfile='') if(kwags.get('show_detail', m_show_detail)) else None
        addlog('arr:\n%s'%str(arr)[:1200])
        return default

#TODO:isnonnumber()
def isnonnumber(x, **kwags):
    b = True
    try:
        float(x)
        b = False
    except:
        pass
    return b
        
#TODO:convert
def convert(x, _type=float, ret=None, logging=False, logfile='', **kwags):
    try:
        ret = ret if(isinstance(ret, dict)) else {}
        ret.update({'converted':_type(x)})
        return True
    except:
        LOGger.addlog('failed!!!', stamps=[convert.__name__, str(x)], logfile=logfile) if(logging) else None
        return False

def clean_numeric_value(value, precision=5, likeIntPrecisionLimit=1e-7):
    """清理數值，處理浮點數精度問題"""
    if isinstance(value, (float, np.floating)):
        if pd.isna(value):
            return value
        # 如果非常接近整數（誤差小於 likeIntPrecisionLimit），轉換為整數
        if abs(value - round(value)) < likeIntPrecisionLimit:
            return round(value)
        # 否則四捨五入到指定精度
        return round(value, precision)
    return value

#TODO:isiterable
def isiterable(a, exceptions=[str, dict], type_stg_exceptions=[]):
    if(sum([isinstance(a, ecpn)+0 for ecpn in exceptions])>0):
        return False
    if(sum([(str(type(a)).find(ecpn)>-1)+0 for ecpn in type_stg_exceptions])>0):
        return False
    try:
        iter(a)
        return True
    except:
        return False
#TODO:extract
def extract(container, index=0, key='', default=None):
    try:
        if(isiterable(container)):
            shape = np.array(container).shape
            printer('[extract][index:%s][container shape:%s]'%(str(index)[:200], shape), 
                    showlevel=5)
            return np.array(container[np.array(index)]) if(isiterable(index)) else (
                    container[index] if(index<shape[0] or -index<=shape[0]) else default)
        elif(type(container)==dict):
            printer('[extract][key:%s][container length:%d]'%(key, len(container)), 
                    showlevel=5)
            return container[key] if(key in container) else default
    except:
        pass
    return default
#TODO:uniquifying
def uniquifying(ar, **unique_kwags):
    #按照原順序原型態，返回把重複的元素留下第一個向量
    org_type = type(ar)
    org_type = (lambda x:np.array(x)) if(org_type==np.ndarray) else org_type
    org_type = (lambda x:pd.DataFrame(x)) if(str(org_type).find('pandas.core.frame')>-1) else org_type
    org_type = (lambda x:pd.Series(x)) if(str(org_type).find('pandas.core.series')>-1) else org_type
    ar_list = list(tuple(ar))
    ar_list = sorted(np.unique(ar, **unique_kwags), key=lambda x:ar_list.index(x))
    return org_type(ar_list)

#TODO:isunique
def isunique(ar, uniquify_method=uniquifying, **kwags):
    ret =kwags.get('ret',{})
    if(np.array(LOGger.flattern_list(ar)).shape[0]>np.unique(np.array(LOGger.flattern_list(ar))).shape[0]):
        ret['new_ar'] = uniquify_method(ar)
        return False
    return True

#TODO:get_all_values
def get_all_values(*args, only_numbers=1, uniquifying_method=m_uniq_thru_set, **kwags):
    criterion=(lambda arg:(str(type(arg)).find('float')>-1)|(str(type(arg)).find('int')>-1)) if(
            only_numbers) else (lambda **kwags:True)
    ret = dearray_process(*args, method=None, criterion=criterion, **kwags)
    ret = (list(set(ret)) if(kwags['uniquify']) else ret) if('uniquify' in kwags) else ret
    return ret

#TODO:dearray_process
#method can only pass non-iterable object
def dearray_process(*args, method=None, criterion=None, **kwags):
    criterion = criterion if(criterion!=None) else (lambda arg, **kwags:True) 
    ret = []
    for arg in args:
        if(isiterable(arg)):
            ret += dearray_process(*arg, method=method, **kwags)
        else:
            ret += [method(arg, **kwags) if(method!=None) else arg] if(criterion(arg)) else []
    return ret

def application_byterm(data, method=None, axis=None, **kwags):
    data_T = kwags.get('data_T', data.T)
    if(isinstance(data, pd.core.frame.DataFrame)):
        if(axis==1):
            return data.applymap(method).any()
        elif(axis==0):
            return data_T.applymap(method).any()
        elif(axis==None):
            return data_T.applymap(method)
        else:
            return None
    if(isinstance(data, pd.core.series.Series)):
        return data.apply(method)
    
def reduction(tensor, mask, axis=0, loc_index=True, should_iterable_object_transpose=True):
    if(isinstance(tensor, pd.core.frame.DataFrame)):
        return (tensor[mask] if(axis==1) else tensor.T[mask].T) if(not loc_index) else (
            tensor.iloc[mask] if(axis==0) else tensor.T.iloc[mask].T)
    elif(isinstance(tensor, pd.core.series.Series)):
        return tensor.iloc[mask]
    elif(isinstance(tensor, np.ndarray)):
        return np.take(tensor, mask, axis)
    elif(isiterable(tensor)):
        org_type = type(tensor)
        np_tensor = np.transpose(np.array(tensor)) if(should_iterable_object_transpose) else np.array(tensor)
        np_tensor_new = reduction(np_tensor, mask, axis)
        tensor_new = np_tensor_new.tolist()
        return org_type(tensor_new)
    else:
        return None
        
#TODO:find_spec_values
def find_spec_values(data, values=[], conditions={}, max_show_length=20, **kwags):
    addlog = kwags.get('addlog', (lambda s, **kwags:LOGger.addlog(s, logfile='')))
    data_T = kwags.get('data_T', data.T)
    for value in values:
        addlog('%s start....'%value)
        columns_value_in = data.columns[application_byterm(data, (lambda s:s==value), 1, data_T=data_T)]
        index_mask = application_byterm(data, (lambda s:s==value), 0, data_T=data_T)
        index_value_in = data_T.columns[index_mask]
        addlog('columns value in:%s'%','.join(list(map(str, columns_value_in[:max_show_length]))))
        addlog('index value in:%s'%','.join(list(map(str, index_value_in[:max_show_length]))))
        data_value_in = data[columns_value_in][index_mask]
        addlog('data focus on value:%s'%str(data_value_in))
    for condition_name, condition in conditions.items():
        addlog('%s start....'%condition_name)
        columns_value_in = data.columns[application_byterm(data, (lambda s:condition(s,**kwags)), 1, data_T=data_T)]
        index_mask = application_byterm(data, (lambda s:condition(s,**kwags)), 0, data_T=data_T)
        index_value_in = data_T.columns[index_mask]
        addlog('columns value in:%s'%','.join(list(map(str, columns_value_in[:max_show_length]))))
        addlog('index value in:%s'%','.join(list(map(str, index_value_in[:max_show_length]))))
        data_value_in = data[columns_value_in][index_mask]
        addlog('data focus on value:%s'%str(data_value_in))

#TODO:locate_spec_data
def locate_spec_data(root, spec_locs=None, spec_values=None, extend=(0, 1), value_extend=(None, None), 
                     extends=None, value_extends = None, metric=(lambda v,s:np.abs(v-s)), fill_interval=False, 
                     root_ascending=True, value_tolerance=1, **kwags):
    spec_locs = [] if(spec_locs==None) else spec_locs
    spec_values = [] if(spec_locs==None) else spec_values
    extends = mylist([] if(isinstance(extends, type(None))) else extends)
    value_extends = mylist([] if(isinstance(value_extends, type(None))) else value_extends)
    len_root = np.array(root).shape[0]
    
    locate_mask = np.array([False]*len_root)
    if(fill_interval):
        for value in spec_values:
            loc = int(np.argmin([metric(v, value) for v in root])//1)
            spec_locs.append(loc)
        min_loc = int(max(min(int(np.min(spec_locs)//1)+extend[0], len_root), 0)//1)
        max_loc = int(min(max(int(np.max(spec_locs)//1)+extend[1], min_loc), len_root)//1)
        locate_mask[min_loc:max_loc] = True
    else:
        for loc in spec_locs:
            min_loc = int(max(min(loc+extend[0], len_root), 0)//1)
            max_loc = int(min(max(loc+extend[1], min_loc), len_root)//1)
            locate_mask[min_loc:max_loc] = True
        for i,value in enumerate(spec_values):
            loc = int(np.argmin([metric(v, value) for v in root])//1)
            spec_locs.append(loc)
            extend_ = extends.get(i,extend)
            value_extend_ = value_extends.get(i,value_extend)
            if(not isinstance(value_tolerance, type(None))):
                if(not np.array(list(map(lambda v:metric(v, value+value_extend_[1])<=value_tolerance, root))).any() and 
                   not np.array(list(map(lambda v:metric(v, value+value_extend_[0])<=value_tolerance, root))).any()):
                    continue
            min_loc = int(max(min((loc+extend_[0]) if(value_extend_[0]==None) else np.argmin(
                        [metric(v, value+value_extend_[0]) for v in root] if(
                            root_ascending) else [metric(v, value+value_extend_[1]) for v in root]), len_root-1), 0)//1)
            max_loc = int(min(max((loc+extend_[1]) if(value_extend_[1]==None) else np.argmin(
                        [metric(v, value+value_extend_[1]) for v in root] if(
                            root_ascending) else [metric(v, value+value_extend_[0]) for v in root]), min_loc+1), len_root)//1)
            locate_mask[min_loc:max_loc] = True
    locs = sorted(spec_locs)
    return locs, locate_mask

#TODO:array_parallel_filter
def array_parallel_filter(*args, spec_locs=None, spec_values=None, root_index=0, extend=(0, 1), 
                          value_extend=(None, None), root=None, metric=(lambda v,s:np.abs(v-s)), **kwags):
    if(root==None if(not isinstance(root, np.ndarray)) else False):
        root = args[root_index]
    spec_locs = [] if(spec_locs==None) else spec_locs
    spec_values = [] if(spec_values==None) else spec_values
    locate_mask = locate_spec_data(root, spec_locs=spec_locs, spec_values=spec_values, extend=extend, 
                     value_extend=value_extend, metric=metric, **kwags)[1]
    new_args, unsel_args = [], []
    for arg in args:
        arg_type = type(arg)
        new_arg = np.array(tuple(arg))[locate_mask]
        # print('%s'%str(new_arg)[:200])
        new_args.append(arg_type(tuple(new_arg)))
        unsel_arg = np.array(tuple(arg))[np.logical_not(locate_mask)]
        unsel_args.append(arg_type(tuple(unsel_arg)))
    return tuple(new_args), tuple(unsel_args)

#TODO:sort_values
def sort_values(data, order, ordered_col='index', ordering_column='index', 
                nanwarning=True, index_revert=True, **kwags):
    str_type = str(type(data))
    if(str_type.find('pandas.core.frame')>-1):
        index = data.index
        data = data.set_index(ordered_col) if(ordered_col!='index') else data
        new_data = data.reindex(order)
#        if(nanwarning and np.isnan(np.array(new_data).reshape(-1)).any()):
#            printer('nan!!!')
#            printer(str(new_data)[:200])
        if(index_revert and ordered_col!='index'):
            new_data = new_data.reset_index()
            new_data = pd.DataFrame(np.mat(new_data), index=index, columns=new_data.columns)
    elif(str_type.find('ndarray')>-1):
        new_data = type_transforming_in_method(sort_values, type_transform=(np2pd, pd2np),
                                    data = data, method_inputs_consts={
                                            'order':order, 'ordered_col':ordered_col, 
                                            'ordering_column':ordering_column, 
                                            'nanwarning':nanwarning, 
                                            'index_revert':index_revert,
                                            **kwags})
    return new_data
#TODO:inherit
def inherit(object_):
    dtype = type(object_)
    if(dtype==np.ndarray):
        return np.array
    else:
        return dtype

class Container():
    def __init__(self, root=[]):
        self.ctype = type(root)
        self.container = root
        
    def get_inherit(self):
        return inherit(self.ctype)
        
    def add(self, object_, **add_kwags):
        success = False
        container_index = list(tuple(getattr(self.container, 'index', [])))
        k = add_kwags.get('k', 'k_%d'%len(getattr(
                        self.container, 'values', list(tuple(self.container)))))
        while(k in self.container):
            k += '_'
        if(hasattr(self.container, 'append')):
            self.container.append(object_, **add_kwags)
            success = True
        elif(self.ctype in [set, tuple]):
            inheritor = self.get_inherit(self.ctype)
            container_list = list(self.container)
            container_list.append(object_, **add_kwags)
            self.container = inheritor(container_list)
            success = True
        elif(self.ctype in [np.ndarray]):
            inheritor = self.get_inherit(self.ctype)
            container_list = list(tuple(self.container))
            container_list.append(object_, **add_kwags)
            self.container = inheritor(container_list)
            success = True
        elif(self.ctype in [pd.core.series.Series]):
            add_kwags['sort'] = add_kwags.get('sort', False)
            inheritor = self.get_inherit(self.ctype)
            container_list = list(tuple(self.container))
            container_list.append(object_, **add_kwags)
            self.container = inheritor(container_list, index=container_index+[k])
            success = True
        elif(self.ctype==dict):
            dict_object_ = {k:object_}
            self.container.update(**dict_object_)
            success = True
        elif(self.ctype in [pd.core.frame.Frame]):
            add_kwags['sort'] = add_kwags.get('sort', False)
            inheritor = self.get_inherit(self.ctype)
            axis = add_kwags.get('axis', 1)
            list_object_ = [object_]*(self.container.shape[1-axis])
            if(axis==1):
                self.container[k] = list_object_
            elif(axis==0):
                pd_object_ = pd.Series([list_object_], index=[k])
                self.container = self.container.append(pd_object_, **add_kwags).copy()
            success = True
        return success
        
#TODO:dict_transpose
def dict_transpose(dic):
    k2_keys = []
    for k1,v in dic.items():
        k2_keys += list(tuple(v.keys()))
    k2_keys = list(tuple(np.unique(k2_keys)))
    
    ret = {k2:{} for k2 in k2_keys}
    for k2 in k2_keys:
        for k1,v in dic.items():
            if(k2 in v):
                ret[k2][k1] = dcp(v[k2])
    return ret

#TODO:mydict
class mydict(dict):
    def __init__(self, dict_root={}, **stg_dict_root):
        dict_root = dict_root if(isinstance(dict_root, dict)) else {}
        self.update(dict_root)
        self.update(stg_dict_root)
    def concatenate(self, dtype=None, rewrite=False, return_self=False, **kwags):
        printer = kwags.get('printer', print)
        item_inlist = list(self.values())
        if(dtype==None):
            dtype = np.array
            uniquifying_dtypes_instg = uniquifying(list(map(lambda x:str(type(x)), item_inlist)))
            dtypes = list(map(inherit, item_inlist))
            if(len(uniquifying_dtypes_instg)==1):
                dtype = dtypes[0]
                printer('using common dtype:%s!!!'%str(dtype))
        kwags['axis'] = kwags.get('axis', 1)
        if(dtype==dict):
            contents_key_inlist = get_all_values([list(tuple(d.keys())) for item_key,d in self.items()])
            if(np.unique(contents_key_inlist).shape[0]<len(contents_key_inlist)):
                printer('contents keys duplicated!!!!')
                if(not rewrite):
                    return {}
            ret = {}
            for item_key,d in self.items():
                ret.update(d)
            return ret
        elif(dtype==tuple):
            item_inlist_flatten = get_all_values(item_inlist)
            return tuple(uniquifying(item_inlist_flatten) if(rewrite) else item_inlist_flatten)
        elif(str(dtype).find('list')>-1):
            item_inlist_flatten = get_all_values(item_inlist)
            return mylist(tuple(uniquifying(item_inlist_flatten) if(rewrite) else item_inlist_flatten))
        elif(dtype==str):
            stg_sep = kwags.get('stg_sep', '||')
            item_inlist_flatten = get_all_values(item_inlist)
            item_inlist_flatten = list(tuple(uniquifying(item_inlist_flatten) if(
                                    rewrite) else item_inlist_flatten))
            return stg_sep.join(list(map(str, tuple(item_inlist_flatten))))
        elif(dtype==np.array):
            return np.concatenate(item_inlist, **kwags)
        elif(dtype in [pd.DataFrame, pd.Series]):
            return pd.concat(item_inlist, **kwags)
        elif(str(type(dtype)).find('dataframeprocedure.collection')>-1):
            ret = collection()
            for v in item_inlist:
                ret.update(v, **kwags)
            return ret
        if(return_self):
            return self
        return None
    def get_items_by_key(self, item_key, return_grp=False):
        ret = {}
        for k in self:
            if(isinstance(self[k], dict)):
                if(item_key in self[k]):
                    ret[k] = self[k][item_key]
            if(k==item_key):
                ret[k] = self[k]
        if(len(ret)==1 and not return_grp):
            return ret[tuple(ret.keys())[0]]
        if(len(ret)==0 and not return_grp):
            return None
        return ret
    
    def get_items_by_index(self, item_index=0, return_grp=False):
        ret = {}
        for k in self:
            if(isinstance(self[k], list)):
                if(item_index < len(self[k])):
                    ret[k] = self[k][item_index]
        if(len(ret)==1 and not return_grp):
            return ret[tuple(ret.keys())[0]]
        if(len(ret)==0 and not return_grp):
            return None
        return ret

#TODO:collection
class collection():
    def __init__(self, item_key=None, **kwags):
        self.collection_keys = []
        for k, v in kwags.items():
            setattr(self, '%s_org'%k, v)
            setattr(self, k, v)
            set_item_key = item_key if(item_key!=None) else 0
            setattr(self, '%s_grp'%k, mydict({set_item_key:v}))
            self.collection_keys.append(k)
            self.collection_keys
    #TODO:collection.add
    def add(self, *args, item_key=None, **kwags):
        for i in range(len(args)):
            k, v = '_%d'%i, args[i]
            self.collection_keys.append(k)
            if(hasattr(self, '%s_org'%k)):
                setattr(self, k, v)
                grp = getattr(self, '%s_grp'%k)
                len_grp = len(grp)
                set_item_key = item_key if(item_key!=None) else len_grp
                while(set_item_key in grp):
                    set_item_key = str(set_item_key) + '_'
                grp[set_item_key] = v
            else:
                setattr(self, '%s_org'%k, v)
                setattr(self, k, v)
                set_item_key = item_key if(item_key!=None) else 0
                setattr(self, '%s_grp'%k, mydict({set_item_key:v}))
        for k, v in kwags.items():
            self.collection_keys.append(k)
            if(hasattr(self, '%s_org'%k)):
                setattr(self, k, v)
                grp = getattr(self, '%s_grp'%k)
                len_grp = len(grp)
                set_item_key = item_key if(item_key!=None) else len_grp
                while(set_item_key in grp):
                    set_item_key = str(set_item_key) + '_'
                grp[set_item_key] = v
            else:
                setattr(self, '%s_org'%k, v)
                setattr(self, k, v)
                set_item_key = item_key if(item_key!=None) else 0
                setattr(self, '%s_grp'%k, mydict({set_item_key:v}))
    #TODO:collection.add
    def integrate(self, *args, flatten_values=False, uni_key=None, item_index=mylist()):
        args = [args] if(type_string(args).find('str')>-1) else list(args)
        args = args if(not flatten_values) else get_all_values(args)
        type_args = mylist(map(lambda v:type_string(v), args))
        if(np.unique(type_args).shape[0]!=1):
            return False
        uni_type = type_args[0] if(uni_key==None) else uni_key
        item_index = mylist(item_index)
        for i, arg in enumerate(args):
            self.add(**{uni_type:arg, 'item_key':item_index.get(i, i, inherit=True)})
        return True
            
    #TODO:collection.pop
    def pop(self, k, item_key=None, item_index=-1, entire=False, grp_type=dict, dtype=None, return_grp=False):
        k = '_%d'%k if(type(k)==int) else k
        grp = getattr(self, '%s_grp'%k, mydict())
        history_keys = dcp(list(tuple(grp).keys()))
        if(not entire):
            item_key = item_key if(item_key!=None) else (history_keys[item_index] if(history_keys) else None)
            if(item_key in history_keys):
                popped_value = grp.pop(item_key, None)
                if(len(grp)>0):
                    setattr(self, k, grp[history_keys[-1]])
                else:
                    delattr(self, k)
                    delattr(self, '%s_org'%k)
                    delattr(self, '%s_grp'%k)
                return popped_value
        else:
            popped_value = dcp(self.get_all(k, grp_type=grp_type, dtype=dtype, return_grp=return_grp))
            for item_key in history_keys:
                self.pop(k, item_key=item_key, item_index=-1, entire=False)
            return popped_value
        return None
    #TODO:collection.get
    def get(self, k, item_key=None, item_index=0):
        k = '_%d'%k if(type(k)==int) else k
        history_keys = list(tuple(getattr(self, '%s_grp'%k, mydict()).keys()))
        item_key = item_key if(item_key!=None) else (history_keys[item_index] if(history_keys) else None)
        if(item_key!=None):
            item = getattr(self, '%s_grp'%k)[item_key]
        else:
            item = None
        return item
    #TODO:collection.get_keys
    def get_keys(self, unique=True, uniquifying_method=m_uniq_thru_set):
            return uniquifying_method(self.collection_keys) if(unique) else self.collection_keys
    #TODO:collection.is_empty
    def is_empty(self):
        if(len(self.get_keys())==0):
            return True
        negative_ret = False
        for k in self.get_keys():
            negative_ret |= (getattr(self, '%s_grp'%k, mydict())!=mydict())
        return not negative_ret
    #TODO:collection.get_all
    def get_all(self, k=None, ks=[], axis=0, grp_type=dict, dtype=None, return_grp=False, 
                item_key=None, item_index=0):
        if(k==None):
            if(len(self.get_keys())==1): #如果其實只有一種類型的物件，那就是要那個類型物件的get_all
                k = self.get_keys()[0]
                return self.get_all(k, ks=ks, axis=axis, grp_type=grp_type, dtype=dtype,
                                    return_grp=return_grp, item_key=item_key, item_index=item_index)
            #每一種類型都拿，裝在grp_type裡
            ret = {}
            ks = ks if(np.array(ks).shape!=()) else self.collection_keys
            for k in ks:
                ret['_%d'%k if(type(k)==int) else k] = self.get_all(k, grp_type=dict, return_grp=True)
            ret = dict_transpose(ret) if(axis==0) else ret
            if(len(ret)==0 and not return_grp):
                return None
            elif(len(ret)==1 and not return_grp):
                return tuple(ret.values())[0]
        else:
            k = '_%d'%k if(type(k)==int) else k
            grp = getattr(self, '%s_grp'%k, mydict())
            inherits_inlist = list(map(inherit, tuple(grp.values())))
            inheritor = dtype #只是換個名字
            if(inheritor==None):
                inheritor = lambda x:x
                uniquifying_dtypes_instg = uniquifying(list(map(lambda x:str(type(x)), inherits_inlist)))
                if(len(uniquifying_dtypes_instg)==1):
                    inheritor = inherits_inlist[0]
                    printer('using common dtype:%s!!!'%str(inheritor))
                else:
                    inheritor = lambda x:x
            if(len(grp)==0 and not return_grp):
                return None
            elif(len(grp)==1 and not return_grp):
                return inheritor(self.get(k, item_key=None, item_index=0))
            ret = mydict({k:inheritor(v) for (k,v) in grp.items()})
#        if(grp_type==None):
#            grp_type = np.array
#            uniquifying_dtypes_instg = uniquifying(inherits_inlist)
#            if(len(uniquifying_dtypes_instg)==1):
#                grp_type = inherits_inlist[0]
#                printer('using common grp_type:%s!!!'%str(grp_type))
        if(grp_type!=dict):
            values = list(tuple(ret.values()))
            ret = grp_type(values)
        return ret
    #TODO:collection.update
    def update(self, clt, rewrite=False):
        clt_collection_keys = clt.get_keys()
        for k in clt_collection_keys:
            collection_k = getattr(clt, '%s_grp'%k, mydict())
            for item_key, v in collection_k.items():
                if(not rewrite):
                    while(item_key in  getattr(self, '%s_grp'%k, mydict())):
                        item_key = str(item_key) + '_'
                item = {k:v}
                self.add(**item, item_key=item_key)
    #TODO:collection.concatenate
    def concatenate(self, k=None, dtype=None, rewrite=False, **kwags):
        if(k==None):
            if(len(self.get_keys())==1): #如果其實只有一種類型的物件，那就是要那個類型物件的concatenate
                k = self.get_keys()[0]
                return self.concatenate(k, dtype=dtype, 
                                        rewrite=rewrite, **kwags)
        else:
            k = '_%d'%k if(type(k)==int) else k
            grp = getattr(self, '%s_grp'%k, mydict())
            if(len(grp)==0):
                dtype = dtype if(dtype!=None) else np.array
                return dtype([])
            if(len(grp)==1):
                item = getattr(self, k) if(dtype==None) else  dtype(getattr(self, k))
                return item
            return grp.concatenate(**kwags)
        
def data_transform(*args, output_type='pkl', **kwags):
    for file in args:
        file_type = os.path.basename(file).split('.')[-1]
        df = import_data(file, file_type=file_type, **kwags)
        if(output_type=='pkl'):
            df.to_pickle(file.replace('.'+file_type, '.pkl'))
        elif(output_type=='csv'):
            df.to_csv(file.replace('.'+file_type, '.csv'))
        elif(output_type=='xlsx'):
            wrt = pd.ExcelWriter(file.replace('.'+file_type, '.xlsx'), engine='xlsxwriter')
            df.to_excel(wrt, sheet_name=kwags.get('sheet_name', kwags.get('sheet', kwags.get('sht', 'main'))))

#TODO:set_header
def set_header(df, hds=mylist(), cell_size=0, cell_subsize=(0,-1), cell_title=None, default_value=0,
                index_sep='#', ret = collection()):
    type_mask = mylist(map(lambda hd:type_string(hd), hds))
    if(not type_mask.find('list') in [-1]):
        if(len(uniquifying(type_mask))>1): #代表接口層次不齊
            return False
        for i,hd_is_list in enumerate(hds):
            if(not mylist(map(lambda hd:type_string(hd), hd_is_list)).find('list') in [-1]): #代表接口太多層
                ret = collection()
                return False
            set_header(df, hds=mylist(hd_is_list), cell_size=cell_size, 
                       cell_subsize=cell_subsize, cell_title=cell_title, 
                       default_value=default_value, index_sep=index_sep, ret=ret)
    hds_stg = list(tuple(np.array(hds)[np.array(list(map(lambda hd:(type_string(hd).find('str')>-1), hds)))]))
    df_hds = df[hds_stg]
    if(cell_size):
        cell_title = 'cell_%d'%len(getattr(ret, 'df_grp', {})) if(cell_title==None) else cell_title
        DF = pd.DataFrame(columns=[cell_title])
        for i in range(0,len(df_hds.index), cell_size):
            Index = str(df_hds.index[i])[:str(df_hds.index[i]).find(index_sep)]
            df_i = df.iloc[i:i+cell_size].copy()
            df_i = df_i[hds_stg]
            df_i = df_i.iloc[cell_subsize[0]:cell_subsize[1]:(1 if(len(cell_subsize)<3) else cell_subsize[2])]
            padding_size = len([v for v in range(*cell_subsize)])
            padding(df_i, 0, padding_size, default_value=default_value)
            DF = append(DF, pd.DataFrame([[df_i]], index=[Index], columns=[cell_title]), sort=False)
    else:
        DF = df[hds_stg]
    ret.add(df = DF) if(hds_stg!=[]) else None
    return True
#TODO:indexing_header_from_large_one
def indexing_header_from_large_one(*args, base=[], need_uniquifying=True, strictly=False, **kwags):
    b = LOGger.flattern_list(mylist(tuple(base)))
    ret = {}
    if(not isunique(b, ret=ret) and need_uniquifying):
        b = ret['new_array']
    
    masks=[]
    for i, arg in enumerate(args):
        arg = uniquifying(arg) if(need_uniquifying) else arg
        index_mask = np.array([b.index(a) for a in arg]) if(
                    strictly) else np.array([b.index(a) for a in arg if a in b])
        masks.append(index_mask)
    ret = kwags.get('ret', {})
    for k,V in kwags.items():
        if(not isiterable(V)):
            continue
        ret[k] = index_mask = np.array([b.index(a) for a in V]) if(
                    strictly) else np.array([b.index(a) for a in V if a in b])
    return tuple(masks)

#TODO:file_checking
def file_checking(file, data_type=None, file_None_return=False, file_type='pkl', **kwags):
    try:
        if(not isinstance(file, str)):
            return file_None_return
        if(not os.path.exists(file)):
            print('檔案不存在:%s'%file)
            return False
        if(data_type!=None):
            data = import_data(file, file_type=file_type)
            if(not isinstance(data, data_type)):
                print('資料類型[%s]錯誤:%s'%(str(data_type), LOGger.type_string(data)))
                return False
    except Exception as e:
        LOGger.exception_process(e, logfile='', stamps=[__file__, 'file_checking'])
        return False
    return True
#TODO:columns_valid
def columns_valid(columns='all', data=pd.DataFrame(), **kwags):
    try:
        if(columns=='all'):
            return True
        if(columns==None):
            return True
        data[columns]
    except:
        print('欄位[%s]不存在'%str(columns))
        return False
    return True
#TODO:transpose_data
def transpose_data(data, thread_cell_size=None, **kwags):
    thds = []
    ret = kwags['ret'] if('ret' in kwags) else {}
    ta = dt.now()
    if(thread_cell_size):
        for i in range(0, len(data.index), thread_cell_size):
            df_cell = data.iloc[i:i+thread_cell_size]
            thd = threading.Thread(target=transpose_data, args=[df_cell], 
                                   kwargs={'thread_key':i, 'ret':ret, 'thread_cell_size':0})
            thd.start()
            thds.append(thd)
        for thd in thds:
            thd.join()
        new_data = pd.DataFrame()
        for i, t in sorted(ret.items()):
            new_data = t if(new_data.empty) else new_data.append(t, sort=False)
        printer('[separable]transposing end:%.2f(s)'%((dt.now() - ta).total_seconds()))
    elif(thread_cell_size==0):
        i = kwags['thread_key']
        key = '[%s]'%kwags['thread_key'] if('thread_key' in kwags) else ''
        print('%stransposing....'%key)
        ta = dt.now()
        ret[i] = data.T.copy()
        print('%stransposing end:%.2f(s)'%(key, (dt.now() - ta).total_seconds()))
        return
    else:
        new_data = data.T.copy()
    return new_data

#TODO:separate_nanull_columns
def separate_nanull_columns(data, return_coln=1, **kwags):
    str_type = str(type(data))
    if(str_type.find('pandas.core')>-1):
        if(not data.applymap(isnonnumber).any().any()):
            columns, other_cols = (data.columns, pd.Index([], dtype=int))
        else:
            print('偵測到空值，進行排除:\n%s...'%str(','.join(list(tuple(data.columns[
                    data.applymap(isnonnumber).any()])))[:20]))
            columns = data.columns[np.logical_not(
                    data.applymap(isnonnumber).any())]
            other_cols = pd.Index(list(set(list(data.columns))-set(columns)))
            if(other_cols.shape[0]==0): 
                other_cols = np.Index([], dtype=int)
        print('other_cols', other_cols.shape)
        data_0, data_1 = data[columns], data[other_cols]
    else:
        pd_data = np2pd(data)
        columns, other_cols = separate_nanull_columns(pd_data, 1)
        pd_data_0, pd_data_1 = pd_data[columns], pd_data[other_cols]
        data_0, data_1 = pd2np(pd_data_0), pd2np(pd_data_1)
    return (columns, other_cols) if(return_coln) else (data_0, data_1)

#TODO:compute_chaos
#如果不要某些欄，可以先drop掉
def compute_chaos(data, drop_columns=[], **kwags):
    str_type = str(type(data))
    if(str_type.find('pandas.core')>-1):
        data = separate_nanull_columns(data, return_coln=0)[0]
        data_sq = ((data - np.average(data, axis=0))**2).drop(drop_columns, axis=1)
        return np.std((np.sum(data_sq, axis=1))**(1/2))
    else:
        kwags.update({'drop_columns':drop_columns})
        pd_data = np2pd(data, **kwags)
        return compute_chaos(pd_data, **kwags)

#TODO:get_column_names_from_txt
def get_column_names_from_txt(file, encoding='utf-8'):
    if(not os.path.isfile(file)):
        return []
    f = open(file, encoding=encoding)
    line = f.read()
    f.close()
    line = line[line.find('['):line.find(']')+1]
    labels = line.split(',')
    labels = [v.replace("'", "").replace(
            ' ', '').replace('[','').replace(']','') for v in labels]
    return labels    

def feature_selecting(data, select_method=SelectKBest, algorithm=chi2, 
                        yheader=None, xheader='full', normalize=False, 
                        data_x=pd.DataFrame(), data_y=pd.DataFrame(),
                        **kwags):
    if(data_x.shape[0]!=data_y.shape[0]):
        data_x, data_y = pd.DataFrame(), pd.DataFrame()
    xheader = list(np.array(xheader).reshape(-1))
    yheader = list(np.array(yheader).reshape(-1))
    data_x = (data[xheader] if(xheader!='full') else data[[
            c for c in data.columns if not c in yheader]]) if(
            data_x.empty) else data_x
    data_y = (data[yheader] if(yheader!=None) else data[data.columns[-1]]) if(
            data_y.empty) else data_y
    str_type_x, str_type_y = str(type(data_x)), str(type(data_y))
    if(str_type_x!=str_type_y):
        return 
    if(str_type_x.find('pandas')>-1):
        dfx, xscr = normalization(data_x) if(normalize) else (data_x.copy(), None)
        dfy, yscr = normalization(data_y) if(normalize) else (data_y.copy(), None)
        skb = select_method()#SelectKBest(chi2, k='all')
        np_new_data_x = skb.fit_transform(dfx, dfy)
        columns_selected = data_x.columns[skb.get_support()]
        df_scores = pd.DataFrame(skb.scores_, 
                             index = dfx.columns).round(2)
        new_data_x = pd.DataFrame(
                np_new_data_x, columns = columns_selected, index = data_x.index)
        new_data = new_data_x.join(data_y)
    if(str_type_x.find('ndarray')>-1):
        method = lambda **kwags:feature_selecting(pd.DataFrame(),
                                select_method=select_method, algorithm=algorithm, 
                                normalize=normalize, **kwags)
        new_data, df_scores = type_transforming_in_method(
                method, type_transform=(np2pd, pd2np), data_x=data_x, data_y=data_y)
    return new_data, df_scores

def selectSubHeaderIndex(sourceHeader=np.arange(3), subHeader=np.array([2,1,None,-1]), notFoundCode=None, **kwags):
    return np.array([LOGger.mylist(sourceHeader).index_default(s,notFoundCode) for s in subHeader])

def selectSubHeaderFromData(data=np.arange(15).reshape(5,3), subHeader=np.array([2,1,None,-1]), sourceHeader=None, defaultValue=None, 
                            notFoundCode=None, isReduceDim=True, **kwags):
    '''
    pd.core frame.DataFrame --> np.ndarray --> main procedure...

    data: np.ndarray or pd.core.frame.DataFrame
    subHeader: list or np.ndarray
    sourceHeader: list or np.ndarray
    defaultValue: Any
    '''
    if(isinstance(data, pd.core.frame.DataFrame)):
        dataTemp = data.values
        if(not isiterable(sourceHeader)):   sourceHeader = list(data.columns)
        subHeaderTemp = selectSubHeaderIndex(sourceHeader, subHeader, notFoundCode=notFoundCode, **kwags) if(isinstance(sourceHeader, list)) else subHeader
        dataArr = selectSubHeaderFromData(dataTemp, subHeaderTemp, sourceHeader=None, defaultValue=defaultValue, **kwags)
        dataTemp = pd.DataFrame(dataArr, index=data.index, columns=subHeader)
        return dataTemp
    sourceHeader = list(tuple(np.arange(data.shape[1]) if(sourceHeader is None) else sourceHeader))
    if(sourceHeader):   subHeader = [(x%len(sourceHeader) if x<0 else x) if(isinstance(x, int)) else x for x in subHeader]
    subHeaderIndex = selectSubHeaderIndex(sourceHeader, subHeader, notFoundCode=notFoundCode, **kwags)
    dataTemp = None
    for i,idx in enumerate(subHeaderIndex):
        if(dataTemp is None):
            if(idx is notFoundCode):
                dataTemp = np.full((data.shape[0], 1), defaultValue)
            else:
                dataTemp = data[:,np.array([idx])]
        else:
            if(idx is notFoundCode):
                dataTemp = np.hstack((dataTemp, np.full((data.shape[0], 1), defaultValue)))
            else:
                dataTemp = np.hstack((dataTemp, data[:,np.array([idx])]))
    if(isReduceDim):    
        if(len(dataTemp.shape)>1):
            if(dataTemp.shape[1]==1):   dataTemp = dataTemp.reshape(-1)
    return dataTemp



def asnumeric(data, axis=1, **kwags):
    str_type = str(type(data))
    if(str_type.find('ndarray')>-1):
        pd_data = pd.DataFrame(data) if(len(data.shape)>1) else pd.Series(data)
        new_data = asnumeric(pd_data, axis=axis, **kwags)
    if(str_type.find('pandas.core')>-1):
        if(len(data.shape)==1):
            new_data = data.astype(float)
        else:
            axis_array = data.columns if(axis==1) else data.index
            new_data = data
            for c in axis_array:
                try:
                    new_data[c] = new_data[c].astype(float)
                except:
                    continue
    return new_data

#method_kwags = {'a':1}
#def method(df_haaa, **kwags):
#    printer(kwags['kwags'])
#    NGtypes = ['crack_NG','hairy_NG','bubbly_NG']
#    for NG in NGtypes:
#        NGstates = np.array(df_haaa[NG].astype(float))
#        NGstate = np.average(NGstates)
#        df_haaa[NG] = NGstate
#    return df_haaa
#refield_sequential_data(df0, 20, 4, method=method, method_kwags=method_kwags)
#refield_sequential_data(df0, 20, 4, method=method)
def refield_sequential_data(df0, new_field_sz, old_field_sz=None, df_new=None, eventname_header='event_index', stamps=None,
                            initial_header=None, initial_headers=None, initial_method='value min', initial_methods=None, **kwags):
    df_new = pd.DataFrame() if(isinstance(df_new, type(None))) else df_new
    stamps = stamps if(isinstance(stamps, list)) else []
    initial_headers = initial_headers if(isinstance(initial_headers, list)) else []
    initial_methods = initial_methods if(isinstance(initial_methods, dict)) else {}
    old_field_sz = float(old_field_sz) if(isinstance(astype(old_field_sz, float, default=None), float)) else df0.shape[0]
    old_data_sz_undup = len(set(df0.index))
    printer('old data size %d with field size:%d'%(old_data_sz_undup, old_field_sz))
    if(new_field_sz<=1 or ((new_field_sz//1)!=new_field_sz)):
        {}['new size error:%.2f'%(new_field_sz)]
    if(((old_data_sz_undup)/old_field_sz)//1 != (old_data_sz_undup/old_field_sz)):
        {}['old size ratio error:%.2f'%(old_data_sz_undup/old_field_sz)]
    df0_T = df0.T.copy()
    new_grp_no=0
    for i in range(0, len(df0.index), old_field_sz):
        indexes_now = df0.index[i:i+old_field_sz].copy()
        df_haa = df0_T[indexes_now].T.copy()
        #|||||...|||||    ← df_haa
        #||...||          ← df_haa.index[0:n_fd_sz]
        # ||...||         ← df_haa.index[1:n_fd_sz+1]
        #  ||...||        ← df_haa.index[2:n_fd_sz+2]
        #      ...                ...
        #      ||...||    ← df_haa.index[-n_fd_sz-1:n_fd_sz-n_fd_sz-1]
        for j,jndex in enumerate(df_haa.index[:(-new_field_sz+1)]):
            jndexes_now = df_haa.index[j:j+new_field_sz].copy()
            df_haaa = df_haa.T[jndexes_now].T.copy()
            if('method' in kwags.keys()):
                df_haaa = kwags['method'](
                        df_haaa, kwags=(kwags['method_kwags'] if(
                                'method_kwags' in kwags.keys()) else None))
            new_grp_no += 1
            np_df_haaa = np.mat(df_haaa)
            df_haaa_index = df_haaa.index.map(lambda x:str(x)+'F%d'%new_grp_no).copy()
            df_haaa = pd.DataFrame(np_df_haaa, columns = df_haaa.columns, 
                                   index = df_haaa_index)
            for ihd in initial_headers + ([initial_header] if(not initial_header in initial_headers) else []):
                if(not ihd in df_haaa):
                    continue
                initial_method_ = dcp(initial_methods.get(ihd, initial_method))
                if(initial_method_ == 'value min'):
                    df_haaa[ihd] = df_haaa[ihd] - np.min(df_haaa[ihd])
                elif(initial_method_ == 'value max'):
                    df_haaa[ihd] = df_haaa[ihd] - np.max(df_haaa[ihd])
                elif(initial_method_ == 'loc min'):
                    df_haaa[ihd] = df_haaa[ihd] - df_haaa[ihd].iloc[0]
                elif(initial_method_ == 'loc max'):
                    df_haaa[ihd] = df_haaa[ihd] - df_haaa[ihd].iloc[-1]
                elif(str(type(initial_method_)).find('func')>-1):
                    df_haaa[ihd] = initial_method_(df_haaa, header=ihd)
            df_haaa[eventname_header] = stamp_process('',stamps+[('' if(len(df0.index)//old_field_sz<=1) else i), 
                                                         ('' if(len(df_haa.index)//new_field_sz<=1) else 'sub%d'%j)],
                                                      '','','','_')
            df_new = df_haaa.copy() if(
                    df_new.empty) else append(df_haaa, df_new, sort=False)
    printer('new data shape:%s'%(str(df_new.shape)))
    return df_new

def refield_sequential_data_by_events(df0, new_field_sz, d=1, event_header=None, event_mask=None, df_new=None, 
                            initial_header=None, initial_headers=None, initial_method='value min', initial_methods=None, **kwags):
    addlog = kwags.get('addlog', LOGger.addloger(logfile=kwags.get('logfile','')))
    df_new = pd.DataFrame() if(isinstance(df_new, type(None))) else df_new
    initial_headers = initial_headers if(isinstance(initial_headers, list)) else [] #不隨時間變動的欄位
    initial_methods = initial_methods if(isinstance(initial_methods, dict)) else {} #不隨時間變動的欄位要怎麼定義
    # old_field_sz = float(old_field_sz) if(isinstance(astype(old_field_sz, float, default=None), float)) else df0.shape[0]
    event_mask = df0[event_header] if(True if(not isiterable(event_mask)) else np.array(event_mask).shape[0]!=df0.shape[0]) else np.array(event_mask)
    events = np.unique(event_mask)
    len_events = len(events)
    old_data_sz_undup = len(set(df0.index))
    addlog('old data size %d with events:%d'%(old_data_sz_undup, len_events))
    if(new_field_sz<=1 or ((new_field_sz//1)!=new_field_sz)):
        {}['new size error:%.2f'%(new_field_sz)]
    new_grp_no=0
    
    for event in events:
        df_haa = df0[df0[event_header]==event].copy()
        #|||||...|||||    ← df_haa
        #||...||          ← df_haa.index[0:n_fd_sz]
        # ||...||         ← df_haa.index[1:n_fd_sz+1]
        #  ||...||        ← df_haa.index[2:n_fd_sz+2]
        #      ...                ...
        #      ||...||    ← df_haa.index[-n_fd_sz-1:n_fd_sz-n_fd_sz-1]
        for j in range(0, df_haa.shape[0] - new_field_sz, d):
            df_haaa = df_haa.iloc[j:j+new_field_sz].copy()
            if(df_haaa.shape!=new_field_sz):
                continue
            if('method' in kwags.keys()):
                df_haaa = kwags['method'](
                        df_haaa, kwags=(kwags['method_kwags'] if(
                                'method_kwags' in kwags.keys()) else None))
            new_grp_no += 1
            np_df_haaa = np.mat(df_haaa)
            df_haaa_index = df_haaa.index.map(lambda x:str(x)+'F%d'%new_grp_no).copy()
            df_haaa = pd.DataFrame(np_df_haaa, columns = df_haaa.columns, 
                                   index = df_haaa_index)
            for ihd in initial_headers + ([initial_header] if(not initial_header in initial_headers) else []):
                if(not ihd in df_haaa):
                    continue
                initial_method_ = dcp(initial_methods.get(ihd, initial_method))
                if(initial_method_ == 'value min'):
                    df_haaa[ihd] = df_haaa[ihd] - np.min(df_haaa[ihd])
                elif(initial_method_ == 'value max'):
                    df_haaa[ihd] = df_haaa[ihd] - np.max(df_haaa[ihd])
                elif(initial_method_ == 'loc min'):
                    df_haaa[ihd] = df_haaa[ihd] - df_haaa[ihd].iloc[0]
                elif(initial_method_ == 'loc max'):
                    df_haaa[ihd] = df_haaa[ihd] - df_haaa[ihd].iloc[-1]
            df_new = df_haaa.copy() if(
                    df_new.empty) else append(df_haaa, df_new, sort=False)
    printer('new data shape:%s'%(str(df_new.shape)))
    return df_new

#TODO:trans_nan
def trans_nan(df, header=None, target_if_true=0, target_if_false=None, target_if_true_method=None, target_if_false_method=lambda x:x):
    header = header if(isiterable(header)) else []
    target_if_false_method = target_if_false_method if(not isinstance(target_if_false_method, type(None))) else lambda x:target_if_false
    target_if_true_method = target_if_true_method if(not isinstance(target_if_true_method, type(None))) else lambda x:target_if_true
    df_new = df.drop(header, axis=1).copy()
    for hd in header:
        df_new[hd] = df[hd].map(lambda s:target_if_true_method(s) if(LOGger.np_isnan(s)) else target_if_false_method(s))
    return df_new

#TODO:find_critical_windows 相關係數矩陣 rel cor
def find_critical_windows(cor, lbd=0.5, ubd=1, half_triangle=True):
    '''
    return selects_index, selects_value, selects_index_tag
    '''
    if(not hasattr(cor, 'shape')):
        return []
    if(cor.shape[0]!=cor.shape[1] if(len(cor.shape)==2) else True):
        return []
    iheader = cor.index if(isinstance(cor, pd.core.frame.DataFrame)) else np.arange(cor.shape[0])
    jheader = cor.columns if(isinstance(cor, pd.core.frame.DataFrame)) else np.arange(cor.shape[1])
    np_cor = np.array(cor) if(not isinstance(cor, np.ndarray)) else cor
    selects_index, selects_value, selects_index_tag = [], [], []
    for i,ih in enumerate(iheader):
        for j,jh in enumerate(jheader):
            if(half_triangle and j>i):
                continue
            if(np_cor[i,j]>=lbd and np_cor[i,j]<ubd):
                selects_index.append((i,j))
                selects_index_tag.append((ih,jh))
                selects_value.append(np_cor[i,j])
    return selects_index, selects_value, selects_index_tag
    
#TODO:summary_cor_header
def summary_cor_header(cor, lbd=0.5, ubd=1, half_triangle=True, sort=True, ascending=False):
    sel = find_critical_windows(cor, lbd=lbd, ubd=ubd, half_triangle=half_triangle)
    table = pd.DataFrame(list(zip(*[sel[1],*list(zip(*(sel[2])))])), columns=['ratio','header1','header2'])
    return table.sort_values(by='ratio', ascending=ascending) if(sort) else table

#TODO:summary_cor_header
def summary_cor_header_bidirectional(cor, lbd=0.5, ubd=1, half_triangle=True, sort=True, ascending=False):
    sel_pos = find_critical_windows(cor, lbd=lbd, ubd=ubd, half_triangle=half_triangle)
    table_pos = pd.DataFrame(list(zip(*[sel_pos[1],*list(zip(*(sel_pos[2])))])), columns=['ratio','header1','header2'])
    sel_neg = find_critical_windows(-cor, lbd=lbd, ubd=ubd, half_triangle=half_triangle)
    table_neg = pd.DataFrame(list(zip(*[list(map(lambda x:-x, sel_neg[1])),*list(zip(*(sel_neg[2])))])), columns=['ratio','header1','header2'])
    return (table_pos.sort_values(by='ratio', ascending=ascending), table_neg.sort_values(by='ratio', ascending=not ascending)) if(
        sort) else (table_pos, table_neg)

def shapeUnify(tensor, shape, default=0.0, **kwags):
    temp = dcp(tensor)
    approShape = (min(shape[0], temp.shape[0]), min(shape[1], temp.shape[1]), min(shape[2], temp.shape[2]))
    tensor = np.full(approShape, default)
    tensor[:approShape[0], :approShape[1], :approShape[2]] = temp[:approShape[0], :approShape[1], :approShape[2]]
    return tensor

def padCropPadEdge(cell, cell_size_target):
    cell = np.array(cell)
    old_shape = cell.shape
    if(len(old_shape)!=len(cell_size_target)):
        return None
    pad_width = []
    for i,sh in enumerate(cell_size_target):
        diff = max(0, sh - old_shape[i])
        pad_width.append([0,diff])
    cell = np.pad(cell, pad_width=pad_width, mode='edge')
    slices = []
    for i,sh in enumerate(cell_size_target):
        ubd = max(0, sh)
        slices.append(slice(0, ubd))
    slices = tuple(slices)
    cell = cell[slices]
    return cell

def padCropPadConstant(cell, cell_size_target, value=0):
    old_shape = cell.shape
    if(len(old_shape)!=len(cell_size_target)):
        return None
    pad_width = []
    for i,sh in enumerate(cell_size_target):
        diff = max(0, sh - old_shape[i])
        pad_width.append(diff)
    cell = np.pad(cell, pad_width=pad_width, mode='constant', constant_values=value)
    slices = []
    for i,sh in enumerate(cell_size_target):
        ubd = max(0, sh)
        slices.append(slice(0, ubd))
    slices = tuple(slices)
    cell = cell[slices]
    return cell

#TODO:merge_header
def merge_header(df, method=lambda tensor,**kwags:np.sign(np.sum(tensor,axis=1)), old_header=None, new_header=None):
    old_header = old_header if(isiterable(old_header)) else ([old_header] if(isinstance(old_header, str)) else [])
    old_header = [k for k in old_header if k in df]
    new_header = new_header if(isinstance(new_header, str)) else 'new_header'
    df_new = df.drop(old_header, axis=1).copy()
    df_new[new_header] = method(df[old_header])
    return df_new

def DeleteColThatHasCells(df_data, **kwags):
    ColThatHasNoCellsMask = np.full(df_data.shape[1], True)
    for i,col in enumerate(df_data.columns):
        values = df_data[col].values
        for v in values:
            if(isiterable(v) or getattr(v, 'shape', ())!=()):
                # LOGger.addDebug('DeleteColThatHasCells %s\n'%str(type(v)), v, stamps=[col])
                ColThatHasNoCellsMask[i] = False
                break
    df_data = df_data[df_data.columns[ColThatHasNoCellsMask]]
    return df_data

def save(df_data, sht=None, fn='', exp_fd='test', seq_sz=1, save_types=['xlsx', 'pkl'], 
         data_counter=None, data_counter_stamp=None, data_difference_to_save=3000, 
         del_nan_data=True, nan_data_allowed_header=None, xlsx_data_size_ubd=60000, rewrite=True, 
         xlsx_dont_save_header=None, isDeleteColThatHasCells=True, **kwags):
    if(df_data.empty):
        return
    addlog = kwags.get('addlog', LOGger.addloger(logfile = kwags.get('logfile', os.path.join(exp_fd, 'log.txt'))))
    if(isinstance(data_counter, dict) and isinstance(data_counter_stamp, str)):
        if(data_counter.get(data_counter_stamp, -np.inf) == df_data.shape[0]):
            return
        if(np.abs(df_data.shape[0] - data_counter.get(data_counter_stamp, -np.inf))<data_difference_to_save):
            # print(data_counter)
            return
        data_counter.update({data_counter_stamp:df_data.shape[0]})
        # print('--||%s'%data_counter)
    if(df_data.index.duplicated().any()):
        addlog('duplicated indexes!!!!!')
        addlog('%s'%(df_data[np.logical_not(
                df_data.index.duplicated(keep=False))]))
    nan_data_allowed_header = nan_data_allowed_header if(isinstance(nan_data_allowed_header, list)) else []
    nan_data_allowed_header = [v for v in df_data.columns]
    if(del_nan_data and df_data.drop(nan_data_allowed_header, axis=1).T.isna().any().any()):
        addlog('find nan values!!!!!')
        addlog('%s'%(df_data[df_data.T.isna().any()]))
        df_data = df_data[np.logical_not(df_data.T.isna().any())]
    if('pkl' in save_types):
        file = os.path.join(exp_fd, '%s.pkl'%(
            stamp_process('',[fn,(sht if(isinstance(sht, str)) else '')],'','','','_',for_file=1)))
        if(not rewrite): file = pathrpt(file)
        LOGger.CreateFile(file, lambda f:df_data.to_pickle(f))
    if('csv' in save_types):
        if(isDeleteColThatHasCells):   df_data = DeleteColThatHasCells(df_data)
        file = os.path.join(exp_fd, '%s.csv'%(
            stamp_process('',[fn,(sht if(isinstance(sht, str)) else '')],'','','','_',for_file=1)))
        if(not rewrite): file = pathrpt(file)
        LOGger.CreateFile(file, lambda f:df_data.to_csv(f))
    if('xlsx' in save_types):
        if(LOGger.isinstance_not_empty(exp_fd, str)):
            if(not os.path.isdir(exp_fd)):
                LOGger.CreateContainer(exp_fd)
        xlsx_dont_save_header = xlsx_dont_save_header if(isinstance(xlsx_dont_save_header, list)) else []
        xlsx_dont_save_header = [v for v in xlsx_dont_save_header if v in df_data.columns]
        sht = sht if(isinstance(sht, str)) else 'Sheet1'
        for i in range(0,df_data.shape[0],xlsx_data_size_ubd):
            file = os.path.join(exp_fd, '%s.xlsx'%(stamp_process(
                '',[fn,('' if(df_data.shape[0]<xlsx_data_size_ubd) else '--%d'%i)],'','','','_',for_file=1)))
            if(not rewrite): file = pathrpt(file)
            eva = evaluation()
            eva.read(file)
            df_data = df_data.drop(xlsx_dont_save_header, axis=1).iloc[i:i+xlsx_data_size_ubd].copy()
            if(isDeleteColThatHasCells):   df_data = DeleteColThatHasCells(df_data)
            eva.storageboxes[sht] = df_data
            eva.export(file)

def nppad(cell, cell_size, default_pad_direction=1, default_pad_front_back=None, **kwags):
    cell = dcp(cell if(getattr(cell, 'shape', ())!=()) else np.full(cell_size, cell))
    if(cell.shape!=cell_size):
        cell_shape = cell.shape
        pad_front_backs = []
        for ii in range(min(len(cell_size), len(cell.shape))):
            pad_length = dcp(max(cell_size[ii] - cell_shape[ii], 0))
            if(default_pad_direction>0):
                pad_front_backs.append((0,pad_length))
            elif(default_pad_direction<0):
                pad_front_backs.append((pad_length,0))
            else:
                pad_front_backs.append(
                    int(pad_length*(default_pad_front_back.get(0,0)/(default_pad_front_back.get(0,0)+default_pad_front_back.get(1,1)))),
                    int(pad_length*(default_pad_front_back.get(1,1)/(default_pad_front_back.get(0,0)+default_pad_front_back.get(1,1)))))
        pad_front_backs = tuple(pad_front_backs)
        cell_paded = np.pad(cell, pad_front_backs, mode=kwags.get('mode','constant'), constant_values=kwags.get('constant_values',0))
    else:
        cell_paded = dcp(cell)
    return cell_paded

def npexpand(cell, cell_size, default_pad_direction=1, default_pad_front_back=None, expand_value=0, **kwags):
    cell = dcp(cell if(getattr(cell, 'shape', ())!=()) else np.full(cell_size, cell))
    if(isiterable(cell_size)):
        cell_expanded = np.full(cell_size, expand_value)
        slices = tuple(slice(0, cell.shape[i]) for i in range(len(cell.shape))) + (0, )*max(len(cell_size) - len(cell.shape), 0)
        cell_expanded[slices] = cell #TODO:為什麼沒用
    else:
        cell_expanded = dcp(cell)
    return cell_expanded

def npcrop(cell, cell_size, default_pad_direction=1, default_pad_front_back=None, **kwags):
    pass

def padding(array, backward_step=0, forward_step=0, preserve_all_data=False,
            default_value=None, key_index=None,
            resize_tend=-1, seq_sym='#', **kwags):
    str_array_type = str(type(array))
    if(isinstance(array, list)):
        forward_step = ((forward_step+len(array)-key_index) if(
                        key_index!=None) else (
                            forward_step+len(array))) if(
                                preserve_all_data) else forward_step
        backward_step = ((backward_step+key_index) if(
                        key_index!=None and preserve_all_data) else (
                            backward_step))
        new_array=[]
        default_value = default_value if(not isinstance(default_value, type(None)) and
                                         not isiterable(default_value)) else 0
        if(default_value==seq_sym):
            new_array = ['%s-%d'%(seq_sym, v+1) for v in reversed(
                    range(backward_step))] + array + [
                         '%s%d'%(seq_sym, v+1) for v in range(forward_step)]
        elif(default_value=='-%s'%seq_sym):
            old_head = array[0]
            old_tail = array[-1]
            new_array = ['%s%s-%d'%(str(old_head), seq_sym, v+1) for v in reversed(
                         range(backward_step))] + array + [
                         '%s%s%d'%(str(old_tail), seq_sym, v+1) for v in range(forward_step)]
        elif(default_value=='extend'):
            old_head = array[0]
            old_tail = array[-1]
            new_array = [old_head for v in reversed(range(backward_step))] + array + [
                         old_tail for v in range(forward_step)]
        else:
            new_array = [default_value]*backward_step + array + [default_value]*forward_step
        if(key_index!=None):
            if(key_index<-1 or key_index>len(array)):
                return []
            if(resize_tend==1):
                new_array = new_array[(key_index+1):(key_index + backward_step + forward_step+2)]
            if(resize_tend==0):
                new_array = new_array[(key_index):(key_index + backward_step + forward_step+1)]
            if(resize_tend==-1):
                new_array = new_array[(key_index):(key_index + backward_step + forward_step)]
    if(isinstance(array, np.ndarray)):
        default_value = default_value if(default_value) else 0
        array = array.tolist()
        new_array = padding(array, backward_step, forward_step, 
                            preserve_all_data=preserve_all_data,
                            default_value=default_value, key_index=key_index,
                            resize_tend=resize_tend, seq_sym=seq_sym, kwags=kwags)
        new_array = np.array(new_array)
    if(str_array_type.find('pandas.core')>-1):
        if(str_array_type.find('series')>-1):
            default_value = default_value if(default_value) else 0
            index_default_values = kwags['index_default_values'] if(
                    'index_default_values' in kwags.keys()) else '-%s'%seq_sym
            values = np.array(array.values)
            indexes = np.array(array.index)
            key_index = list(indexes).index(kwags['Index']) if(
                    'Index' in kwags.keys()) else key_index
            new_values = padding(values, backward_step, forward_step, 
                                 preserve_all_data = preserve_all_data,
                                 default_value=default_value, key_index=key_index,
                                 resize_tend=resize_tend, seq_sym=seq_sym, kwags=kwags)
            indexes = padding(indexes, backward_step, forward_step, 
                              preserve_all_data = preserve_all_data,
                              default_value=index_default_values, key_index=key_index,
                              resize_tend=resize_tend, seq_sym=seq_sym, kwags=kwags)
            new_array = pd.Series(new_values, index = indexes)
        if(str_array_type.find('frame')>-1):
            default_value = default_value if(default_value) else ''
            index_default_values = kwags['index_default_values'] if(
                    'index_default_values' in kwags.keys()) else '-%s'%seq_sym
            indexes = array.index.copy()
            key_index = list(indexes).index(kwags['Index']) if(
                    'Index' in kwags.keys()) else key_index
            key_index = indexes[0] if(key_index==None) else key_index
            new_array = pd.DataFrame()
            for col in array.columns:
                array_col = array[col].copy()
                excute_default_value = (kwags['default_dict'][col] if(
                        col in kwags['default_dict'].keys()) else default_value) if(
                        'default_dict' in kwags) else default_value
                new_array_col = padding(array_col, backward_step, forward_step, 
                        preserve_all_data = preserve_all_data,
                        default_value=excute_default_value, key_index=key_index,
                        resize_tend=resize_tend, seq_sym=seq_sym, kwags=kwags, 
                        Index=(kwags['Index'] if(
                                'Index' in kwags.keys()) else key_index))
                new_array[col] = new_array_col.copy()
    return new_array

def customization_curve_fit(f, xdata, ydata, p0=None, sigma=None, 
                       absolute_sigma=False, check_finite=True, 
                       method=None, jac=None, return_executor=True,
                       default_method='linear', ret=None, maxfev=5000, full_output=False, **kwags):
    addlog = LOGger.execute('addlog', kwags, default=LOGger.addloger(logfile=''), not_found_alarm=False)
    ret = ret if(isinstance(ret, dict)) else {}
    ret['msg'] = ''
    default_method = (lambda t,a,b:a*t+b) if(
            default_method=='linear') else default_method
    executor, popt, pcov, infodict = None, None, None, {}
    if(len(set(ydata))==1):
        const = list(set(ydata))[0]
        executor = lambda t:const
    else:
        try:
            if(full_output):
                popt, pcov, infodict = curve_fit(f, xdata, ydata, p0=p0, maxfev=maxfev, full_output=True)
                ret.update(infodict)
            else:
                popt, pcov = curve_fit(f, xdata, ydata, p0=p0, maxfev=maxfev)
            executor = (lambda t: f(t, *popt))
        except RuntimeError as e:
            addlog('RuntimeError!!! using default_method!!!')
            LOGger.exception_process(e, logfile='')
            if(default_method):
                addlog('Using default_method!!!')
                popt, pcov = curve_fit(default_method, xdata, ydata, p0=None, **kwags)
                f = default_method
            else:
                return False
    ret.update({'popt':popt, 'pcov':pcov})
    if(return_executor):
        ret['executor'] = executor
    return True

def resolSequentialView(*datas, ret=None):
    all_values = get_all_values([datas])
    ymdn = np.median(all_values)
    datasResed = []
    for i,data in enumerate(datas):
        if(np.array(data).shape[0]==0):
            if(isinstance(ret, dict)):  ret[i] = {}
            datasResed.append([])
            continue
        y_tomdn = np.array(data) - ymdn
        ystd2 = np.nanmedian(np.abs(y_tomdn))
        dataResed = list(tuple(np.where(
            np.abs(y_tomdn) > 3*ystd2, np.where(y_tomdn > 0, ymdn + 3*ystd2, ymdn - 3*ystd2), ymdn + y_tomdn)))
        if(isinstance(ret, dict)):
            infrm = {}
            infrm['mdn'] = y_tomdn
            infrm['std'] = ystd2
            infrm['end'] = data[-1] if(np.array(data).shape[0]>0) else np.nan
            infrm['min'] = np.nanmin(data) if(np.array(data).shape[0]>0) else np.nan
            infrm['max'] = np.nanmax(data) if(np.array(data).shape[0]>0) else np.nan
            try:
                infrm['nan'] = np.isnan(data).any()
            except Exception as e:
                infrm['nan'] = '-'
            ret[i] = dcp(infrm)
        datasResed.append(dataResed)
    return tuple(datasResed)    

def scipyMinizingTarget(score_fun=(lambda x:x**2+1), X0=0, method='Nelder-Mead', history=None, stamps=None,
                        bounds=None, options={'disp': True}, ret=None, callback=None, 
                        draw_loss_curve_method=None, exp_fd='.', xheader=None, fixHeader=None, tol=1e-4, **kwags):
    addlog = LOGger.execute('addlog', kwags, default=LOGger.addloger(logfile=''), not_found_alarm=False)
    fixHeader = fixHeader if(isinstance(fixHeader,dict)) else {}
    score_fun_exact = score_fun
    
    xheader = LOGger.mylist(tuple(
        xheader if(isiterable(xheader)) else (np.arange(np.array(X0).shape[0]) if(isiterable(X0)) else [])))
    if(fixHeader and isiterable(xheader)):
        fixIndexes = {x:xheader.index(x) for x in fixHeader if x in xheader}
        X0 = [x for i,x in enumerate(X0) if i not in fixIndexes] if(isiterable(X0)) else X0
        def score_fun_exact(x):
            __x__ = dcp(x.astype(float))
            for hd in fixHeader:
                __x__ = np.insert(__x__, int(fixIndexes[hd]), fixHeader[hd])
            return score_fun(__x__)
    stamps=stamps if(isinstance(stamps, list)) else []
    res = LOGger.mystr()
    res.nit = 0
    res.success = False
    res.fun = np.nan
    res.x = np.full(np.array(X0).shape, np.nan) if(hasattr(X0,'shape')) else np.nan
    if(isinstance(ret, dict)):  ret['res'] = res
    try:
        res = minimize((lambda X:mylist([score_fun_exact(X)]).get_all()[-1]), X0, method=method, 
                       callback=(lambda d:callback(
                           d, score_fun_exact, history=history, stamps=stamps, exp_fd=exp_fd, 
                           draw_loss_curve_method=draw_loss_curve_method, 
                           PAW_Standard=kwags.get('PAW_Standard',None))) if(callback) else None, 
                           bounds=bounds, options=options, tol=tol)
    except Exception as e:
        LOGger.exception_process(e, logfile=os.path.join(exp_fd, 'log.txt'), stamps=['scipy_local_optimizing'])
        addlog('%s優化失敗%s'%(LOGger.stamp_process('',stamps), stamp_process('',{
            'X0 shape':np.array(X0).shape, 'X0':','.join(list(map(lambda s:parse(s,2), X0))[:30])})))
        messages = []
        try:
            addlog(' ',stamps={'score_fun(X0)':parse(score_fun(X0), 4)})
        except Exception as e:
            LOGger.exception_process(e, logfile=os.path.join(exp_fd, 'log.txt'))
            messages += ['score fun is not compatible with X0!']
        if(bounds!=None):
            if(not isiterable(bounds) and not isinstance(bounds, Bounds)):
                messages += ['bounds type error: %s'%LOGger.type_string(bounds)]
            else:
                if(np.array(bounds).shape[0]!=np.array(X0).shape[0]):
                    messages += ['bound size is not compatible with inputs!']
        res.message = LOGger.stamp_process('',messages,'','','','\n')
        return False
    if(isinstance(ret, dict)):  ret['res'] = res
    if(res.nit==0):
        res.message = LOGger.stamp_process('',[res.message, 'No iterations!!!'],'','','','\n')
        return False
    return True

def scipyMinizingTargets(score_fun=(lambda x:x**2+1), targets=np.array([0]), starts=None, algorithms=None,boundses=None,tols=None,
                         return_success=False,return_message=False,measure=None,
                         default_start=0, default_algorithm='Nelder-Mead',default_bound=None,default_tol=None,
                         stamps=None,history=None,exp_fd='.',ret=None,**kwags):
    """
    
    Parameters
    ----------
    score_fun : callable
        output dimesion must be the same with targets. 
    targets : iterable of constants
        dimension must be the same with outputs of score_fun. 
    default_start : TYPE, optional
        DESCRIPTION. The default is 0.
    default_algorithm : TYPE, optional
        DESCRIPTION. The default is 'Nelder-Mead'.
    default_bound : TYPE, optional
        DESCRIPTION. The default is None.
    default_tol : TYPE, optional
        DESCRIPTION. The default is None.
    measure : TYPE, optional
        沒有指定measure，就採用歐式測距測量輸出目標(假定輸出與目標同樣的維度). The default is None.
    history : TYPE, optional
        回傳歷史分數，要用`dict` key 就是targets的index. The default is None.
    **kwags : TYPE
        DESCRIPTION.

    Returns
    -------
    results : np.ndarray

    """    
    starts = LOGger.mylist(starts if(isinstance(starts, list)) else [default_start])
    algorithms=LOGger.mylist(algorithms if(isinstance(algorithms, list)) else [])
    boundses=LOGger.mylist(boundses if(isinstance(boundses, list)) else [])
    tols=LOGger.mylist(tols if(isinstance(tols, list)) else [])
    stamps=stamps if(isinstance(stamps, list)) else []
    history=history if(isinstance(history, dict)) else {}
    kwags['options']=kwags.get('options', {'disp': False})
    
    if(measure==None):
        #沒有指定measure，就採用歐式測距測量輸出目標(假定輸出與目標同樣的維度)
        shape = getattr(score_fun(starts[0]), 'shape', ())
        if(shape==()):
            measure = lambda p:(lambda x:abs(score_fun(x) - p))
        else:
            measure = lambda p:(lambda x:np.sqrt(np.sum(score_fun(x).reshape(-1) - p.reshape(-1))**2))
    suggest_infrms, reTemp=[], {}
    for i,p in enumerate(targets):
        reTemp.clear()
        scipyMinizingTarget(measure(p), X0=starts.get(i,default_start), 
                            method=algorithms.get(i,default_algorithm), 
                            bounds=boundses.get(i,default_bound),tol=tols.get(i,default_tol),
                            history=history.get(i,None), stamps=stamps+[i],
                            ret=reTemp, exp_fd=exp_fd, **kwags)
        res = dcp(reTemp['res'])
        suggest_infrm = [res.x,res.fun]
        reTemp['success'] = res.success
        reTemp['message'] = res.message
        if(isinstance(ret,dict)):   ret[i] = dcp(reTemp)
        suggest_infrms.append(suggest_infrm)
    return np.array(suggest_infrms)

def reshapeThruFlatten(array, shape=None, default_pad_value=0, default=None, slicesByChannel=None, **kwags):
    if(not isiterable(shape)):
        return array
    
    if(not isinstance(slicesByChannel, dict)):
        slicesByChannel = LOGger.mylist(slicesByChannel if(isinstance(slicesByChannel, list)) else [])
    arrayFlatten = np.array(LOGger.get_all_values(getattr(array, 'values', array)))
    dtype = getattr(getattr(array,'values',array),'dtype',float)
    shapeProduct = np.product(shape)
    array_temp = np.full(shapeProduct, default_pad_value, dtype=dtype)
    
    slices = slice(min(shapeProduct, arrayFlatten.shape[0]))
    array_temp[slices] = arrayFlatten[slices]
    array_temp = array_temp.reshape(*shape)
    return array_temp

def reshapeThruDimensions(array, shape=None, default_pad_value=0, default=None, slicesByChannelOutcome=None, slicesByChannelResource=None, **kwags):
    if(not isiterable(shape)):
        return array
    
    if(not isinstance(slicesByChannelOutcome, dict)):  
        slicesByChannelOutcome = LOGger.mylist(slicesByChannelOutcome if(isinstance(slicesByChannelOutcome, list)) else [])
    if(not isinstance(slicesByChannelResource, dict)):  
        slicesByChannelResource = LOGger.mylist(slicesByChannelResource if(isinstance(slicesByChannelResource, list)) else [])
    
    dtype = getattr(getattr(array,'values',array),'dtype',float)
    array_temp = np.full(shape, default_pad_value, dtype=dtype)
    new_slices = [slicesByChannelOutcome.get(j, slice(min(shape[j], array.shape[j]))) if(j<len(array.shape)) else 0 for j,sh in enumerate(shape) if(j<len(shape))]
    old_slices = [slicesByChannelResource.get(j, slice(min(shape[j], array.shape[j]))) if(j<len(shape)) else 0 for j,sh in enumerate(array.shape)]
    array_temp[tuple(new_slices)] = array[tuple(old_slices)]
    return array_temp

def transformCellToFlat(data, cell_size=None, editor=None, pad_value=0, **kwags):
    addlog = LOGger.addloger(logfile='')
    cell_size = cell_size if(isiterable(cell_size)) else data.shape
    column_dimension = int(np.product(cell_size))
    if(isinstance(data,np.ndarray)):
        data = dcp(data)
    elif(isiterable(data)):
        data = np.array(data)
    else:
        addlog('transformCellToFlat data type error', colora=LOGger.FAIL)
        joblib.dump({'data':data, 'cell_size':cell_size}, 'debug.pkl')
        sys.exit(1)
    cellTemp = dcp(data)
    if(hasattr(editor, '__call__')):
        cellTemp = editor(data)
    cellTemp = reshapeThruFlatten(cellTemp, shape=(column_dimension,), default_pad_value=pad_value) if(tuple((column_dimension,))!=data.shape) else dcp(data)
    return cellTemp

def annilateGroupStructure(tensor, groupIndex=1, **kwags):
    cell_size_annil_group = tuple(map(int, np.array(tensor.shape)[np.array([int(i) for i,x in enumerate(tensor.shape) if i!=groupIndex],dtype=int)]))
    return reshapeThruFlatten(tensor, shape=cell_size_annil_group)

def transformCellsToFlatBatch(tensor, cell_size=None, editor=None, pad_value=0, feaVarAxis=None, **kwags):
    cell_size = cell_size if(isiterable(cell_size)) else tensor.shape[1:]
    if(feaVarAxis is not None):
        totalShape = (np.array(tensor).shape[0], *cell_size)
        tensorTemp = dcp(tensor)
        if(not hasattr(editor, '__call__')):
            # 不用 editor 走這裡比較快
            if(feaVarAxis%len(totalShape)!=-1%len(totalShape)): tensorTemp = np.moveaxis(tensor, feaVarAxis, -1)
            otherAxises = np.array(totalShape)[np.arange(len(totalShape))!=feaVarAxis%len(totalShape)]
            return tensorTemp.reshape(np.product(otherAxises), -1)
        else:
            if(feaVarAxis%len(totalShape)!=-1%len(totalShape)): tensorTemp = np.moveaxis(tensor, feaVarAxis, 0)
            totalShapeTemp = (totalShape[feaVarAxis], *(np.array(totalShape)[np.where(np.arange(len(totalShape))!=feaVarAxis%len(totalShape))]))
            tensorTemp = transformCellsToFlatBatch(tensorTemp, cell_size=totalShapeTemp[1:], editor=editor, pad_value=pad_value, feaVarAxis=0, **kwags)
            return np.transpose(tensorTemp)
    method = lambda x:transformCellToFlat(x, cell_size=cell_size, editor=editor, pad_value=pad_value, **kwags)
    return transformByBatch(tensor, method)

def transformGroupofCellsToFlatBatch(tensor, cell_size=None, editor=None, groupIndex=1, pad_value=0, **kwags):
    tensor = annilateGroupStructure(tensor, groupIndex=1)
    tensorFlat = transformCellsToFlatBatch(tensor, cell_size=cell_size, editor=editor, pad_value=pad_value, **kwags)
    return tensorFlat

def transformFlatToCell(data, cell_size=None, editor=None, pad_value=0, **kwags):
    addlog = kwags.get('addlog', LOGger.addloger(logfile=''))
    # cell_size = cell_size if(isiterable(cell_size)) else data.shape
    cellTemp = dcp(np.full(cell_size, pad_value))
    if(isinstance(data,np.ndarray)):
        if(len(data.shape)!=1):
            addlog("transformFlatToCell input shape error:", str(getattr(data,'shape',())), colora=LOGger.Fore.RED)
            sys.exit(1)
        cellTemp = dcp(data)
    if(isiterable(cell_size)):
        cellTemp = reshapeThruFlatten(cellTemp, shape=cell_size, pad_value=pad_value)
    if(hasattr(editor, '__call__')):
        cellTemp = editor(cellTemp)
    return cellTemp

def transformFlatBatchToCells(tensor, cell_size=None, editor=None, preprocessor=None, pad_value=0, feaVarAxis=None, **kwags):
    cell_size = cell_size if(isiterable(cell_size)) else tensor.shape[1:]
    tensorTemp = dcp(tensor)
    ShapeDimProduction = np.product(tensor.shape)
    batchDim = int(ShapeDimProduction//np.product(cell_size))
    totalShape = (batchDim, *cell_size)
    if(feaVarAxis is not None):
        tensorTemp = tensorTemp.reshape(batchDim, np.product(cell_size))
        tensorTemp = transformFlatBatchToCells(tensorTemp, cell_size=cell_size, editor=editor, preprocessor=preprocessor, 
                                               pad_value=pad_value, feaVarAxis=None, **kwags)
        if(feaVarAxis%len(totalShape)!=-1%len(totalShape)):
            tensorTemp = np.moveaxis(tensorTemp, -1, feaVarAxis)
        return tensorTemp
    method = lambda x:transformFlatToCell(x, cell_size=cell_size, editor=editor, pad_value=pad_value, **kwags)
    tensorTemp = transformByBatch(tensor, method)
    return tensorTemp

def transformByBatch(tensor, method, stamps=None):
    # 将输入转换为 NumPy 数组
    if isinstance(tensor, pd.DataFrame) or isinstance(tensor, pd.Series):
        tensor_array = tensor.to_numpy()
    elif hasattr(tensor, 'numpy'):  # 处理 TensorFlow 张量
        tensor_array = tensor.numpy()
    else:
        tensor_array = np.array(tensor)
    
    # 调用 method 的批量操作（假设 method 支持矢量化）
    try:
        # tensor_temp = method(tensor_array)
        tensor_temp = np.apply_along_axis(method, axis=np.arange(len(tensor_array.shape))[1:], arr=tensor_array)
    except Exception:
        # 如果 method 不支持矢量化，则分批处理
        tensor_temp = np.array([method(cell) for cell in tensor_array])
    
    return tensor_temp

def transformFlatBatchToGroupofCells(tensor, cell_size=None, preprocessor=None, groupIndex=1, pad_value=0, **kwags):
    tensorCells = transformFlatBatchToCells(tensor, cell_size=cell_size, preprocessor=preprocessor, pad_value=pad_value, **kwags)
    tensorOutput = np.expand_dims(tensorCells, axis=groupIndex)
    return tensorOutput

def transformCells(data, cell_size=None, preprocessor=None, pad_value=0, editor=None, thruMethod=None, **kwags):
    editorAug = editor
    if(isiterable(cell_size)):
        if(tuple(cell_size)!=data.shape[1:]):
            def editorAug(x):
                if(editor): x = editor(x)
                x = thruMethod(x, shape=cell_size, pad_value=pad_value)
                return x
    # 通常inputs有data frame結構，所以shape=(None, 1, *cell_size) or (None, 1)
    data = transformGroupofCellsToFlatBatch(data, cell_size=cell_size, pad_value=pad_value, editor=editorAug, **kwags) if(
        isiterable(cell_size)) else transformCellsToFlatBatch(data, cell_size=cell_size, pad_value=pad_value, editor=editorAug, **kwags)
    if(preprocessor):   data = preprocessor.transform(data)
    data = transformFlatBatchToCells(data, cell_size=cell_size, pad_value=pad_value, **kwags)
    return np.array(data)

def inverse_transformFlatToCell(data, cell_size=None, editor=None, pad_value=0, **kwags):
    return transformFlatToCell(data, cell_size=cell_size, editor=editor, pad_value=pad_value, **kwags)

def inverse_transformCellsToFlatBatch(tensor, cell_size=None, editor=None, pad_value=0, **kwags):
    return transformCellsToFlatBatch(tensor, cell_size=cell_size, editor=editor, pad_value=pad_value, **kwags)

def inverse_transformFlatBatchToCells(tensor, cell_size=None, editor=None, preprocessor=None, pad_value=0, **kwags):
    cell_size = cell_size if(isiterable(cell_size)) else tensor.shape[1:]
    method = lambda x:inverse_transformFlatToCell(x, cell_size=cell_size, editor=editor, pad_value=pad_value, **kwags)
    return transformByBatch(tensor, method)

def inverse_transformFlatBatchToGroupofCells(tensor, cell_size=None, preprocessor=None, groupIndex=1, pad_value=0, **kwags):
    tensorCells = inverse_transformFlatBatchToCells(tensor, cell_size=cell_size, preprocessor=preprocessor, pad_value=pad_value, **kwags)
    tensorOutput = np.expand_dims(tensorCells, axis=groupIndex)
    return tensorOutput

def inverse_transformCells(data, cell_size=None, preprocessor=None, pad_value=0, editor=None, thruMethod=None, **kwags):
    editorAug = editor
    if(isiterable(cell_size)):
        if(tuple(cell_size)!=data.shape[1:]):
            def editorAug(x):
                if(editor): x = editor(x)
                x = thruMethod(x, shape=cell_size, pad_value=pad_value)
                return x
    # 通常inputs有data frame結構，所以shape=(None, 1, *cell_size) or (None, 1)
    data = inverse_transformCellsToFlatBatch(data, cell_size=cell_size, pad_value=pad_value, **kwags)
    if(preprocessor):   data = preprocessor.inverse_transform(data)
    data = inverse_transformFlatBatchToGroupofCells(data, cell_size=cell_size, pad_value=pad_value, editor=editorAug, **kwags) if(
        isiterable(cell_size)) else inverse_transformFlatBatchToCells(data, cell_size=cell_size, pad_value=pad_value, editor=editorAug, **kwags)
    return np.array(data)

#TODO:packaging
def packaging(df, cell_size, is_padding=False, Index=[], **kwags):
    dF = pd.DataFrame()
    for I, i in enumerate(range(0, len(np.array(df).shape[0], cell_size))):
        df_i = df.iloc[i:i+cell_size]
        if(np.array(df_i).shape<cell_size):
            d = cell_size - np.array(df_i)
            if(is_padding):
                padding(df_i, 0, d)
            else:
                printer('[%d][%s]模組化不整除%d(%d)!!!'%(i, df_i.index[i], np.array(df_i).shape, cell_size))
                continue
        dI = I if(I>=np.array(Index).shape[0]) else Index[I]
        dF = pd.DataFrame([df_i.copy()], index=[dI]) if(dF.empty) else append(dF, df_i)
    return dF

#TODO:inverse_packaging
def inverse_packaging(dF, cell_size, default_col=0, is_attach_index=False):
    df = pd.DataFrame()
    for dI in dF.index:
        df_i = dF[default_col][dI]
        if(np.array(df_i).shape!=cell_size):
            printer('[%s]反模組化不整除%d(%d)!!!'%(dI, np.array(df_i).shape, cell_size))
        if(is_attach_index):
            df_i['set_index'] = list(map(lambda s: '%s_%s'%(dI, s), df_i.index))
            df_i = df_i.set_index('set_index')
        df = df_i.copy() if(df.empty) else append(df, df_i)
    return df
    
def interpolation(xxx, yyy, method='linear', x_apply=None,
                  show_graph=False, ret=None, stamps=None, **kwags):
    ret = ret if(isinstance(ret, dict)) else {}
    stamps = stamps if(isinstance(stamps, list)) else []
    stamp = LOGger.stamp_process('',stamps,'','','','_')
    method = (lambda t,a,b:a*t+b) if(method=='linear') else method
    np_xxx = np.array(xxx)
    np_yyy = np.array(yyy)
    if(len(np_yyy.shape)>1):
        ys, ects = [], {}
        for j in range(np_yyy.shape[1]):
            yyyy = np_yyy[:,j]
            kwags.update({'j':j})
            if(not interpolation(xxx, yyyy, method=method, x_apply=x_apply,
                  show_graph=show_graph, stamps=[j], ret=ret, **kwags)):
                return False
        ret['executor'] = lambda xxxx:np.transpose(np.array([ret['executors'][parse(j)](xxxx) for j in range(np_yyy.shape[1])]))
        if(not isinstance(x_apply, type(None))):
            ret['ans'] = ret['executor'](x_apply)
    else:
        #測試y是否為數字
        try:
            np_yyy = np_yyy.astype(float)
        except Exception as e:
            LOGger.exception_process(e, logfile='')
            return False
        else:
            if(not customization_curve_fit(method, np_xxx, np_yyy, ret=ret, **kwags)):
                print('customization_curve_fit failedd')
                return False
        x_apply = x_apply if(not isinstance(x_apply, type(None))) else np_xxx
        np_ppp = ret['executor'](x_apply)
        if(stamp):
            if(not 'executors' in ret):
                ret['executors'] = {}
            ret['executors'][stamp] = dcp(ret['executor'])
        # print(str(np_ppp)[:500])
        ret['ans'] = np_ppp.reshape(-1,1) if(isinstance(ret.get('ans'), type(None))) else np.hstack([ret['ans'], np_ppp.reshape(-1,1)])
    return True

display_max_columns = 0
display_max_rows = 0
def data_polyinterpolation(data, n_anchor=5, anchors=None, main_axis=0, main_coln=None, 
                  interpolation_method = 'linear', keephead=True, keeptail=True, 
                  numeric=True, base_diff_ubd=np.inf, **kwags):
    if(len(data.index)==n_anchor):
        return data
    main_coln = data.columns[main_axis] if(main_coln==None) else main_coln
    main_col = data[main_coln].copy()
    
    
    crtpts = tuple(np.where(np.diff(main_col)>base_diff_ubd)[0])
    if(isinstance(anchors, type(None)) and crtpts):
        crtpts += (np.array(main_col).shape[0],)
        n_anchor_section = int(n_anchor//len(crtpts))
        new_data =  pd.DataFrame()
        for i,crtpt in enumerate(crtpts):
            crtpt_a = dcp(crtpts[i-1]+1) if(i>0) else 0
            data_crtsec = data.iloc[crtpt_a:(crtpt+1)].copy()
            if(data_crtsec.shape[0]<2):
                return None
            new_data_crtsec = data_polyinterpolation(data_crtsec, n_anchor=n_anchor_section, anchors=None, main_axis=main_axis, main_coln=main_coln, 
                                                     interpolation_method = interpolation_method, keephead=keephead if(i==0) else True, 
                                                     keeptail=keeptail if(i==len(crtpts)-1) else True, 
                                                     numeric=numeric, **kwags)
            if(isinstance(new_data_crtsec, type(None))):
                return None
            new_data = new_data_crtsec.copy() if(new_data.empty) else append(new_data, new_data_crtsec, sort=False)
    else:
        A = np.min(main_col)
        B = np.max(main_col)
        new_data = pd.DataFrame(data.iloc[0]).T if(keephead) else pd.DataFrame()
        
        main_col_ached = np.linspace(A, B, n_anchor)
        anchors = anchors if(not isinstance(anchors, type(None))) else [v for v in main_col_ached if not v in [A, B]]
        for i, anchor in enumerate(anchors):
            if(data[main_col<=anchor].empty):
                return None
            yaaa = data[main_col<=anchor].iloc[-1] #比anchor小的最後一筆真實資料
            ybbb = data[main_col>anchor].iloc[0] ##比anchor大的第一筆真實資料
            a = yaaa.pop(main_coln)
            b = ybbb.pop(main_coln)
            yaaa.index=yaaa.index.map(lambda s:str(s)+'_a')
            ybbb.index=ybbb.index.map(lambda s:str(s)+'_b')
            try:
                ret = {}
                if(not interpolation([a, b], [yaaa, ybbb], x_apply=anchor, method = interpolation_method, ret=ret)):
                    return None
                np_yccc = ret['ans']
            except Exception as e:
                LOGger.exception_process(e, logfile='')
                pd.set_option('display.max_columns', None)
                pd.set_option('display.max_rows', None)
                printer('[%d] anchor=%.2f'%(i, anchor))
                printer('[%d]\n%s'%(i, str(yaaa)))
                printer('[%d]\n%s'%(i, str(ybbb)))
    #            printer('np_yccc.shape:%s'%str(np_yccc.shape))
                pd.set_option('display.max_columns', display_max_columns)
                pd.set_option('display.max_rows', display_max_rows)
                return None
            np_yccc = np.hstack([anchor, np_yccc])
            
            yccc = pd.DataFrame(np_yccc, index=[main_coln]+[v for v in data.columns if v!=main_coln], 
                                columns=[str(yaaa.name) + '..%d'%i]).T
            new_data = yccc.copy() if(new_data.empty) else append(new_data, yccc, sort=False)
        new_data = new_data.append(
                pd.DataFrame(data.iloc[-1]).T, sort=False) if(
                        keeptail) else new_data
    new_data = numeric_as_you_can(new_data) if(numeric) else new_data
    return new_data

#TODO:numeric_as_you_can
def numeric_as_you_can(data, columns=[]):
    exe_columns = columns if(columns) else data.columns
    for coln in exe_columns:
        try:
            data[coln] = data[coln].astype(float)
        except:
            continue
    return data

#TODO:unifrom_xbase
def unifrom_xbase(xdata, *ydata, n_sam=100):
    np_xdata_new, np_ydata_new = None, None
    if(len(ydata)>1):
        np_ydata = []
        for y_ in ydata:
            np_x, np_y = unifrom_xbase(xdata, y_, n_sam=100)
            if(np_xdata_new is None):
                np_xdata_new = dcp(np_x)
            np_ydata.append(np_y)
        np_ydata_new = np_ydata
        return (np_xdata_new, *np_ydata_new)
    np_xdata_new, np_ydata_new = None, None
    np_ydata = np.array(ydata[0])
    if(xdata is None):  xdata = np.arange(np_ydata.shape[0])
    if(isiterable(xdata)):
        np_xdata = np.array(xdata)
        np_xdata_new = np.linspace(np_xdata[0], np_xdata[-1], n_sam)
        if(np_xdata.shape[0]==np_ydata.shape[0] and np_xdata.shape[0]>0):
            interp_function = interp1d(np_xdata, np_ydata)
            np_ydata_new = interp_function(np_xdata_new)
    return np_xdata_new, np_ydata_new

def fft_to_function(y_fft):
    def f(x):
        N = len(y_fft)
        a_0 = y_fft[0].real
        result = a_0 / 2
        for n in range(1, N // 2):
            a_n = y_fft[n].real
            b_n = y_fft[n].imag
            result += a_n * np.cos(2 * np.pi * n * x / N) - b_n * np.sin(2 * np.pi * n * x / N)
        return result
    return f

#TODO:fft_scanerio
def fft_scenario(ydata, xdata=None, sampling_rate=None, sampling_rate_default=1, ret=None, xyuniform=True):
    unitimestep = 1/sampling_rate_default
    if(isiterable(xdata)):
        if(not (np.diff(xdata)<0).any()):
            unitimestep = np.median(np.diff(xdata))
        sampling_rate = sampling_rate if(sampling_rate!=None) else 1/unitimestep
        if(xyuniform):
            xdata, ydata = unifrom_xbase(xdata, ydata, n_sam=sampling_rate)
    if(not isiterable(ydata)):
        return False if(isinstance(ret, dict)) else None
    size = ydata.shape[0]
    unitimestep = 1/sampling_rate
    bfreq = fftfreq(ydata.shape[0], unitimestep)[:size//2]
    y_fft = fft(ydata)
    y_ifft = ifft(y_fft)
    fft_function = fft_to_function(y_fft)
    ifft_function = fft_to_function(y_ifft)
    if(isinstance(ret, dict)):
        spectrum = 2/size * np.abs(y_fft[:size//2])
        ret.update({'y_fft':y_fft, 'x_uni':xdata, 'y_ifft':y_ifft, 
                    'fft_function':fft_function, 'ifft_function':ifft_function,
                    'bfreq':bfreq, 'spectrum':spectrum})
        return True
    return fft_function

#TODO:unduplicate
def unduplicate(data, spec_axis=None, keep='first', alarm_monitor=False):
    root = data[spec_axis].copy() if(spec_axis!=None) else data.copy()
    if(root.duplicated().any().any() if(isiterable(root.duplicated().any())) else root.duplicated().any()):
        LOGger.addlog('偵測到資料重複, n:%d!!!!\nspec axis:%s'%(
            dict(zip(*np.unique(root.duplicated(), return_counts=1)))[True], 
            stamp_process('',spec_axis,'','','',',')), logfile='')
        if(alarm_monitor):
            {}['']
        return data[np.logical_not(root.duplicated(keep=keep))].copy()
    return data

#TODO:unduplicate
def is_duplicated(data, ret=None, unduplicate_method=unduplicate, **kwags):
    ret = ret if(isinstance(data, dict)) else {}
    unduplicated_data = unduplicate_method(data, **kwags)
    ret.upadte({'unduplicated_data':unduplicated_data})
    if(unduplicated_data.shape[0]!=data.shape[0]):
        return True
    return False

def lookup(df, dueTos=None, dueToKeys=None, targetKey=None, default=None, defaultIndex=0):
    if(not isinstance(df, pd.core.frame.DataFrame)):
        return default
    if(isiterable(targetKey)):
        for t in targetKey:
            if(not t in df.columns):    return default
    elif(not targetKey in df.columns):    return default
    dfCond = dcp(df)
    if(isiterable(dueTos)):
        if(not isiterable(dueToKeys)):
            dueToKeys = tuple(np.full(dueToKeys, np.array(dueTos).shape[0]))
        for i,dueTo in enumerate(dueTos):
            dfCond = dfCond[dfCond[dueTo]==dueToKeys[i]]
        return lookup(dfCond, None, None, targetKey, default=default, defaultIndex=defaultIndex)
    if(dueTos is not None):   dfCond = dfCond[dfCond[dueTos]==dueToKeys]
    elif(dueToKeys is not None):    dfCond = dfCond[dfCond.index==dueToKeys]
    if(dfCond[targetKey].shape[0]==0): return default
    if(isinstance(defaultIndex, int)):
        defaultIndex = defaultIndex%dfCond.shape[0]
        return dfCond[targetKey].values[defaultIndex] 
    return dfCond[targetKey].values
    
#step是過去的那1步
def transition(array, step, same_size=1, error_counter={}, **kwags):
    str_array_type = str(type(array))
    if(str_array_type.find('pandas')>-1 and 
       str_array_type.find('Series')>-1):
        if(step==0 or step==None):
            return array
        new_array = (pd.Series([
            array.iloc[0]]*step, 
            index=[str(array.index[0])+'#-%db'%(step-i) 
                   for i in range(step)]).append(array)) if(
                    step>=0) else (
            array.append(pd.Series([
            array.iloc[-1]]*(-step), 
            index=[str(array.index[-1])+'#%db'%(i+1)
                   for i in range(-step)])))
        if(same_size):
            new_array = new_array.iloc[:-step] if(
                    step>=0) else new_array.iloc[-step:]
        return new_array
    if(str_array_type.find('pandas')>-1 and 
       str_array_type.find('DataFrame')>-1):
        kwags.update({'columns':array.columns}) if(len(array.columns)==1) else None
        if(not 'columns' in kwags):
            LOGger.addlog('forget columns??', stamps=[transition.__name__, type(array)], logfile='', 
                          log_counter=error_counter, log_counter_stamp='error', log_counter_ubd=3)
            return None
        LOGger.addlog('', logfile='', log_counter=error_counter, log_counter_stamp='error', reset_log_counter=True)
        if(len(np.array(kwags['columns']).shape)==0):
            if(kwags['columns']=='full'):
                kwags['columns'] = list(array.columns)
            else:
                kwags['columns'] = [kwags['columns']]
        new_array = pd.DataFrame()
        for i, coln in enumerate(kwags['columns']):
            array_col = array[coln].copy()
            new_array[coln] = transition(
                    array_col, step=step, same_size=same_size).copy()
        if(same_size==1):
            for i, coln in enumerate(list(set(array.columns) - set(kwags['columns']))):
                new_array[coln] = array[coln]
#        return new_array[array.columns]
        return new_array[kwags['columns']]
    if(str_array_type.find('list')>-1):
        old_size = len(array)
        if(old_size==0):
            return []
        if('padding_method' in kwags.keys()):
            padding_inputs = kwags['padding_inputs'] if(
                    'padding_inputs' in kwags.keys()) else range(abs(step))
            padding_param = {'a_0':array[0],'a_-1':array[-1]}
            padding_param.update(kwags['padding_param'] if(
                    'padding_param' in kwags.keys()) else {})
            padding_outputs = [kwags['padding_method'](
                    v, padding_param=padding_param) for v in padding_inputs]
            array = (padding_outputs + array) if(step>=0) else (array + padding_outputs)
        else:
            v = (array[0] if(step>=0) else array[-1]) if(
                    not 'default_value' in kwags.keys()) else kwags['default_value']
            array = (([v]*step) + array) if(step>=0) else (array + ([v]*(-step)))
        if(same_size):
            array = array[:old_size] if(step>=0) else (array[-old_size:])
            return array
    if(str_array_type.find('array')>-1):
        array1 = [v for v in array]
        return np.array(transition(array1, step=step, same_size=same_size))
    if(str_array_type.find('tuple')>-1):
        array1 = [v for v in array]
        return tuple(transition(array1, step=step, same_size=same_size))

def suddenly(array, critical_value=0, 
             greater=True, equal_contain=True, step=1,
             mask_size=0, **kwags):
    array_trans = kwags['array_trans'] if(
            'array_trans' in kwags.keys()) else transition(array, step)
    if(np.array([(type(v)==str) for v in array]).any()):
        printer('there is str type elements in array')
        np_array = np.array(array.astype(str))
        np_array_trans = np.array(array_trans.astype(str))
        mask = (np_array!=np_array_trans)
    else:
        printer('>:%s; =:%s'%(greater, equal_contain))
        if(greater and equal_contain):
            mask = (np.array(array_trans)<critical_value)&(np.array(
                       array)>=critical_value)
        if(not greater and equal_contain):
            mask = (np.array(array_trans)>critical_value)&(np.array(
                       array)<=critical_value)
        if(not greater and not equal_contain):
            mask = (np.array(array_trans)>=critical_value)&(np.array(
                       array)<critical_value)
        if(greater and not equal_contain):
            mask = (np.array(array_trans)<=critical_value)&(np.array(
                       array)>critical_value)
    if(mask_size):
        mask = mask[:mask_size]
    return mask
    

def self_array_process(array, tm_array=pd.DataFrame(), step=1, gap=60, critical_value=0,
                       return_extend=(0, 1), return_mask=0, return_package=0, **kwags):
    str_array_type = str(type(array))
    if(str_array_type.find('pandas')==-1):
        if(str_array_type.find("'list'")>-1 or
           str_array_type.find("tuple")>-1 or
           str_array_type.find("ndarray")>-1):
            array = pd.Series(array)
            return_things = self_array_process(array, tm_array = tm_array, step=step, gap=gap,
                               critical_value = critical_value, 
                               return_extend = return_extend,
                               return_mask = return_mask,
                               return_package= return_package, **kwags)
            new_series = return_things[0]
            if(str_array_type.find("'list'")>-1):
                new_array = list(new_series.T)
            if(str_array_type.find("tuple")>-1):
                new_array = tuple(new_series.T)
            if(str_array_type.find("ndarray")>-1):
                new_array = np.array(new_series.T)[0]
            return (new_array, *return_things[1:])
        return None
    method_param = kwags['method_param'] if(
            'method_param' in kwags) else {}
    if(not 'transition_method' in kwags):
        def transition_method(array, **kwags):
            return transition(array, step=step)
    else:
        kwag_transition_method = kwags['transition_method']
        def transition_method(array, **kwags):
            return kwag_transition_method(array, method_param=method_param)
    if(not 'method' in kwags):
        def method(array, transition_method, **kwags):
            array_trans = transition_method(array)
            return np.array(array) - np.array(
                    array_trans)>gap
    else:
        if(type(kwags['method'])==str):
            if(kwags['method'].find('suddenly')>-1):
                greater = (kwags['method'].find('>')>-1 or 
                           kwags['method'].find('greater')>-1)
                equal_contain = (kwags['method'].find('=')>-1 or 
                                 kwags['method'].find('euqal contain')>-1)
                def method(array,**kwags):
                    return suddenly(array, critical_value=critical_value, 
                             greater = greater, equal_contain=equal_contain)
            elif(kwags['method'].find('abs_err')>-1):
                def method(array, transition_method, **kwags):
                    array_trans = transition_method(array)
                    return np.abs(np.array(array) - np.array(
                            array_trans))>gap
            
            elif(kwags['method']=='>='):
                def method(array, transition_method, **kwags):
                    array_trans = transition_method(array)
                    return np.array(array) - np.array(
                            array_trans)>=gap
            elif(kwags['method']=='<='):
                def method(array, transition_method, **kwags):
                    array_trans = transition_method(array)
                    return np.array(array) - np.array(
                            array_trans)<=gap
            elif(kwags['method']=='<'):
                def method(array, transition_method, **kwags):
                    array_trans = transition_method(array)
                    return np.array(array) - np.array(
                            array_trans)<gap
            elif(kwags['method']=='>'):
                def method(array, transition_method, **kwags):
                    array_trans = transition_method(array)
                    return np.array(array) - np.array(
                            array_trans)>gap
        else:
            kwags_method = kwags['method']
            def method(array, **kwags):
                array_trans = transition_method(array)
                method_param['array_trans'] = array_trans
                return kwags_method(
                        array,
                        method_param = method_param)
    np_critical_mask = method(array, 
                              transition_method=transition_method,
                              method_param=method_param)
    package = {}
    selected_indexes = array[np_critical_mask].index
    if(return_extend!=(0, 1)):
        for index in selected_indexes:
            if(list(array.index).count(index)==1):
                index_loc = list(array.index).index(index)
            if(list(array.index).count(index)>1):
                locs = list(np.where(np.array(
                        array.index)==index)[0].astype(str))
                str_locs = ','.join(locs)
                printer('[%s]index duplicated!! locs:\n%s'%(
                        index, str_locs[:50]+'...'))
            package[index] = array.iloc[
                    (index_loc+return_extend[0]):(
                            index_loc+return_extend[1])].copy()
    _return = [array[np_critical_mask]] if(
        tm_array.empty) else [tm_array[np_critical_mask]]
    _return = ([array[np_critical_mask]] if(
            tm_array.empty) else [tm_array[np_critical_mask]]) if(
            str_array_type.find('pandas')>-1) else []
    _return += [np_critical_mask] if(
            return_mask and str_array_type.find('pandas')>-1) else []
    _return += [package] if(return_extend!=(0, 1)) else []
    _return = tuple(_return)
    return _return

def uniqueByIndex(array, **kwags):
    kwags['return_index'] = True
    unique, indices = np.unique(array, **kwags)
    unique_in_order = unique[np.argsort(indices)]
    return unique_in_order

def deNoise(data, default_value=np.nan, replaced_values=None, replace_mapping=None, **kwags):
    replaced_values = replaced_values if(isinstance(replaced_values, list)) else m_noise_values
    print(replaced_values)
    for replaced_value in replaced_values:
        if(replaced_value in [np.nan]): 
            data = np.where(np.vectorize(lambda x:parse(x).lower().find('nan')>-1)(data), default_value, data)
        else:
            data = np.where(data==replaced_value, default_value, data)
    return data

def enNoise(data, default_value=np.nan, reveal_value=None, **kwags):
    if(default_value in [np.nan]):
        data = np.where(np.vectorize(lambda x:parse(x).lower().find('nan')>-1)(data), reveal_value, data)
    else:   
        data = np.where(data==default_value, reveal_value, data)
    return data

def data_centrelize(data, method=None, is_restore_nonnumeric_data=True, **kwags):
    addlog = kwags.get('addlog', LOGger.addloger(logfile=''))
    columns_nonnumeric = [v for v in data.columns if data[v].map(lambda s:isinstance(astype(s), type(None))).any()]
    addlog('columns_nonnumeric:%s'%(','.join(columns_nonnumeric[:20]))) if(len(columns_nonnumeric)>0) else None
    columns_numeric = [v for v in data.columns if not v in columns_nonnumeric]
    data_numeric = data[columns_numeric].astype(float)
    if(method!=None):
        data_centre = method(data_numeric)
    else:
        #取眾數
        if(len(columns_nonnumeric)>0):
            data_nonnumeric = data[columns_nonnumeric].astype(str)
            hlbc = HLBC_fit(data_nonnumeric)
            data_nonnumeric_lbced = data_nonnumeric.copy()
            for i,hlbced in enumerate(hlbc.transform(data_nonnumeric)):
                data_nonnumeric_lbced[data_nonnumeric.columns[i]] = hlbced
            data_lbced = data_numeric.join(data_nonnumeric_lbced, sort=False).copy()
        u,c = np.unique(data_lbced, axis=0, return_counts=1) if(len(columns_nonnumeric)>0) else np.unique(data, axis=0, return_counts=1)
        data_centre = u[np.argmax(c)]
    if(is_restore_nonnumeric_data and len(columns_nonnumeric)>0):
        data_centre = np.hstack([data_centre[:-len(columns_nonnumeric)], 
                   hlbc.inverse_transform(data_centre[-len(columns_nonnumeric):]).reshape(-1)])
        addlog('data_centre:\n%s'%str(data_centre))
    return data_centre

def tablelog_process(a_data, stamps=None, exp_fd='IO_tables', a_index=None, **kwags):
    a_index = dt.now().strftime('%H-%M-%S-%f') if(a_index==None) else a_index
    a_index = LOGger.mylist([a_index]).get_all()[0]
    stamps = stamps if(isinstance(stamps, list)) else []
    exp_fd = exp_fd if(isinstance(exp_fd, str)) else 'IO_tables'
    topic_stamp = os.path.basename(exp_fd)
    file = os.path.join(exp_fd,'%s.pkl'%stamp_process('',[topic_stamp]+stamps,'','','','_',for_file=1))
    df_history = pd.DataFrame()
    if(os.path.exists(file)):
        try:
            df_history = pd.read_pickle(file)
        except Exception as e:
            LOGger.exception_process(e, logfile=kwags.get('logfile', os.path.join('log','log_%t.txt')))
            return False
    pd_a_data = None
    if(isinstance(a_data, dict)):
        pd_a_data = pd.DataFrame.from_dict({a_index:a_data}).T
    elif(isinstance(a_data, pd.core.series.Series)):
        pd_a_data = pd.DataFrame(a_data)
    elif(not isiterable(a_data)):
        printer('illegal data type:%s'%LOGger.type_string(a_data))
        return False
    if(isinstance(a_data, dict)):
        pd_a_data = pd.DataFrame.from_dict({a_index:a_data}).T
    else:
        np_a_data = np.array(a_data)
        pd_a_data = pd.DataFrame(np_a_data.reshape(1,-1), index=[a_index], 
                                 columns=getattr(pd_a_data, 'columns', range(np_a_data.shape[1])))
    if(not df_history.empty):
        for col in pd_a_data:
            if(not col in df_history):
                df_history[col] = None
        df_history = pd_a_data[df_history.columns].append(df_history, sort=False)
    else:
        df_history = pd_a_data.copy()
    df_history = df_history.applymap(lambda s: astype(s,default_method=lambda x:x))
    LOGger.CreateFile(file, lambda f:df_history.to_pickle(f))
    return True

#TODO:ismonotonic
def ismonotonic(array, direction=0, strict=True, **kwags):
    ret=None
    str_type = str(type(array))
    if(str_type.find('ndarray')>-1):
        if(direction==0):#是否單調
            ret = (np.prod(np.sign(np.diff(np.array(array))))>0) if(
                    strict) else np.prod(np.sign(np.diff(np.array(array))))>=0
        if(direction!=0):
            array *= direction
            ret = np.logical_not(np.sign(np.diff(np.array(array)))>0).any() if(
                    strict) else np.logical_not(np.sign(np.diff(np.array(array)))>=0).any()
            output = kwags.get('output',{})
            if(output):
                output['new_array'] = monotonization(array, increasing=direction>0, 
                                                     reverse=kwags.get('reverse',False), 
                                                     self_array_process_method=kwags.get('self_array_process_method','<'))
    if(str_type.find('list')>-1 or str_type.find('tuple')>-1 or 
       str_type.find('pandas.core')>-1):
        np_array = np.array(array)
        ret = ismonotonic(np_array, direction=direction, strict=strict, **kwags)
    return ret

def monotonization(data, increasing=True, reverse=False, 
                   self_array_process_method='<', **kwags):
    if('abort_condition' in kwags):
        abort_condition = kwags['abort_condition']
        if(abort_condition(data)):
            return None
    str_array_type = str(type(data))
    if(str_array_type.find('pandas')==-1):
        if(str_array_type.find("'list'")>-1 or
           str_array_type.find("tuple")>-1 or
           str_array_type.find("ndarray")>-1):
            data = pd.DataFrame(data)
            new_series = monotonization(
                    data, increasing=increasing, reverse=reverse, **kwags)
            if(str_array_type.find("'list'")>-1):
                new_data = list(new_series.T)
            if(str_array_type.find("tuple")>-1):
                new_data = tuple(new_series.T)
            if(str_array_type.find("ndarray")>-1):
                new_data = np.array(new_series.T)[0]
        else:
            return None
    else:
        new_data = data.iloc[::-1].copy() if(reverse) else data.copy()
        data_index = list(data.index)
        crt_timings = self_array_process(
                new_data, method=self_array_process_method, gap=0, step=1)[0]
        while(not crt_timings.empty):
#                    {}['stop']
            index = crt_timings.index[0]
            loc = data_index.index(index)
            if(loc<2):
                return None
            remain_frame = new_data[:loc].copy()
            remain_frame = pd.Series(
                    [0]*len(remain_frame.index), 
                    index=remain_frame.index)
            
            accm_frame = new_data[loc:].copy()
            value = new_data[loc-1] - new_data[loc-2]
            accm_frame = pd.Series(
                    [value]*len(accm_frame.index), 
                    index=accm_frame.index)
            addings = np.array(
                    remain_frame.append(accm_frame))
#                        {}['stop']
            new_data = new_data + addings 
            crt_timings.pop(index)
        new_data = new_data.iloc[::-1].copy() if(reverse) else new_data
    return new_data

def tensorizing_data(X_data, y_data, seq_sz=1, key='', lbc=None):
    if(type(y_data)==dict):
        np_y_data_box={}
        for key in y_data.keys():
            np_X_data, np_y_data = tensorizing_data(
                    X_data, y_data[key], seq_sz=seq_sz, key=key, lbc=lbc)
            np_y_data_box[key] = np_y_data
        return np_X_data, np_y_data_box
    else:
        if(seq_sz>1):
            np_X_data = bcd.cast_to_floatx([[[(X_data[0][i])[k][j] 
            for k in X_data[0][i].columns] 
            for j in X_data[0][i].index] 
            for i in X_data.index])
            np_y_data = bcd.cast_to_floatx(np.array(y_data))
        else:
            y_data_execute = lbc.transform(y_data) if(
                    key=='clf' and lbc!=None) else y_data
            np_X_data = bcd.cast_to_floatx(np.array(X_data))
            np_y_data = bcd.cast_to_floatx(np.array(y_data_execute))
        return np_X_data, np_y_data

def inverse_normalization(data, scr, columns='full', is_seq=False):
    if(scr):
        if(is_seq):
            columns = (data[0].iloc[0]).columns if(columns=='full') else columns
            data_exp = pd.DataFrame()
            for seq_index in data.index:
                data_execute = data[0][seq_index][columns].copy()
                data_wait = data[0][seq_index].drop(columns, axis=1)
                np_data_recover = scr.inverse_transform(data_execute)
                data_recover = pd.DataFrame(
                        np_data_recover,  columns=columns, index=data[0][seq_index].index)
                data_exp = pd.DataFrame(
                        [data_recover.join(data_wait)], columns=[0], 
                        index=[seq_index]) if(
                                data_exp.empty) else data_exp.append(
                                        pd.DataFrame(
                                        [data_recover.join(data_wait)], 
                                        columns=[0], 
                                        index=[seq_index]), 
                                        sort=False)
        else:
            columns = data.columns if(columns=='full') else columns
            data_execute = data[columns].copy()
            data_wait = data.drop(columns, axis=1)
            np_data_recover = scr.inverse_transform(data_execute)
            data_recover = pd.DataFrame(
                    np_data_recover,  columns=columns, index=data.index)
            data_exp = data_recover.join(data_wait)
        return data_exp
    else:
        return data
#TODO:normalization
def normalization(df0, cell_size=0, columns_separate_nanull = True,
                  file = '', lbfile='', scr=None, 
                  scrsrc=None, columns_name_file='', columns_name_json='', showlog=True, **kwags):
    if(cell_size):
        df0 = inverse_packaging(df0, cell_size)
    columns, other_cols = getattr(df0,'columns',np.arange(df0.shape[1])), []
    if(columns_name_file):
        columns, other_cols = get_column_names_from_txt(columns_name_file), list(
                                    set(df0.columns) - set(get_column_names_from_txt(columns_name_file)))
    if(columns_name_json):
        columns, other_cols = LOGger.load_json(columns_name_json), list(
                                    set(df0.columns) - set(LOGger.load_json(columns_name_json)))
    df0 = df0[columns] if(isinstance(df0, pd.core.frame.DataFrame)) else df0[:,columns]
    if(columns_separate_nanull):
        columns, other_cols = separate_nanull_columns(df0, return_coln=1)
    df0_unnrm = df0[other_cols] if(isinstance(df0, pd.core.frame.DataFrame)) else df0[:,other_cols]
    df0_nrm = df0[columns] if(isinstance(df0, pd.core.frame.DataFrame)) else df0[:,columns]
    if(showlog):
        printer('normalize colums:%s\n'%str(columns)+(str(getattr(df0_nrm, 'values', df0_nrm)[:5])))
    if(scr==None):
        if(scrsrc):
            scr =  scrsrc if(str(type(scrsrc)).find('Scaler')>-1) else joblib.load(scrsrc)
        else:
            SCALER = kwags.get('SCALER', Stdscr)
            scr = SCALER().fit(df0_nrm)
            scr.header = columns
            printer('set header:%s'%(','.join(list(map(str, columns)))))
    df0_nrm = pd.DataFrame(scr.transform(df0_nrm), columns=columns, index=df0.index) if(
        isinstance(df0, pd.core.frame.DataFrame)) else scr.transform(df0_nrm)
    df0 = df0_nrm.join(df0_unnrm) if(
        isinstance(df0, pd.core.frame.DataFrame)) else np.hstack([df0_nrm, df0_unnrm])
    if(showlog):
        printer('after normalizaztion.....\n'+(
                str(getattr(df0_nrm, 'values', df0_nrm)[:5])))
    if(cell_size):
        df0 = packaging(df0, cell_size)
    scr.headers = columns
    if(file):
        printer('儲存常態器:%s'%file)
        joblib.dump(scr, file)
        if(not lbfile):
            lbfile = file.replace(file[file.find('.'):],'.json')
        LOGger.save_json(list(tuple(columns)), lbfile)
    return df0, scr

def correl(x, y, n_sigma=3, n_sigmas=None, ret=None):
    try:
        ret = ret if(isinstance(ret, dict)) else {}
        ret['count'] = np.array(y).shape[0]
        ret['cor'] = np.corrcoef(x, y)[1,0]
        # ret['W'], ret['p'] = stats.kstest(data, 'norm', args=(ret['mean'], ret['std']))#shapiro(data)
        # 计算常態分布的临界值
        # if(astype(n_sigma)!=None):
        #     if(n_sigma>0):
        #         ret['norm_lower'] = ret.get('norm_lower', ret['mean'] - n_sigma * ret['std'])
        #         ret['norm_upper'] = ret.get('norm_upper', ret['mean'] + n_sigma * ret['std'])
        #         ret['E'] = np.sum((np.array(data) < ret['norm_lower'])|(np.array(data) > ret['norm_upper']))/np.array(data).shape[0]
        # n_sigmas = n_sigmas if(isinstance(n_sigmas, list)) else []
        # for n_sig in n_sigmas:
        #     if(astype(n_sig)!=None):
        #         if(n_sig>0):
        #             ret['norm_lower_%s'%parse(n_sig)] = ret['mean'] - n_sig * ret['std']
        #             ret['norm_upper_%s'%parse(n_sig)] = ret['mean'] + n_sig * ret['std']
    except Exception as e:
        LOGger.exception_process(e, logfile='')
        return False
    return True
#TODO:reg_metrics
reg_std_metrics = {}
reg_std_metrics['mse'] = lambda f, p, cell_size=None: mean_squared_error_with_exception_process(f, p, cell_size=cell_size)
# reg_std_metrics['mse'] = lambda f, p: np.mean((f - p)**2)
reg_std_metrics['rmse'] = lambda f, p, cell_size=None: (reg_std_metrics['mse'](f, p, cell_size=cell_size))**(1/2)
reg_std_metrics['r2'] = lambda f, p, cell_size=None: r2_score_with_exception_process(f, p, cell_size=cell_size)
def mean_squared_error_with_exception_process(f, p, cell_size=None):
    try:
        cell_size = cell_size if(isiterable(cell_size)) else np.array(f).shape[1:]
        if(len(cell_size)>0):
            flat_shape = np.prod(cell_size)
            f = np.array(f).reshape((-1, flat_shape))
            p = np.array(p).reshape((-1, flat_shape))
        return skm.mean_squared_error(f, p)
    except Exception as e:
        LOGger.exception_process(e, '', stamps=['mean_squared_error'])
        LOGger.show_vector(np.array(f)[np.array(tuple(map(lambda v: isinstance(v,str), f)))], stamps=['fact str elements'], logfile='')
        LOGger.show_vector(np.where(np.abs(np.array(f).reshape(-1))==np.inf), stamps=['fact inf elements locs'], logfile='')
        LOGger.show_vector(np.where(np.array(f).reshape(-1)==np.nan), stamps=['fact nan elements locs'], logfile='')
        LOGger.show_vector(np.array(p)[np.array(tuple(map(lambda v: isinstance(v,str), p)))], stamps=['pred str elements'], logfile='')
        LOGger.show_vector(np.where(np.abs(np.array(p).reshape(-1))==np.inf), stamps=['pred inf elements locs'], logfile='')
        LOGger.show_vector(np.where(np.array(p).reshape(-1)==np.nan), stamps=['pred nan elements locs'], logfile='')
        return np.nan #-1
def r2_score_with_exception_process(f, p, cell_size=None):
    try:
        cell_size = cell_size if(isiterable(cell_size)) else np.array(f).shape[1:]
        if(len(cell_size)>0):
            flat_shape = np.prod(cell_size)
            f = np.array(f).reshape((-1, flat_shape))
            p = np.array(p).reshape((-1, flat_shape))
        return skm.r2_score(f, p)
    except Exception as e:
        LOGger.exception_process(e, '', stamps=['r2_score'])
        LOGger.show_vector(np.array(f)[np.array(tuple(map(lambda v: isinstance(v,str), f)))], stamps=['fact str elements'], logfile='')
        LOGger.show_vector(np.where(np.abs(np.array(f))==np.inf), stamps=['fact inf elements locs'], logfile='')
        LOGger.show_vector(np.where(np.array(f)==np.nan), stamps=['fact nan elements locs'], logfile='')
        LOGger.show_vector(np.array(p)[np.array(tuple(map(lambda v: isinstance(v,str), p)))], stamps=['pred str elements'], logfile='')
        LOGger.show_vector(np.where(np.abs(np.array(p))==np.inf), stamps=['pred inf elements locs'], logfile='')
        LOGger.show_vector(np.where(np.array(p)==np.nan), stamps=['pred nan elements locs'], logfile='')
        return np.nan #-np.inf
    
#TODO:OKratio
def OKratio(f, p, equal_included=True, pdt_side=0, tol=None):
    tol = (np.std(f) if(np.array(f).shape[0]>2) else 1) if(tol==None) else tol
    if(not np.isnan(astype(tol, d_type=float, default=np.nan))):
        if(pdt_side>0):
            return np.sum(np.ones(f.shape[0])[
                            p - f >= -tol])/f.shape[0] if(
                            equal_included) else np.sum(np.ones(f.shape[0])[
                            p - f > -tol])/f.shape[0]
        elif(pdt_side<0):
            return np.sum(np.ones(f.shape[0])[
                            p - f <= tol])/f.shape[0] if(
                            equal_included) else np.sum(np.ones(f.shape[0])[
                            p - f < tol])/f.shape[0]
        elif(pdt_side==0):
            return np.sum(np.ones(f.shape[0])[
                    np.abs(p - f) <= tol])/f.shape[0] if(
                            equal_included) else np.sum(np.ones(f.shape[0])[
                    np.abs(p - f) < tol])/f.shape[0]
        else:
            {}['error pdt side:%s'%pdt_side]
    else:
        #for example...... tol = lambda a,b: np.abs(a-b)<=a*0.03
        return np.sum(np.ones(f.shape[0])[tol(f, p)])/f.shape[0]

def sumup_regression_norms(y_real, y_pred, tol=-1, names=(tuple(reg_std_metrics.keys())+('OKR',)), 
                           xdim=0, equal_included=False, pdt_side=0,
                           additional_norms={}, additional_kwagses={}, **kwags):
    y_pred = np.array(y_pred)
    y_real = np.array(y_real)
    tol = np.std(y_real) if(tol==-1 or tol==None) else tol
    ret = {k: reg_std_metrics[k](y_real, y_pred) for k in names if k in reg_std_metrics}
    # print(ret)
    if('r2' in names and xdim!=0):
        ret['adr2'] = 1-((1-ret['r2'])*(y_real.shape[0]-1)/(y_real.shape[0]-xdim-2))
    if('OKR' in names):
        ret['OKR'] = OKratio(y_real, y_pred, equal_included=equal_included, pdt_side=pdt_side, tol=tol)
    for k,f in additional_norms.items():
        additional_kwags = additional_kwagses.get(k, {})
        ret[k] = f(y_real, y_pred, **additional_kwags)
    ret['tol'] = tol
    ret['size'] = np.array(y_real).shape[0]
    return ret

def MannwhitneyuTest(sample1 = np.random.normal(loc=50, scale=5, size=30), sample2 = np.random.normal(loc=53, scale=5, size=30), 
                     alternative='two-sided', threshold=None, ret=None, **kwags):
    """
    檢定兩獨立樣本的秩和差異。U越大差異越大、p越小越顯著

    Parameters
    ----------
    sample1 : TYPE, optional
        DESCRIPTION. The default is np.random.normal(loc=50, scale=5, size=30).
    sample2 : TYPE, optional
        DESCRIPTION. The default is np.random.normal(loc=53, scale=5, size=30).
    alternative : TYPE, optional
        DESCRIPTION. The default is 'two-sided'.
    ret : TYPE, optional
        DESCRIPTION. The default is None.
    **kwags : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    addlog = LOGger.execute('addlog', kwags, default=LOGger.addloger(logfile=''), not_found_alarm=False)
    ret = ret if(isinstance(ret, dict)) else {}
    
    u_stat, p_value = stats.mannwhitneyu(sample1, sample2, alternative=alternative)
    threshold = threshold if(not np.isnan(astype(threshold, d_type=float, default=np.nan))) else np.std(
        np.hstack([sample1, sample2]))
    addlog(f"U-statistic: {u_stat:.2f}, P-value: {p_value:.4f}")
    ret['u_stat'] = u_stat
    ret['p_value'] = p_value
    ret['threshold'] = threshold
    
    
    
    # 判断置信区间是否在 [-4, 4] 之内
    # if u_stat > threshold:
    #     addlog("threshold:%.4f; 分布差異大"%threshold)
    #     return True
    # else:
    #     addlog("threshold:%.4f; 無法接受分布差異大"%threshold)
    #     return False
    return True

def KSTest(sample1 = np.random.normal(loc=50, scale=5, size=30), sample2 = np.random.normal(loc=53, scale=5, size=30), 
           threshold=0.05, ret=None, **kwags):
    """
    檢定兩獨立樣本的秩和差異。U越大差異越大、p越小越顯著

    Parameters
    ----------
    sample1 : TYPE, optional
        DESCRIPTION. The default is np.random.normal(loc=50, scale=5, size=30).
    sample2 : TYPE, optional
        DESCRIPTION. The default is np.random.normal(loc=53, scale=5, size=30).
    alternative : TYPE, optional
        DESCRIPTION. The default is 'two-sided'.
    ret : TYPE, optional
        DESCRIPTION. The default is None.
    **kwags : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    """
    addlog = LOGger.execute('addlog', kwags, default=LOGger.addloger(logfile=''), not_found_alarm=False)
    ret = ret if(isinstance(ret, dict)) else {}
    
    ks_stat, p_value = stats.ks_2samp(sample1, sample2)
    threshold = threshold if(not np.isnan(astype(threshold, d_type=float, default=np.nan))) else np.std(
        np.hstack([sample1, sample2]))
    addlog(f"KS-statistic: {ks_stat:.2f}, P-value: {p_value:.4f}")
    ret['ks_stat'] = ks_stat
    ret['p_value'] = p_value
    ret['threshold'] = threshold
    
    
    
    if p_value < threshold:
        addlog("threshold:%.4f; 分布差異大"%threshold)
        return True
    else:
        addlog("threshold:%.4f; 無法接受分布差異大"%threshold)
        return False
    # if u_stat > threshold:
    #     addlog("threshold:%.4f; 分布差異大"%threshold)
    #     return True
    # else:
    #     addlog("threshold:%.4f; 無法接受分布差異大"%threshold)
    #     return False

def NormalMeanDifferenceTest(sample1 = np.random.normal(loc=50, scale=5, size=30), sample2 = np.random.normal(loc=53, scale=5, size=30), 
                             MeanGap=4, ret=None, **kwags):
    addlog = LOGger.execute('addlog', kwags, default=LOGger.addloger(logfile=''), not_found_alarm=False)
    ret = ret if(isinstance(ret, dict)) else {}
    # 计算样本均值和标准差
    mean1 = np.mean(sample1)
    mean2 = np.mean(sample2)
    std1 = np.std(sample1, ddof=1)
    std2 = np.std(sample2, ddof=1)
    
    addlog(f"Sample 1 mean: {mean1:.2f}, std: {std1:.2f}")
    addlog(f"Sample 2 mean: {mean2:.2f}, std: {std2:.2f}")
    
    # 执行独立样本 t 检验
    t_stat, p_value = stats.ttest_ind(sample1, sample2, equal_var=False)
    addlog(f"T-statistic: {t_stat:.2f}, P-value: {p_value:.4f}")
    
    # 计算95%置信区间
    alpha = 0.05
    n1 = len(sample1)
    n2 = len(sample2)
    se1 = std1 / np.sqrt(n1)
    se2 = std2 / np.sqrt(n2)
    se_diff = np.sqrt(se1**2 + se2**2)
    df = ((se1**2 + se2**2)**2) / ((se1**2 / (n1 - 1)) + (se2**2 / (n2 - 1)))
    ci_low, ci_high = stats.norm.interval(1 - alpha, loc=(mean1 - mean2), scale=se_diff)
    
    addlog(f"95% Confidence interval of the difference: [{ci_low:.2f}, {ci_high:.2f}]")
    ret['ci_low'] = ci_low
    ret['ci_high'] = ci_high
    ret['t_stat'] = t_stat
    ret['p_value'] = p_value
    # 判断置信区间是否在 [-4, 4] 之内
    if ci_low > -MeanGap and ci_high < MeanGap:
        addlog("The difference in means is within ±4g.")
        return True
    else:
        addlog("The difference in means is not within ±4g.")
        return False

def normfit(data, n_sigma=3, n_sigmas=None, ret=None, annil=False):
    try:
        ret.clear() if(annil) else None
        ret = ret if(isinstance(ret, dict)) else {}
        ret['count'] = np.array(data).shape[0]
        ret['mean'], ret['std'] = stats.norm.fit(data)
        ret['W'], ret['p'] = stats.kstest(data, 'norm', args=(ret['mean'], ret['std']))#shapiro(data)
        # 计算常態分布的临界值
        if(astype(n_sigma)!=None):
            if(n_sigma>0):
                ret['norm_lower'] = ret.get('norm_lower', ret['mean'] - n_sigma * ret['std'])
                ret['norm_upper'] = ret.get('norm_upper', ret['mean'] + n_sigma * ret['std'])
                ret['E'] = np.sum((np.array(data) < ret['norm_lower'])|(np.array(data) > ret['norm_upper']))/np.array(data).shape[0]
        n_sigmas = n_sigmas if(isinstance(n_sigmas, list)) else []
        for n_sig in n_sigmas:
            if(astype(n_sig)!=None):
                if(n_sig>0):
                    ret['norm_lower_%s'%parse(n_sig)] = ret['mean'] - n_sig * ret['std']
                    ret['norm_upper_%s'%parse(n_sig)] = ret['mean'] + n_sig * ret['std']
    except Exception as e:
        LOGger.exception_process(e, logfile='')
        return False
    return True

def normfit_alpha(data, alpha=0.05, alphas=None, ret=None):
    try:
        ret = ret if(isinstance(ret, dict)) else {}
        ret['mean'], ret['std'] = stats.norm.fit(data)
        ret['W'], ret['p'] = shapiro(data)
        # 计算常態分布的临界值
        if(astype(alpha)!=None):
            if(alpha>0):
                ret['norm_lower'] = ret['mean'] + stats.norm.ppf(alpha / 2) * ret['std']
                ret['norm_upper'] = ret['mean'] + stats.norm.ppf(1 - alpha / 2) * ret['std']
        alphas = alphas if(isinstance(alphas, list)) else []
        for al in alphas:
            ret['norm_lower_%s'%parse(al)] = ret['mean'] + stats.norm.ppf(al / 2) * ret['std']
            ret['norm_upper_%s'%parse(al)] = ret['mean'] + stats.norm.ppf(1 - al / 2) * ret['std']
    except Exception as e:
        LOGger.exception_process(e, logfile='')
        return False
    return True
    
def chi2fit(data, df=2, alpha=0.05, alphas=None, ret=None):
    try:
        ret = ret if(isinstance(ret, dict)) else {}
        ret['count'] = np.array(data).shape[0]
        ret['df'], ret['loc'], ret['scale'] = stats.chi2.fit(data)
        df = ret['df'] if(df==None) else df
        ret['W'], ret['p'] = stats.kstest(data, 'chi2', args=(df,))
        # 计算卡方分布的临界值
        if(astype(alpha)!=None):
            if(alpha<1 and alpha>0):
                ret['alpha'] = alpha
                ret['chi2_upper'] = stats.chi2.ppf(1 - alpha, df=df)
                ret['chi2_upper_adjusted'] = ret['loc'] + stats.chi2.ppf(1 - alpha, df=df) * ret['scale']
        
        alphas = alphas if(isinstance(alphas, list)) else []
        for al in alphas:
            if(astype(al)!=None):
                if(al<1 and al>0):
                    ret['chi2_upper_%s'%parse(al)] = stats.chi2.ppf(1 - al, df=df)
                    ret['chi2_upper_adjusted_%s'%parse(al)] = ret['loc'] + stats.chi2.ppf(1 - al, df=df) * ret['scale']
    except Exception as e:
        LOGger.exception_process(e, logfile='')
        return False
    return True    

def returned(*args, method=None, default=None, **kwags):
    ret_ = kwags.pop('ret', default)
    if(method(*args, ret=ret_, **kwags)):
        return default
    return ret_

def normfited(data, default={}, **kwags):
    return returned(data, method=normfit, default=default, **kwags)

def chi2fited(data, default={}, **kwags):
    return returned(data, method=chi2fit, default=default, **kwags)

def annilate_normfit_extremes(data, n_sigma=3, n_sigmas=None, ret=None, stamps=None, condition=None, refuse_nan=False, sign=0):
    if(data.shape[0]<1):
        return False
    if(not isiterable(condition)):
        condition = np.full(data.shape[0], True)
    if(np.array(condition).shape[0]!=data.shape[0]):
        return False
    ret = ret if(isinstance(ret, dict)) else {}
    stamps = stamps if(isinstance(stamps, list)) else []
    if(len(data.shape)>1):
        if(data.shape[1]>1):
            LOGger.addlog('before annil data size:%d'%data.shape[0], stamps=stamps, logfile='')
            for i,hd in enumerate(getattr(data, 'columns', range(data.shape[1]))):
                np_data = dcp(np.array(data)[:,i])
                if(not annilate_normfit_extremes(np_data, n_sigma=n_sigma, n_sigmas=n_sigmas, ret=ret, condition=condition, stamps=stamps+[hd])):
                    return False
                condition &= ret['condition']
            ret['condition'] = dcp(condition)
            ret['new_data'] = dcp(data[condition])
            return True
    nonnum = np.array(tuple(map((lambda x:np.isnan(astype(x,default=np.nan))), data)))
    data_normfit = data[np.logical_not(nonnum)]
    if(nonnum.any()):
        LOGger.addlog('detect nan', stamps=stamps, logfile='')
        if(not refuse_nan):
            ret['condition'] = dcp(condition)
            ret['new_data'] = dcp(data[condition])
            return True
        else:
            condition &= nonnum
            if(np.array(data_normfit).shape[0]):
                return False
    ret_normfit = {}
    if(not normfit(data_normfit, ret=ret_normfit, n_sigma=n_sigma, n_sigmas=n_sigmas)):
        return False
    ubd = ret_normfit['norm_upper']
    lbd = ret_normfit['norm_lower']
    condition &= ((data<=ubd)|(sign<0))&((data>=lbd)|(sign>0))
    LOGger.addlog('after annil data size[%s,%s]:%d'%(parse(lbd), parse(ubd), data[condition].shape[0]), 
                  stamps=stamps, logfile='')
    if(data.shape[0]<1):
        return False
    ret['condition'] = dcp(condition)
    ret['new_data'] = dcp(data[condition])
    return True
    

def renew_version(version, sep='-', new_vname=None, loc_signed=None, **kwags):
    if(version.find('v')>-1):
        v_stg_loc = version.find('v')
        version_1 = version.replace('v','')
        out_version_1 = renew_version(version_1, sep=sep, 
                              new_vname=new_vname, 
                              loc_signed=loc_signed, **kwags)
        list_out_version_1 = list(out_version_1)
        list_out_version_1.insert(v_stg_loc, 'v')
        out_version = ''.join(list_out_version_1)
    else:
        splited_list = version.split('-')
        new_splited_list = list(splited_list)
        for loc, stg in enumerate(splited_list):
            try:
                num = int(stg)
            except:
                continue
            else:
                last_num = num + 1 if(
                        new_vname==None) else str(new_vname)
                last_num_loc = int(loc) if(
                        loc_signed==None) else loc_signed
        new_splited_list.pop(last_num_loc)
        new_splited_list.insert(last_num_loc, '%s'%last_num)
        out_version = '-'.join(new_splited_list)
    return out_version

def sigmoid(x, center=0, k=1):
    return 1 / (1 + np.exp(-k * (x - center)))

def diagonal_semibelt_matrix(b=3,channel_dim=3,axis=1,pattern=None):
    if(axis<0):
        return np.transpose(diagonal_semibelt_matrix(b,channel_dim,axis=1,pattern=pattern))
    if(not isiterable(pattern)):
        zero_dim = (channel_dim-1)*b
        if(zero_dim<0):
            return None
        init_roll = [1]*b + [0]*zero_dim
        shift=b
    else:
        pattern = mylist(pattern).get_all()
        b = np.array(pattern).shape[0]
        zero_dim = (channel_dim-1)*b
        init_roll = pattern + [0]*zero_dim
        shift=len(pattern)
    return special_matrix_rolling_array(init_roll, shift=shift)

#TODO:diagonal_belt_matrix
def diagonal_belt_matrix(b=1, dim=5):
    roller_dim = dim - b - 1#(dim+b-1)+1-2*b-1 
    # if(roller_dim<0):
    #     return None
    init_roll = [1]*(2*b+1) + [0]*roller_dim
    return special_matrix_rolling_array(init_roll)[:-b,b:]

#TODO:frequenize
def frequenize(ay, b=30):
    M = diagonal_belt_matrix(int(b//2), np.array(ay).shape[0])
    return np.matmul(M, ay)

#TODO:frequency_densize
def frequency_densize(ay, b=30):
    #把每一筆資料當中心點，往左右開展b//2格，加總起來取平均
    M = diagonal_belt_matrix(int(b//2), np.array(ay).shape[0])
    return np.matmul(M, ay)/np.matmul(M, np.ones(shape=np.array(ay).shape[0]))

def rendering_ones(ay, b=1):
    dim = ay.shape[0]
    M = diagonal_belt_matrix(dim=dim, b=1)
    return np.sign(np.sign(np.matmul(M, ay)))

def dimension_monotonized(array, group_size=2):
    if(array.shape[1]==0):
        return np.full(array.shape, np.nan)
    elif(array.shape[1]%group_size!=0):
        return np.full(array.shape, np.nan)
    elif(array.shape[1]>group_size):
        dtype=array.dtype
        outputs = np.ones(shape=(array.shape[0],0))
        for i in range(0,array.shape[1],group_size):
            array_plane = array[:,i:i+group_size]
            array_plane_ed = dimension_monotonized(array_plane)
            outputs = np.hstack([outputs, array_plane_ed])
        return np.array(outputs, dtype=dtype)
    else:
        array = np.sort(array, axis=1)
        return array

def intersection_area(y_true, y_pred, egn=np):
    if(y_true.shape!=y_pred.shape):
        return None
    elif(y_true.shape[1]==0):
        return None
    elif(y_true.shape[1]%2!=0):
        return egn.full(y_true.shape, egn.nan)
    elif(y_true.shape[1]>2):
        area = 1
        for i in range(0,y_true.shape[1],2):
            y_true_plane = y_true[:,i:i+2]
            y_pred_plane = y_pred[:,i:i+2]
            area *= intersection_area(y_true_plane, y_pred_plane, egn=egn)
        return area
    else:
        y_true_egned = egn.sort(y_true, axis=1)
        y_pred_egned = egn.sort(y_pred, axis=1)
        lbd = egn.max(egn.stack([y_true_egned[:,0], y_pred_egned[:,0]], axis=0), axis=0)
        ubd = egn.min(egn.stack([y_true_egned[:,1], y_pred_egned[:,1]], axis=0), axis=0)
        dtype = (ubd - lbd).dtype
        return egn.max(egn.stack([egn.zeros(egn.shape(ubd - lbd), dtype=dtype), ubd - lbd], axis=0), axis=0)
    
def union_area(y_true, y_pred, egn=np):
    if(y_true.shape!=y_pred.shape):
        return None
    elif(y_true.shape[1]==0):
        return None
    elif(y_true.shape[1]%2!=0):
        return egn.full(y_true.shape, egn.nan)
    elif(y_true.shape[1]>2):
        y_true_area, y_pred_area = 1, 1
        for i in range(0,y_true.shape[1],2):
            y_true_area *= egn.abs(y_true[:,i] - y_true[:,i+1])
            y_pred_area *= egn.abs(y_pred[:,i] - y_pred[:,i+1])
        return y_true_area + y_pred_area - intersection_area(y_true, y_pred, egn=egn)
    else:
        y_true_diff = egn.abs(y_true[:,1] - y_true[:,0])
        y_pred_diff = egn.abs(y_pred[:,1] - y_pred[:,0])
        return y_true_diff + y_pred_diff - intersection_area(y_true, y_pred, egn=egn)

def iou(y_true, y_pred, egn=np, get_zeros_method=lambda x:np.zeros(x.shape[0])):
    if(y_true.shape!=y_pred.shape):
        return None
    elif(y_true.shape[1]%2!=0):
        return egn.full(y_true.shape, egn.nan)
    return intersection_area(y_true, y_pred, egn=egn)/union_area(y_true, y_pred, egn=egn)

def predicted_area(y_true, y_pred, egn=np):
    if(y_pred.shape[1]==0):
        return None
    elif(y_pred.shape[1]%2!=0):
        return egn.full(y_pred.shape, egn.nan)
    elif(y_pred.shape[1]>2):
        area = 1
        for i in range(0,y_pred.shape[1],2):
            y_pred_plane = y_pred[:,i:i+2]
            area *= predicted_area(None, y_pred_plane, egn=egn)
        return area
    else:
        return egn.abs(y_pred[:,1] - y_pred[:,0])
    
def iop(y_true, y_pred, egn=np, get_zeros_method=lambda x:np.zeros(x.shape[0])):
    if(y_true.shape!=y_pred.shape):
        return None
    elif(y_true.shape[1]%2!=0):
        return egn.full(y_true.shape, egn.nan)
    return intersection_area(y_true, y_pred, egn=egn)/predicted_area(y_true, y_pred, egn=egn)

#使data 中最靠近new_end的地方成為最後一筆資料
def renew_endpt_of_an_array(data, new_end, main_col=None, main_cols={}, 
                            turn_to_zero=1, **kwags):
    seq_sz = kwags['seq_sz'] if('seq_sz' in kwags) else 0
    str_array_type = str(type(data))
    if(str_array_type.find('ndarray')>-1):
        if(len(data.shape)>1 and data.shape[0]>1):
            new_data = data[:,-seq_sz:]
            if(main_cols!={}):
                new_index={}
                for main_col in main_cols:
                    new_end = main_cols[main_col]
                    new = renew_endpt_of_an_array(
                        new_data[main_col,:], main_col=None, main_cols=[], 
                        new_end=new_end, turn_to_zero=turn_to_zero, **kwags)
                    new_data[main_col,:] = new['new_data']
                    new_index[main_col] = new['new_index']
            else:
                new_data[main_col,:], new_index = tuple(renew_endpt_of_an_array(
                        data[main_col,:], main_col=None, main_cols=[], 
                        new_end=new_end, turn_to_zero=turn_to_zero, **kwags).values())
        else:
            np_abs_dis_buffer = np.abs(data - new_end)
            new_index = np.where(
                    np_abs_dis_buffer==np.min(np_abs_dis_buffer))[0][0]
            new_data = transition(data, 
                     step=data.shape[0]-new_index-1)
            if(turn_to_zero):
                new_data = new_data - new_end
            if(seq_sz):
                new_data = new_data[-seq_sz:]
    if(str_array_type.find('list')>-1 or str_array_type.find('tuple')>-1):
        data = np.array(data)
        new_data, new_index = tuple(renew_endpt_of_an_array(
                data, main_col=None, main_cols=[], 
                new_end=new_end, turn_to_zero=turn_to_zero, **kwags).values())
        new_data = [v for v in new_data]
    if(str_array_type.find('pandas.core.series')>-1):
        new = renew_endpt_of_an_array(
                np.array(data), main_col=None, main_cols=[], 
                new_end=new_end, turn_to_zero=turn_to_zero, **kwags)
        new_data = pd.Series(
                new['new_data'], name = data.name, index=data.index[-seq_sz:])
        new_index = new['new_index']
    if(str_array_type.find('pandas.core.frame')>-1):
        main_col = list(data.columns).index(main_col) if(main_col!=None) else None
        main_cols = {list(data.columns).index(v) for v in main_cols} if(
                main_cols!={}) else {}
        new = renew_endpt_of_an_array(
                np.array(data.T), main_col=main_col, main_cols=main_cols, 
                new_end=new_end, turn_to_zero=turn_to_zero, **kwags)
        new_data = pd.DataFrame(
                new['new_data'], index = data.columns, columns=data.index[-seq_sz:]).T
        new_index = new['new_index']
    return {'new_data':new_data, 'new_index':new_index}

# 指定'tm'欄中，擷取[-36,1)的範圍，公差為6，以最後一筆為基準，作間隔搜取
# interval_picking(data=df_partial, d=6, a=-36, b=1, 
#                  main_col='tm', key_index=df_partial.index[-1])
def interval_picking(data, d=1, ispadding=1, default_value = None,
                     **kwags):
    seq_sym = kwags['seq_sym'] if('seq_sym' in kwags) else 'P'
    a,b = ((kwags['a'] if('a' in kwags) else 0), 
                 (kwags['b'] if('b' in kwags) else len(data)))
    str_array_type = str(type(data))
    if(str_array_type.find("'list'")>-1):
        if(ispadding):
            kwags['key_index'] = kwags['key_index'] if(
                    'key_index' in kwags) else 0
            data = padding(data, default_value=default_value, 
                           seq_sym=seq_sym, **kwags)
        new_data = [data[i] for i in range(a,b,d)]
    if(str_array_type.find('numpy.ndarray')>-1):
        data = [v for v in data]
        new_data = interval_picking(data, d=d,
                                    default_value=default_value,
                                    **kwags)
        new_data = np.array(new_data)
    if(str_array_type.find('pandas.core.series')>-1):
        index_default_values = kwags['index_default_values'] if(
                'index_default_values' in kwags.keys()) else '-%s'%seq_sym
        values = np.array(data.values)
        indexes = np.array(data.index)
        new_values = interval_picking(values, d=d,
                                    default_value=default_value,
                                    **kwags)
        new_indexes = interval_picking(indexes, d=d,
                                    default_value=index_default_values,
                                    **kwags)
        new_data = pd.Series(new_values, index = new_indexes)
    if(str_array_type.find('pandas.core.frame')>-1):
        if(('main_col' in kwags) and ('key_index' in kwags)):
            main_col = kwags['main_col']
            key_index = kwags['key_index']
            key_value = data[main_col][key_index]
            data[main_col] = data[main_col].map(lambda s:s-key_value)
            data_b = data[(np.abs(data[main_col])<=b)&
                          a<=(np.abs(data[main_col]))]
            select_mask = np.zeros(len(data_b.index))>0
            for i in range(a,b,d):
                np_abs_data_b_tm = (np.abs(data_b[main_col] - i))
                select_mask |= (np_abs_data_b_tm == np.min(np_abs_data_b_tm))
            new_data = data_b[select_mask]
        else:
            new_data = pd.DataFrame()
            data_columns = kwags['columns'] if('columns' in kwags) else data.columns
            for col in data_columns:
                data_col = data[col].copy()
                excute_default_value = (kwags['default_dict'][col] if(
                        col in kwags['default_dict'].keys()) else default_value) if(
                        'default_dict' in kwags) else default_value
                new_data_col = interval_picking(data_col, d=d,
                                                default_value=excute_default_value,
                                                **kwags)
                new_data[col] = new_data_col.copy()
    return new_data
#TODO:str_multi_contains
def str_multi_contains(series, contents):
    contents = [str(c).replace('[','\[') for c in contents]
    mask = np.zeros(len(series))>0
    for content in contents:
        mask |= series.str.contains(content)
    return mask

def index_ret(s, L, ret=None, start=-1):
    try:
        ret = {} if(ret==None) else ret
        ret.update({s:list(tuple(L)).index(s, ret.get(s, start) + 1)})
    except:
        if(s in L):
            locs = np.where((np.array(L).reshape(-1)==s))
            print('init loc %s > all locs: ...%s'%(ret.get(s, np.nan), str(tuple(locs[-10:]))))
            return np.nan
        else:
            return None
    return dcp(ret[s])

#TODO:indexing
#repeat_index允許重複的index出現，也就是如果有相同元素出現在不同位置，都只會選第一個出現的位置
def indexing(referentor, reference=None, repeat_index=False):
    ret={}
    core_operator = (lambda s:list(tuple(reference)).index(s)) if(
        repeat_index) else (lambda s:index_ret(s, reference, ret=ret))
    return np.array(tuple(map(core_operator, referentor)))

def adjust_stratify(stgy, default='*', showcount = 5, showdetail=False, class_max=0):
    printer('分類依照的性質種類(個數:%d)如下:'%(len(list(set(stgy)))))
    printer(','.join([str(s) for s in list(set(stgy))[:showcount]])+(
            '' if(len(list(set(stgy)))<=showcount) else ', ...'))
    isolatedtype = []
    counts={default:0}
    for t in set(stgy):
        if(stgy.count(t)<2):
            isolatedtype.append(t)
            counts[default] += 1
            if(showdetail):
                printer('偵測到孤立點:%s'%(str(t)))
        else:
            counts[t] = int(stgy.count(t))
    if(len(isolatedtype)==1):
        default = list(set(stgy)-{isolatedtype[0]})[0]
    printer('孤立點個數:%d. 以下種類均被轉化為[%s]'%(
            len(isolatedtype), default)+'\n'+'\n'.join(
            [str(s) for s in isolatedtype[:showcount]])+(
                    '' if(len(isolatedtype)<=showcount) else '\n...'))
    new_stgy = []
    for s in stgy:
        new_stgy.append(default if(s in isolatedtype) else s)
    if(class_max>0):
        #代表需要確認test_sz>=class_sz
        n_class = len(set(new_stgy))
        if(n_class>class_max):
            printer('class number[%d] > class max size[%d]!!! correcting...'%(
                    n_class, class_max))
            sort_counts = list(counts.items())
            sort_counts.sort(key=lambda item: item[1])
            c_class = int(n_class)
            need_merge =[]
            pt_number = 0
            for item in sort_counts:
                need_merge.append(item[0])
                c_class -= 1 if(item[0]!='*') else 0
                pt_number+=item[1]
                if(c_class<=class_max):
                    break
            printer('受減班點個數:%d. 以下種類均被轉化為[%s]'%(
                    pt_number, default)+'\n'+'\n'.join(
            [str(s) for s in need_merge[:showcount]])+(
                    '' if(len(need_merge)<=showcount) else '\n...'))
            opt_stgy = []
            for s in new_stgy:
                opt_stgy.append(default if(s in need_merge) else s)
        else:
            opt_stgy = list(new_stgy)
    else:
        opt_stgy = list(new_stgy)    
    return opt_stgy

def save_dict(dcn, file):
    with open(file, 'wb') as f:
        pickle.dump(dcn, f)
        
def load_dict(file):
    with open(file, 'rb') as f:
        inputs = pickle.load(f)
    return inputs

def linsptrsfmMachine(a,b,c,d,exp_type=None,yubd=None,ylbd=None,xubd=None,xlbd=None):
    def trsfm(x):
        if(xubd!=None):
            x = (np.where(x>xubd,xubd,x)) if(isiterable(x)) else min(x,xubd)
        if(xlbd!=None):
            x = (np.where(x<xlbd,xlbd,x)) if(isiterable(x)) else max(x,xlbd)
        y=(x-a)/(b-a)*(d-c)+c
        if(yubd!=None):
            y = (np.where(y>yubd,yubd,y)) if(isiterable(y)) else min(y,yubd)
        if(ylbd!=None):
            y = (np.where(y>ylbd,ylbd,x)) if(isiterable(y)) else min(y,ylbd)
        return (np.vectorize(exp_type)(x) if(isiterable(y)) else exp_type(y)) if(exp_type!=None) else y
    return trsfm

#TODO:import_data 對於.pkl檔，還要再設定data=None 才會joblib.load它，不然會用read_pickle讀它
def import_data(path='', keyword='', xlsx_formula_header=None, xlsx_head_loc=0, **kwags):
    if((path.split(path_sep)[-1].find('.')==-1) and 'file_type' in kwags):
        #多檔案
        file_type = kwags['file_type']
        print(':%s'%file_type)
        df = pd.DataFrame()
        for root, dirs, files in os.walk(path, topdown=False):
            for file in files:
                if(file_type!='*'):
                    if(file.find('.%s'%file_type)==-1):
                        continue
                if(keyword!=''):
                    if(file.find(keyword)==-1):
                        continue
                total_file = os.path.join(root, file)
                df = import_data(
                        path=total_file, data=df, xlsx_formula_header=xlsx_formula_header, **kwags)
    else:
        #單一檔案
        sheet_name = kwags.get('sheet_name', kwags.get('sheet', kwags.get('sht', 0)))
        paths = kwags['paths'] if('paths' in kwags) else {path:path}
        df = kwags['data'] if('data' in kwags) else pd.DataFrame()
        for key in paths:
            path = paths[key]
            if(path.find('.pkl')>-1):
                if(str(type(df)).find('frame.DataFrame')>-1):
                    path = path%sheet_name if(sheet_name) else path
                    df = pd.read_pickle(path) if(
                            df.empty) else df.append(
                                    pd.read_pickle(path), sort=False)
                    printer('data size:%s'%str(df.shape))
                else:
                    df = joblib.load(path)
            elif(path.find('.json')>-1):
                if(isiterable(df) and str(type(df)).find('frame.DataFrame')==-1):
                    d_type = type(df)
                    df = df+d_type([*LOGger.load_json(path)])
                else:
                    df = LOGger.load_json(path)
            elif(path.find('.h5')>-1):
                df = load_model(path)
            elif(path.find('.xlsx')>-1):
                book_values = openpyxl.load_workbook(path, data_only = True)
                if(not sheet_name in book_values.sheetnames):
                    sheet_name = book_values.sheetnames[0]
                if('sheets' in kwags):
                    sheets = kwags.pop('sheets')
                    sheets = sheets if(isinstance(sheets, list)) else book_values.sheetnames
                    df = pd.DataFrame()
                    for sht in sheets:
                        kwags['sht'] = sht
                        df = import_data(
                                path=path, xlsx_formula_header=xlsx_formula_header, **kwags) if(
                            df.empty) else df.append(
                                import_data(
                                        path=path, xlsx_formula_header=xlsx_formula_header, **kwags),
                                    sort = False)
                else:
                    xlsx_formula_header = xlsx_formula_header if(isinstance(xlsx_formula_header, list)) else []
                    book_formula = openpyxl.load_workbook(path, data_only = False) if(xlsx_formula_header) else book_values
                    index_col = kwags.get('index_col', kwags.get('index', None))
                    xlsx_head_loc = LOGger.extract(xlsx_head_loc, default=xlsx_head_loc)
                    df_values = pd.DataFrame(book_values[sheet_name].values).iloc[xlsx_head_loc:]
                    header = df_values.iloc[0]
                    xlsx_formula_header = [hd for hd in xlsx_formula_header if hd in tuple(header.values)]
                    df_values_res = pd.DataFrame(np.array(df_values.iloc[1:]), columns=header, index=df_values.index[1:])
                    df_values_res.set_index(index_col, inplace=True) if(index_col in df_values_res) else None
                    
                    df_ = df_values_res.copy()
                    if(xlsx_formula_header):
                        df_formula = pd.DataFrame(book_formula[sheet_name].values)
                        df_formula_res = pd.DataFrame(np.array(df_formula.iloc[1:]), columns=header, index=df_formula.index[1:])
                        df_formula_res.set_index(index_col, inplace=True) if(index_col in df_formula_res) else None
                        for hd in xlsx_formula_header:
                            df_[hd] = df_formula_res[hd].copy()
                    df = df_.copy() if(
                            df.empty) else df.append(df_, sort=False).copy()
                printer('data size:%s'%str(df.shape))
            elif(path.find('.csv')>-1):
                index_col = kwags.get('index')
                delimiter = kwags.get('delimiter',',')
                df = pd.read_csv(
                            path, index_col=index_col, delimiter=delimiter) if(
                            df.empty) else df.append(pd.read_csv(
                            path, index_col=index_col, delimiter=delimiter), sort=False)
            else:
                df = joblib.load(path)
    return df

#TODO:explore_folder
def explore_folder(folder, conditions=None, full=True):
    all_files, all_dirs = [], []
    for dirPath, dirNames, fileNames in os.walk(folder):
        dirct = os.path.join(dirPath)
        if(conditions(dirct) if(conditions) else True):
            all_dirs.append(dirct)
        for f in fileNames:
            file = dcp(os.path.join(dirPath, f) if(full) else f)
            if(conditions(file) if(conditions) else True):
                all_files.append(file)
    return all_files, all_dirs
    
#TODO:inverse_encapsulation
def inverse_encapsulation(data, cell_size, sample_mask=None, inverse_encaped_header=None, target_header=None, 
                          index=None, default=None, **kwags):
    """
    

    Parameters
    ----------
    data : TYPE
        DESCRIPTION.
    cell_size : TYPE
        DESCRIPTION.
    sample_mask : TYPE, optional
        DESCRIPTION. The default is None.
    inverse_encaped_header : TYPE, optional
        DESCRIPTION. The default is None.
    target_header : TYPE, optional
        當cell_size為整數時，當作index的header. The default is None.
    index : TYPE, optional
        DESCRIPTION. The default is None.
    default : TYPE, optional
        DESCRIPTION. The default is None.
    **kwags : TYPE
        DESCRIPTION.

    Returns
    -------
    TYPE
        DESCRIPTION.

    """
    addlog = LOGger.addloger(logfile='')
    if(isiterable(cell_size)):
        if(isinstance(data, pd.core.frame.DataFrame)):
            if(target_header==None):    #target_header不指定，就用第一個
                target_header = data.columns[0]
            elif(isinstance(target_header, int)):
                target_header = target_header % len(data.columns)
            data_target_header = data[target_header] #series
            data_target_header_incaped = inverse_encapsulation(
                data_target_header, cell_size, sample_mask=sample_mask, 
                inverse_encaped_header=inverse_encaped_header, target_header=target_header, index=data.index, **kwags)
            data_inverse_encaped = data.drop(target_header, axis=1).join(data_target_header_incaped, sort=False).copy()
        elif(isinstance(data, pd.core.series.Series)):
            np_data = np.array([np.array(s) for s in data])
            np_data_inencped = np_data.reshape(-1, np.product(cell_size))
            data_inverse_encaped = pd.DataFrame(np_data_inencped, columns=inverse_encaped_header, index=index)
        elif(isinstance(data, np.ndarray)):
            data_inverse_encaped = np.array(data).reshape(-1, np.product(cell_size))
        else:
            return default
    elif(LOGger.isinstance_not_empty(cell_size, int)):
        if(isinstance(sample_mask, int)):
            return data
        sample_mask = np.array([True]*cell_size) if(isinstance(sample_mask, type(None))) else sample_mask
        sample_mask = sample_mask[:cell_size]
        sample_mask = np.array(tuple(map((lambda x:True if(x) else False), sample_mask)))
        data_inverse_encaped = None
        for i in range(data.shape[0]):
            default_value = kwags.get('default_value', np.nan)
            df_cell = dcp(data[data.columns[0]].iloc[i] if(isinstance(data, pd.core.frame.DataFrame) or 
                                                           isinstance(data, pd.core.series.Series)) else data[i,:,:])
            if(np.sum(sample_mask.shape[0]) < df_cell.shape[0]):
                addlog('sample_mask length %d < df_cell length %d'%(np.sum(sample_mask.shape[0]), df_cell.shape[0]), stamps=[i])
                print(df_cell)
                return 
            outheader_size = mylist(df_cell.shape).get(1, 1)
            default_value_cell = np.array([default_value]*(cell_size*outheader_size)).reshape(cell_size, outheader_size)        
            df_cell_inverse_sampled = np.where(np.repeat(sample_mask.reshape(-1,1), df_cell.shape[1], axis=1), df_cell, default_value_cell)
            np_df_cell = np.array(df_cell_inverse_sampled)
            data_inverse_encaped = dcp(np.vstack([data_inverse_encaped, np_df_cell]) if(
                not isinstance(data_inverse_encaped, type(None))) else np_df_cell)
        if(isinstance(data, pd.core.frame.DataFrame) or isinstance(data, pd.core.series.Series)):
            data_inverse_encaped = pd.DataFrame(
                data_inverse_encaped, columns=inverse_encaped_header, 
                index=LOGger.get_all_values(
                    [[LOGger.stamp_process('',[i,ii],'','','','_') for ii in range(df_cell.shape[0])] for i in data.index]))
    else:
        return data
    return data_inverse_encaped

def cell_size_configuring(data):
    if(not isiterable(data)):
        return None
    elif(isinstance(data, list) or isinstance(data, tuple)):
        if(len(data)==0):
            return None
        elif(len(np.unique([str(np.array(x).shape) for x in data]))==1):
            if(np.unique([np.array(x).shape for x in data])[0]==()):
                return np.array(data).shape[1] if(len(np.array(data).shape)>0) else None
            return np.array(data).shape[1:]
        else:
            return len(data)
    return np.array(data).shape[1:]

def encapsulation_intCellSize(data, cell_size, sample_mask=None, calibration_tm_header_name=None, 
                              calibration_method = lambda d:(d - np.min(d)).apply(lambda x:x.total_seconds()), **kwags):
    sample_mask = np.array([True]*cell_size) if(isinstance(sample_mask, type(None))) else sample_mask
    if(isinstance(data, np.ndarray)):
        if(len(np.array(data).shape)==1):    
            data = data.reshape(-1, 1)
    if(isinstance(sample_mask, np.ndarray)):
        data_encaped = []
        kwags.update({'backward_step':0,'forward_step':0, 
                      'preserve_all_data':str(type(data)).find('pandas.core')>-1})
        for i in range(0, data.shape[0], cell_size):
            df_cell = data.iloc[i:i+cell_size] if(isinstance(data, pd.core.frame.DataFrame)) else data[i:i+cell_size,:]
            sample_mask = sample_mask[:df_cell.shape[0]]
            sample_mask = np.array(tuple(map((lambda x:True if(x) else False), sample_mask)))
            df_cell_sampled = padding((df_cell[sample_mask] if(
                not isinstance(data, pd.core.frame.DataFrame)) else df_cell.iloc[sample_mask]), **kwags)
            if(calibration_tm_header_name in df_cell_sampled):
                calibration_tm(df_cell_sampled, calibration_tm_header_name, method=calibration_method)
            new_df_cell = np.array(df_cell_sampled) if(
                not isinstance(data, pd.core.frame.DataFrame)) else df_cell_sampled.copy()
            data_encaped.append(new_df_cell)
    elif(isinstance(sample_mask, int)):
        sample_mask = sample_mask % cell_size
        data_encaped = data.iloc[sample_mask : data.shape[0] : cell_size] if(
            isinstance(data, pd.core.frame.DataFrame)) else data[sample_mask : data.shape[0] : cell_size, : ]
    return data_encaped

#TODO:encapsulation
def encapsulation(data, cell_size, sample_mask=None, calibration_tm_header_name=None, encaped_name='main',
                  calibration_method = lambda d:(d - np.min(d)).apply(lambda x:x.total_seconds()), default=None, 
                  out_type=None, is_bcd_out=False, **kwags):
    out_type = out_type if(out_type!=type(None)) else m_encap_out_type
    data_encaped = None
    if(isiterable(cell_size)):
        np_data = np.array(data)
        np_data_encped = np_data.reshape(-1, *cell_size)
        if(False if(isinstance(out_type, type(None))) else issubclass(out_type, pd.core.frame.DataFrame)):
            data_encaped = pd.DataFrame([[v] for v in np_data_encped], columns=[encaped_name])
        elif(False if(isinstance(out_type, type(None))) else issubclass(out_type, pd.core.series.Series)):
            data_encaped = pd.Series([v for v in np_data_encped], name=encaped_name)
        else:
            data_encaped = np_data_encped
    elif(LOGger.isinstance_not_empty(cell_size, int)):
        data_encaped = encapsulation_intCellSize(
            data, cell_size, sample_mask=sample_mask, calibration_tm_header_name=calibration_tm_header_name, 
            calibration_method = calibration_method, **kwags)
    else:
        data_encaped = data.copy()
    if(is_bcd_out):
        data_encaped = bcd.cast_to_floatx(data_encaped)
    return data_encaped

#把data frame裡面的innershape轉成np.ndarray
def deinnershaping(data, cell_size=None, padding_params = None, editor=None, **kwags):
    try:
        data_loc = None
        padding_params = padding_params if(isinstance(padding_params, dict)) else {}
        if(cell_size!=None):
            if(isinstance(cell_size, int)):
                if(cell_size>0):
                    if(not isinstance(padding_params, dict)):
                        padding_params = {'backward_step': 0,
                                          'forward_step': cell_size,
                                          'preserve_all_data': True}
                        padding_params.update(kwags)
        np_data = []
        headers, indexes = mylist(), mylist()
        indexes = LOGger.mylist()
        for i in range(len(data.index) if(hasattr(data, 'index')) else data.shape[0]):
            data_loc = dcp(data.iloc[i].values[0] if(hasattr(data, 'index')) else data[i])
            if(isinstance(cell_size, int)):
                if(cell_size>0):
                    data_loc = padding(data_loc, **padding_params)
            if(callable(editor)):
                data_loc = editor(np.array(data_loc))
            np_data.append(np.array(data_loc))
            header, index = None, None
            if(isinstance(header, type(None)) and (hasattr(data_loc, 'columns') or hasattr(data_loc, 'index'))):
                header = LOGger.mylist(data_loc.columns if(len(data_loc.shape)>1) else data_loc.index)
            if(isinstance(header, type(None)) and (hasattr(data_loc, 'name') or hasattr(data_loc, 'index'))):
                index = LOGger.mylist(data_loc.index if(len(data_loc.shape)>1) else [data_loc.name])
            headers.append(header)
            indexes.append(index)
        np_data = np.array(np_data)
        # ret.update({'header':header, 'indexes':indexes}) if(isinstance(ret, dict)) else {}
    except Exception as e:
        LOGger.exception_process(e, logfile='')
        print('data_loc', data_loc)
        return None, None, None
    return np_data, headers, indexes

def testTensorTransform(handler=LOGger.mystr(), testingFunction=lambda x:encapsulation(x,cell_size=3), stamps=None):
    addlog = LOGger.addloger(logfile='')
    stamps = stamps if(isinstance(stamps, list)) else []
    stamp = LOGger.stamp_process('',stamps,'','','','_')
    setattr(handler,'msgs',[]) if(isinstance(handler, LOGger.mystr)) else None
    for k,v in m_allTensors.items():
        handlerTemp = LOGger.mystr(k) if(isinstance(handler, LOGger.mystr)) else None
        handlerTemp.output = None
        try:
            w = testingFunction(v)
            addlog(type(v), str(getattr(np.array(v),'shape',None)), '->', 
                   type(w), str(getattr(np.array(w),'shape',None)), stamps=stamps+[k], handler=handlerTemp)
            handlerTemp.output = dcp(w)
            handler.msgs.append([handlerTemp, True, stamp]) if(isinstance(handler, LOGger.mystr)) else None
        except Exception as e:
            LOGger.exception_process(e,logfile='', handler=handlerTemp)
            addlog(type(v), str(getattr(np.array(v),'shape',None)), stamps=stamps+[k], handler=handlerTemp)
            handler.msgs.append([handlerTemp, False, stamp]) if(isinstance(handler, LOGger.mystr)) else None
        None if(isinstance(handler, LOGger.mystr)) else addlog('================================================================')
    handler.msgs = pd.DataFrame(handler.msgs, columns=['msg','success', 'stamp'])
    return True

def innershaping(data, header=None, indexes=None, cell_size=None, **kwags):
    try:
        indexes = mylist(indexes if(isiterable(indexes)) else [])
        if(cell_size!=None):
            if(cell_size>0):
                padding_params = {'backward_step': 0,
                                  'forward_step': cell_size,
                                  'preserve_all_data': True}
                padding_params.update(kwags)
                padding_params = padding_params.update(kwags.get('padding_params', {}))
        pd_data = pd.DataFrame()
        for i in data.shape[0]:
            data_loc = data[i,:,:]
            if(cell_size):
                data_loc = padding(data_loc, **padding_params)
            pd_data_loc = pd.DataFrame(data_loc, columns=header, index=indexes.get(i, None))
            pd_data = (pd_data_loc if(pd_data.empty) else pd_data.append(pd_data_loc, sort=False)).copy()
    except Exception as e:
        LOGger.exception_process(e, logfile='')
        return pd.DataFrame()
    return pd_data

def oscn(data, tdata=None, use_abs=True, tdata_lock=True, **kwags):
    if(np.array(data).shape==0):
        return -1
    tdata = tdata if(isiterable(tdata)) else []
    is_tdata = np.array(tdata).shape[0]==np.array(data).shape[0]
    ret, added = 0, 0
    for i in range(np.array(data).shape[0]-1):
        delta_t = (tdata[i+1] - tdata[i]) if(is_tdata) else 1
        if(delta_t>0):
            numerator = abs(data[i+1]-data[i]) if(
                    use_abs) else data[i+1]-data[i]
            ret += (numerator/delta_t)
        else:
            if(tdata_lock):
                {}['[i:%d]delta t error!!!!'%i]
            return np.nan
        added += 1
    return ret/(added if(added>0) else 1)

def make_connected_masks(mask_old, criterion=lambda x:x, **kwags):
    trues_loc = np.where(criterion(mask_old))[0]
    trues_loc_diff = np.diff(trues_loc)
    groups = np.split(trues_loc, np.where(trues_loc_diff != 1)[0] + 1)
    return [g for g in groups if len(g) > 0]

def RiemannSum(data, base, base_gap_criterion=None, default_value=np.nan, dataCount_lbd=3):
    '''
    

    Parameters
    ----------
    base_gap_criterion : TYPE, optional
        base_gap_criterion(area_height) \
            True=>calculate/False=>default_value. The default is None. 
    data : TYPE
        DESCRIPTION.
    base : TYPE
        DESCRIPTION.

    Returns
    -------
    TYPE
        DESCRIPTION.

    '''
    if(len(np.array(base).shape)>1):
        if(np.array(base).shape[1]!=1):
            return default_value
        base = np.array(base.values).reshape(-1)
    if((np.diff(base)<0).any()):
        print('base decreasing!!!!')
        return default_value
    if(len(np.array(data).shape)==1):
        data = np.array(data).reshape(-1, 1)
    if(np.array(data).shape[0]!=np.array(base).shape[0]):
        return default_value
    if(np.array(data).shape[0]<dataCount_lbd):
        return default_value
    np_data = np.array(data)
    area = 0
    for i,inx in enumerate(getattr(data,'index',range(np_data.shape[0]))[:-1]):
        area_height = base[i+1] - base[i]
        if(base_gap_criterion):
            if(not base_gap_criterion(area_height)):
                return default_value
        area += (np_data[i,:] + np_data[i+1,:])*area_height
    return area/2

def redataframeUniqGroupTag(data, grouptagHeader, valueHeader, uniq_method=m_uniq_thru_np):
    if(isinstance(data, (pd.core.frame.DataFrame, pd.core.frame.Series))):
        data = data[np.logical_not(data[grouptagHeader].isna())].copy()
        data[grouptagHeader] = data[grouptagHeader].astype(str)
    groupTags = uniq_method(data[grouptagHeader])
    new_data = pd.DataFrame()
    for i,groupTag in enumerate(groupTags):
        if(groupTag):
            n = data[data[grouptagHeader]==groupTag].shape[0]
            row_data = pd.concat([pd.DataFrame([[groupTag]], index=[valueHeader]), data[data[grouptagHeader]==groupTag][[valueHeader]].T], axis=1)
            row_data.columns = [grouptagHeader, *tuple(list(map(lambda x:x+1, range(n))))]
            row_data.index = [i]
            new_data = (row_data if(new_data.empty) else new_data.append(row_data, sort=False, ignore_index=True)).copy()
    return new_data

def delNanArray(data):
    dtype = (lambda x:np.array(tuple(x))) if(isinstance(data, np.ndarray)) else type(data)
    np_data = np.array(data)
    np_data = np_data[np.logical_not(np.isnan(np_data))]
    np_data = np_data[np.abs(np_data)<np.inf]
    return dtype(map(lambda x:x, np_data))
    

def save_buffer(obj, stamps=None, fn=None, exp_fd='save_buffer'):
    LOGger.CreateContainer(exp_fd) if(not os.path.isdir(exp_fd)) else None
    stamps = stamps if(isinstance(obj, list)) else ['fn']
    fn = fn if(LOGger.isinstance_not_empty(fn, str)) else stamp_process('',stamps,'','','','_',for_file=1)
    if(isinstance(obj, np.ndarray)):
        file = pathrpt(os.path.join(exp_fd, '%s.bin'%fn))
        LOGger.CreateFile(file, lambda f:joblib.dump(obj, file=f))
    elif(isinstance(obj, (pd.core.frame.DataFrame, pd.core.series.Series))):
        file = pathrpt(os.path.join(exp_fd, '%s.pkl'%fn))
        save(obj, save_types=['xlsx'], fn=os.path.basename(file).split('.')[0], exp_fd=exp_fd)
    elif(isiterable(obj)):
        contents = []
        for v in obj:
            contents.append(parse(v))
        file = pathrpt(os.path.join(exp_fd, '%s.json'%fn))
        LOGger.CreateFile(file, lambda f:LOGger.save_json(contents, file=f))
    elif(isinstance(obj, dict)):
        contents = {}
        for k,v in obj.items():
            contents[k] = parse(v)
        file = pathrpt(os.path.join(exp_fd, '%s.json'%fn))
        LOGger.CreateFile(file, lambda f:LOGger.save_json(contents, file=f))
    else:
        file = pathrpt(os.path.join(exp_fd, '%s.json'%fn))
        LOGger.CreateFile(file, lambda f:LOGger.save_json(contents, file=f))

def drawhistogram(seq, file, xlb='', ylb='', title='', xtkrot=0, figsize=(), 
                  maxmin=(), hratio=1, barwidth=0.6, color='#0504aa'):
    n = len(seq)
    if(isinstance(seq, list)):
        n, bins, patches = plt.hist(x=seq, bins='auto', color=color,
                            alpha=0.7, rwidth=barwidth)
        plt.grid(axis='y', alpha=0.75)
        #plt.text(23, 45, r'$\mu=15, b=3$')
        maxfreq = n.max()
        # Set a clean upper y-axis limit.
        plt.ylim(ymax=np.ceil(maxfreq / 10) * 10 if maxfreq % 10 else maxfreq + 10)
    elif(isinstance(seq, dict)):
        #pos = np.arange(len(seq.keys()))
        adseq={str(k):float(seq[k]) if (type(seq[k])==int or type(seq[k])==float or type(seq[k])==str
                 ) else (len(seq[k]) if (subscrtb(seq[k])) else seq[k]) for k in seq.keys()}
        # gives histogram aspect to the bar diagram
        friendly_height = (len(set(list(adseq.values())))*0.02+15)*hratio
        yvaluemax = np.array(list(adseq.values())).max() if (len(maxmin)<2) else maxmin[1]
        yvaluemin = np.array(list(adseq.values())).min() if (len(maxmin)<2) else maxmin[0]
        figsz = (n*0.2+4, friendly_height) if(figsize==()) else tuple(figsize)
        f, ax = plt.subplots(figsize=figsz) # set the size that you'd like (width, height)
        if(len(maxmin)<2):
            ystd = np.std(np.array(list(adseq.values())))
            yunit = ystd // 10 if ystd>=10 else (int(ystd) if ystd>1 else 1)
            plt.ylim(ymax = np.ceil(yvaluemax / yunit) * yunit if yvaluemax % yunit else yvaluemax + yunit,
                     ymin = (yvaluemin // yunit) * yunit if yvaluemin % yunit else yvaluemin - yunit)
        else:
            plt.ylim(maxmin[0], maxmin[1])
        plt.bar(adseq.keys(), adseq.values(), barwidth, color=color)
    plt.title(title)
    plt.xlabel(xlb)
    plt.ylabel(ylb)
    plt.xticks(rotation=xtkrot)
    if(not file==''):
        plt.savefig(file)
    plt.ioff()
    plt.show()

def diet(ipdf, xcols, maxct=3, showlog=False, re_detail=False):
    opr_df = ipdf[xcols]
    wait_df = ipdf.drop(xcols, axis=1)
    new_dic = {}
    count = {}
    for i in opr_df.index:
        x=tuple(opr_df.T[i])
        count[x]=1 if (not x in count.keys()) else count[x]+1
        if(count[x]<=maxct):
            new_dic[i] = {k: opr_df[k][i] for k in opr_df.columns}
        else:
            if(showlog):
                printer('偵測到%s已累積%d筆，不採納'%(str(x),count[x]))
    opt_df = (pd.DataFrame.from_dict(new_dic).T).join(wait_df)
    if(re_detail):
        newct = {x: count[x] if count[x]<=maxct else 3 for x in count.keys()}
        return opt_df, count, newct
    else:
        return opt_df

def project_record_ending(dic, exp_fd, sheet_name='test', stamps=[], theme='', **kwags):
    addlog = kwags.get('addlog', LOGger.addloger(logfile=os.path.join(exp_fd, 'log.txt')))
    if(not LOGger.isinstance_not_empty(dic, dict)):
        addlog('dic empty!!!', stamps=stamps)
        return False
    dic1 = {}
    for k,v in dic.items():
        if(isiterable(v) or isinstance(v, dict)):
            continue
        dic1.update({k:v})
    prm_fn = 'records%s.xlsx'%(stamp_process('_', [theme], '','','','_', location=-1))
    log = '讀取參數檔案:%s.....'%(prm_fn)
    addlog(log, stamps=stamps)
    evl = evaluation()
    evl.add_export(dic1, prm_fn, sheet_name = sheet_name)
    return True

def make_pathhyperlink_list(source_fd, target_extfiletypes=None):
    files, _ = explore_folder(source_fd)
    file_stgs = []
    for file in files:
        if(target_extfiletypes!=None):
            extfiletype = file[file.rfind('.')+1:]
            if(mylist(target_extfiletypes).find(extfiletype, strictly=True)==None):
                continue
        file_stgs.append('=HYPERLINK("%s", "%s")'%(file, os.path.basename(file)))
    return file_stgs

def export_pathhyperlink_list(source_fd, target_extfiletypes=None, index=None, exp_fd='.', save_types=None, fn=None,
                              stamps=None, **kwags):
    file_stgs = make_pathhyperlink_list(source_fd, target_extfiletypes=target_extfiletypes)
    report = pd.DataFrame(index=index)
    report['path'] = file_stgs
    stamps = stamps if(isinstance(stamps, list)) else []
    fn = fn if(isinstance(fn, str)) else stamp_process('',[os.path.basename(os.path.abspath(source_fd))]+stamps,'','','','_',for_file=1)
    save_types = save_types if(isinstance(save_types, list)) else ['xlsx']
    save(report, fn=fn, exp_fd=exp_fd, save_types=save_types, **kwags)
    return True

def evaluateprd(y_pred, y_real, tol, xdim=0):
    ret={}
    ret['mse'] = skm.mean_squared_error(y_real, y_pred)
    ret['rmse'] = ret['mse']**(1/2)
    ret['r2'] = skm.r2_score(y_real, y_pred)
    if(xdim!=0):
        ret['adr2'] = 1-((1-ret['r2'])*(len(y_real)-1)/(len(y_real)-xdim-2))
    nodelay_count = 0
    noerror_count = 0
    nodelay_and_noerror_count = 0
    for i in range(len(y_pred)):
        if(y_pred[i]>y_real[i]):
            nodelay_count += 1
        if(abs(y_pred[i]-y_real[i])<tol):
            noerror_count += 1
        if(y_pred[i]>y_real[i] and y_pred[i]<y_real[i]+tol):
            nodelay_and_noerror_count  += 1
    size = len(y_real)
    ret['OKR1'] = nodelay_count*100/size
    ret['OKR2'] = noerror_count*100/size
    ret['OKR3'] = nodelay_and_noerror_count*100/size
    return ret

def data_df2csv(DFs, file, colname=[], isindex=True, axis=1):
    collen=0
    oDF = pd.DataFrame(index=DFs[0].index)
    for DF in DFs:
        collen += len(pd.DataFrame(DF).columns)
        oDF = pd.concat([oDF, DF], axis=axis)
        
    
    if(type(colname)==list):
        if(not colname==[] and len(colname)==collen):
            header = colname
        else:
            header = True
    elif(type(colname)==bool):
        header = colname
     
    oDF.to_csv(file, header = header, index = isindex)
    
def data_df2xlsx(DFs, writer, sheet, colname=[], isindex=True, axis=1):
    collen=0
    oDF = pd.DataFrame(index=DFs[0].index)
    for DF in DFs:
        collen += len(pd.DataFrame(DF).columns)
        oDF = pd.concat([oDF, DF], axis=axis)
        
    
    if(type(colname)==list):
        if(not colname==[] and len(colname)==collen):
            header = colname
        else:
            header = True
    elif(type(colname)==bool):
        header = colname
     
    oDF.to_excel(writer, sheet, header = header, index = isindex)
    
def tplize(obj):
    if(type(obj)==str):
        tp = (obj,) 
    else:
        try:
           tp = tuple(obj)
        except TypeError:
           tp = tuple([obj])
    return tp

def drop(dic, key):
    ddic = dic.copy()
    if(len(tplize(key))>1):
        for k in key:
            ddic = drop(ddic, k)
    else:
        ddic.pop(key)
    return ddic

def drawparams(grid_search, file='', xbd=(), ybd=(0.6, 1.1), figsize=(), showmodes=False):
    results = pd.DataFrame(grid_search.cv_results_)
    best = np.argmax(results.mean_test_score.values)
    if(figsize==()):
        pfy = len(grid_search.best_params_)*8
        pfx = 60
        figsize = (pfx, pfy)
    plt.figure(figsize = figsize)
    xbd = (-1, len(results)) if xbd==() else xbd
    plt.xlim(xbd[0], xbd[1])
    plt.ylim(ybd[0], ybd[1])
    #marker_best = 0
    for i, (_, row) in enumerate(results.iterrows()):
        scores = row[['split%d_test_score' % j for j in range(5)]]
        marker_cv, = plt.plot([i] * 5, scores, '^', c='gray', markersize=5,alpha=.5)
        marker_mean, = plt.plot(i, row.mean_test_score, 'v', c='none', alpha=1,
                                markersize=10, markeredgecolor='k')
        if i == best:
            marker_best, = plt.plot(i, row.mean_test_score, 'o', c='red',
                                    fillstyle="none", alpha=1, markersize=20,
                                    markeredgewidth=1)
    if(not showmodes):
        grid_par = [drop(d,'kernel') for d in grid_search.cv_results_['params']]
    else:
        grid_par = grid_search.cv_results_['params']
    plt.xticks(range(len(results)), [str(x).strip("{}").replace("'", "") for x
                                     in grid_par],rotation=90)
    plt.ylabel("Validation score")
    plt.xlabel("Parameter settings")
    plt.legend([marker_cv, marker_mean, marker_best],
               ["cv score", "mean score", "best parameter setting"],
               loc=(1.05, .4))
    #plt.legend([marker_cv, marker_mean],
    #           ["cv accuracy", "mean accuracy"],
    #           loc=(1.05, .4))
    if(file!=''):
        plt.savefig(file)
    plt.show()
    plt.ioff()

def mergeisolatedgroup(stgy, default_ts):
    printer('分類依照的性質種類如下:')
    printer(str(set(stgy)))
    isolatedtype = []
    for t in set(stgy):
        if(stgy.count(t)<2):
            isolatedtype.append(t)
    checkround=0
    while(isolatedtype!=[] and checkround<len(default_ts)):
        printer('按性質分類時發現某些種類的訓練個數只有1個:')
        printer(str(isolatedtype))
        printer('將這些資料都歸類給類別:%s'%
               str(list(default_ts.keys())[checkround]))
        for t in isolatedtype:
            stgy = [(default_ts[list(default_ts.keys())[checkround]])(s) 
                        if s==t else s for s in stgy]
        printer('分類依照的性質種類如下:')
        printer(str(set(stgy)))
        checkround += 1
        isolatedtype = []
        for t in set(stgy):
            if(stgy.count(t)<2):
                isolatedtype.append(t)
    return stgy

def unduplicate_data(ipt_data, exp_fd='', file='', ipt_tp='xlsx', exp_tp='xlsx', replace=0):
    '''------------------匯入---------------------------------'''
    pathsep = os.path.join('a','b')[1]
    if(ipt_tp=='xlsx'):
        file = ipt_data.split(pathsep)[-1] if(file=='') else file
        ipt_df = {}
        book = openpyxl.load_workbook(ipt_data)
        for sht in book.sheetnames:
            df0 = pd.read_excel(ipt_data, sht, index_col=0)
            ipt_df[sht] = df0.copy()
        if(exp_fd==''):
            exp_fd = (os.path.join(*tuple(ipt_data.split(pathsep)[:-1])) if(
                            ipt_data.split(pathsep)[:-1]!=[]) else '') if(
                                                        exp_fd=='') else exp_fd
    
    '''------------------匯出---------------------------------'''
    if(exp_tp=='xlsx'):
        if(file==''):
            {}['err:no file name']
        newfile = pathrpt(os.path.join(exp_fd, file.replace('.','_undup.'))) if(
                    replace==0 or type(ipt_data)!=str) else ipt_data
        wrt = pd.ExcelWriter(newfile, engine='xlsxwriter')
        for sht in ipt_df:
            df0 = ipt_df[sht].copy()
            df1 = df0.loc[~df0.index.duplicated(keep='first')].copy()
            df1.to_excel(wrt, sht)
        wrt.save()
#TODO:pathrpt
def pathrpt(apath, sep='_'):
    adir, abanme = os.path.split(apath)
    eext = (path_sep if abanme[-1]==path_sep else '') if(abanme.find('.')==-1) else abanme[abanme.rfind('.'):]
    cond = lambda p: os.path.isfile(p) if eext!='' else os.path.isdir(p)
    p = (apath[:-1] if apath[-1]==path_sep else apath) if eext=='' else apath[:apath.rfind('.')]
    while(cond('%s%s'%(p, eext) if eext!='' else p)):
        '''如果p已存在，則繼續修改..........'''
        final_loc = p[p.rfind(path_sep)+1:]
        if (final_loc.rfind(sep)==-1 or (not final_loc[(final_loc.rfind(sep)+1):].isnumeric())):
            '''最內一層「沒有「<_sep_>」字元存在 或者 <_sep_>後面接著不是純數目」的話'''
            p = p+sep+'1'
        else:
            '''最內一層「有「<_sep_>」字元存在 且 <_sep_>後面接著是純數目」的話'''
            n = int(p[(p.rfind(sep)+1):]) + 1 #找出已存在複本的編序，然後+1
            rpstr = p[p.rfind(sep):] #找出要修改的名稱段落
            p = p[:p.rfind(path_sep)+1] + final_loc.replace(rpstr, '%s%d'%(sep,n)) 
    return '%s%s'%(p, eext)

def check_bounds(params, data_header, bound=(-np.inf,np.inf), bound_tables=None, 
                 stamps=None, anno_table=None, min_header_anno='min' , max_header_anno='max',
                 tags_header_anno='Tags', **kwags):
    host = kwags.get('host', None)
    log_counter = getattr(host, 'log_counter', kwags.get('log_counter', None))
    exp_fd = getattr(host, 'exp_fd', kwags.get('exp_fd', None))
    logfile = getattr(host, 'logfile', kwags.get('logfile', (os.path.join(exp_fd, 'log.txt') if(os.path.isdir(exp_fd)) else '')))
    addlog = getattr(host, 'addlog', LOGger.addloger(logfile=logfile))
    stamps = kwags.get('stamps', [])
    table_stamps = kwags.get('table_stamps', data_header)
    bound_tables = bound_tables if(isinstance(bound_tables, dict)) else getattr(host, 'bound_tables', {})
    bound_ = bound_tables.get(table_stamps, bound)
    if(isinstance(anno_table, pd.core.frame.DataFrame)):
        if(min_header_anno in anno_table and max_header_anno in anno_table):
            if(anno_table[[min_header_anno,max_header_anno]][anno_table[tags_header_anno]==data_header].shape[0]==1):
                try:
                    map(float, anno_table[[min_header_anno,max_header_anno]][anno_table[tags_header_anno]==data_header].values)
                    bound_ = tuple(map(float, anno_table[[
                        min_header_anno,max_header_anno]][anno_table[tags_header_anno]==data_header].values))
                    log_counter.update({'check bounds error':0}) if(isinstance(log_counter, dict)) else None
                except:
                    addlog('anno file content may error:\n%s'%anno_table[[
                           'min','max']][anno_table[tags_header_anno]==data_header], 
                           stamps=stamps+[log_counter.get('check bounds error', 0)],
                           log_counter=log_counter, log_counter_stamp='check bounds error')
    # addlog('bound:%s'%str(bound_), stamps=[data_header])
    # df_opr_for_debug = list(tuple(ctor.df_opr[data_header].values))
    if(not data_header in params):
        addlog('header:%s 不存在!!!!!!!'%data_header, stamps=stamps+[
            check_bounds.__name__, log_counter.get('status unactivated', 0)], 
            log_counter=log_counter, log_counter_stamp='status unactivated')
        return False #TODO:可能要exit()
    if(params[data_header]<bound_[0] or params[data_header]<bound_[1]):
        addlog('超出範圍(%s,%s)!!!!'%tuple(map(parse, bound_)), 
               stamps=stamps+[data_header, log_counter.get('status unactivated', 0)], 
               log_counter=log_counter, log_counter_stamp='status unactivated')
        return False
    return True

def mergetoindex(in_df, key_property, axis=0, ret='index'):
    if(axis==0):
        dataindex = [str(tuple(str(in_df[key][i]) 
                        for key in key_property)).replace("'",
                                                "").replace(", ",
                                                "_").replace("(",
                                                "").replace(")",
                                                "")
                        for i in in_df.index]
    else:
        dataindex = [str(tuple(str(in_df[j][key]) 
                        for key in key_property)).replace("'",
                                                "").replace(", ",
                                                "_").replace("(",
                                                "").replace(")",
                                                "")
                        for j in in_df.columns]
    return dataindex if(ret=='index') else ((replaceindex(in_df, dataindex) 
                                            if(axis==0) else (replaceindex(in_df.T, dataindex)).T),
                                            dataindex)

#TODO:calibration_tm
def calibration_tm(data, old_tm_header, new_tm_header='tm', time_format='%Y-%m-%d %H:%M:%S.%f', 
                   method=lambda d:(d - np.min(d)).apply(lambda x:x.total_seconds()), **kwags):
    if(isinstance(data[old_tm_header].iloc[0], str)):
        data_parsed = data[old_tm_header].map(lambda s:dt.strptime(s,time_format))
    elif(isinstance(data[old_tm_header].iloc[0], dt)):
        data_parsed = data[old_tm_header]
    elif(isinstance(data[old_tm_header].iloc[0], dtC.time)):
        data_parsed = data[old_tm_header].map(lambda s:dt.combine(dt.now().date(), s))
    elif(isinstance(data[old_tm_header].iloc[0], dtC.date)):
        data_parsed = data[old_tm_header].map(lambda s:dt.combine(s, dt.min.time()))
    else:
        data_parsed = data[old_tm_header].copy()
    data[new_tm_header] = method(data_parsed)
    return True

def bind(dataframelist, axis=0):
    ret_df = dataframelist[0]
    if(axis==0 or axis=='r'):
        for df in dataframelist:
            ret_df = ret_df.append(df)
    elif(axis==1 or axis=='c'):
        for df in dataframelist:
            ret_df = ret_df.join(df)
    return ret_df

def replaceindex(in_df, index):
    if(len(index)!=len(in_df.index)):
        printer('index數量%d與新的index不符!!!')
        return in_df
    out_df = pd.DataFrame(np.mat(in_df), index=index, columns=in_df.columns)
    return out_df

def gettimeinterval(in_df, sttimelb, stimepstg, edtimelb, edimepstg, unit='seconds', index='', columns=''):
    if(unit=='seconds'):
        setimeinterval = [(dt.strptime(str(in_df[edtimelb][i]), edimepstg) -
                             dt.strptime(str(in_df[sttimelb][i]), stimepstg)).seconds 
                                for i in in_df.index]
    elif(unit=='minutes'):
        setimeinterval = [(dt.strptime(str(in_df[edtimelb][i]), edimepstg) -
                             dt.strptime(str(in_df[sttimelb][i]), stimepstg)).minutes 
                                for i in in_df.index]
    elif(unit=='hours'):
        setimeinterval = [(dt.strptime(str(in_df[edtimelb][i]), edimepstg) -
                             dt.strptime(str(in_df[sttimelb][i]), stimepstg)).hours 
                                for i in in_df.index]
    elif(unit=='days'):
        setimeinterval = [(dt.strptime(str(in_df[edtimelb][i]), edimepstg) -
                             dt.strptime(str(in_df[sttimelb][i]), stimepstg)).days 
                                for i in in_df.index]
    elif(unit=='months'):
        setimeinterval = [(dt.strptime(str(in_df[edtimelb][i]), edimepstg) -
                             dt.strptime(str(in_df[sttimelb][i]), stimepstg)).months 
                                for i in in_df.index]
    elif(unit=='years'):
        setimeinterval = [(dt.strptime(str(in_df[edtimelb][i]), edimepstg) -
                             dt.strptime(str(in_df[sttimelb][i]), stimepstg)).years 
                                for i in in_df.index]
    else:
        printer('沒有這種時間單位:%s'%unit)
    return pd.DataFrame(setimeinterval, 
                        index=in_df.index if(index=='') else index, 
                        columns=in_df.columns if(columns=='') else columns)

#TODO:read_excel
def read_excel(file, sheet_name, index_col=0, data_only = False):
    book_formula = openpyxl.load_workbook(file, data_only = data_only)
    printer('讀取到[%d]張原始分頁..................'%len(book_formula.sheetnames))
    df = pd.DataFrame()
    for sht in book_formula.sheetnames:
        if(sheet_name==sht):
            df0 = pd.DataFrame(book_formula[sht].values)
            print('df0:\n%s'%df0)
            df0 = df0.set_index(index_col)
            header = df0.iloc[0]
            df = pd.DataFrame(np.array(df0.iloc[1:]), columns=header, index=df0.index[1:])
            print('df:\n%s'%df)
            break
    return df

class evaluation():
    def __init__(self):
        self.storage=pd.DataFrame()
        self.storageboxes={}
        
    def hey(self):
        printer('__name__:',__name__)
        printer(self.import_method)
        printer(self.storage)
    
    def read(self, file, sheet_name=None, import_method='xlsx', index_col=0):
        if(type(import_method)=='dict'):
            import_method = list(import_method.values())
        if(np.array(import_method).shape!=()):
            filetype = (import_method)[0]
        else:
            filetype = import_method
        if(np.array(import_method).shape!=()):
            sheet_name = (import_method)[1]
        if(filetype=='xlsx'):
            def method(f):
                book_formula = openpyxl.load_workbook(f, data_only = False)
                printer('讀取到[%d]張原始分頁..................'%len(book_formula.sheetnames))
                df = pd.DataFrame()
                for sht in book_formula.sheetnames:
                    df0 = pd.DataFrame(book_formula[sht].values)
                    df0 = df0.set_index(index_col)
                    header = df0.iloc[0]
                    df0 = pd.DataFrame(np.array(df0.iloc[1:]), columns=header, index=df0.index[1:])
                    if(sheet_name==sht):
                        df = df0.copy()
                    else:
                        self.storageboxes[sht] = df0.copy()
                return df
        elif(filetype=='csv'):
            def method(file):
                return pd.read_csv(file, index_col = index_col)
        try:
            df = method(file)
            printer(df)
            if(sheet_name!=None):
                self.storageboxes[sheet_name] = df.copy()
            else:
                self.storage = df.copy()
            return df, ''
        except:
            log1 = 'Unexpected error:%s'%(sys.exc_info()[0])
            printer(log1)
            log2 = '[SHEET:%s]讀取異常:%s'%(str(sheet_name), file)
            printer(log2)
            return None, [log1,log2]
    def add(self, data, index=None, columns=None, sheet_name=None):
        if(sheet_name!=None):
            if(sheet_name in self.storageboxes.keys()):
                adddata = pd.DataFrame(data, 
                                       index=index, 
                                       columns=(self.storageboxes[sheet_name].columns) if(
                                               columns==None) else columns)
                self.storageboxes[sheet_name] = adddata.append(
                        self.storageboxes[sheet_name], sort=False)
            else:
                adddata = pd.DataFrame(data, index=index, columns=columns)
                self.storageboxes[sheet_name] = adddata
        else:
            adddata = pd.DataFrame(data, index=index, 
                                       columns=self.storage.columns if(
                                       columns==None) else columns)
            self.storage = adddata.append(self.storage, sort=False)
        
        printer('%sstorage總量:%d'%('' if(sheet_name==None) else '[%s]'%sheet_name,
                                len(self.storage if(sheet_name==None) else(
                                self.storageboxes[sheet_name]))))
        
    def export(self, file, sheetes=[], export_method = 'xlsx', 
               index=None, columns=None, showindex=True, showcolumns=True):
        if(type(sheetes)!=list):
            printer('sheetes型態錯誤:%s'%(str(type(sheetes))))
            return  
        if(len(sheetes)==0 and self.storageboxes=={}):
            df = self.storage
            if(index!=None):
                npdf = np.mat(df)
                df = pd.Dataframe(npdf, index=index)
            if(columns!=None):
                npdf = np.mat(df)
                df = pd.Dataframe(npdf, columns=columns)
        filetype = file[(file.rfind('.')+1):]
        if(filetype=='xlsx'):
            wrt = pd.ExcelWriter(file, engine='xlsxwriter')
            if(len(sheetes)>0):
                printer('sheet張數:%d'%len(sheetes))
                for sheet in sheetes:
                    if(sheet not in self.storage.keys()):
                        printer('[%s]storge內沒有這張sheet!!!'%sheet)
                        continue
                    df = self.storageboxes[sheet]
                    df.to_excel(wrt, sheet, header = showcolumns, index = showindex)
            else:
                if(self.storageboxes=={}):
                    df = self.storage
                    df.to_excel(wrt, header = showcolumns, index = showindex)
                else:
                    for sheet in self.storageboxes:
                        df = self.storageboxes[sheet]
                        df.to_excel(wrt, sheet, header = showcolumns, index = showindex)
            wrt.save()
        elif(filetype=='csv'):
            df.to_csv(file, header = showcolumns, index=showindex)
    
    def add_export(self, dic, prm_fn, sheet_name, **kwags):
        self.read(prm_fn, sheet_name = sheet_name, **kwags)
        self.add(dic, sheet_name = sheet_name,
                index=[dt.now().strftime('%Y-%m-%d %H:%M:%S')], 
                columns = list(dic.keys()), **kwags)
        self.export(prm_fn, **kwags)

def xlsx2list(file, **kwags):
    sheet_name = kwags.get('sheet_name', kwags.get('sheet', kwags.get('sht')))
    index_col = kwags.get('index_col')

    # 使用 pandas 讀取 Excel 文件
    df = pd.read_excel(file, sheet_name=sheet_name, index_col=index_col)

    # 逐行處理數據
    output = []
    for index, row in df.iterrows():
        # 這裡的 'row' 是一個 Series 對象，代表當前行的數據
        # 您可以通過列名來訪問具體的值，例如: row['column_name']
        sys.exit(1)
        row = np.where(row.isna(), 0, row)
        while(row[-1]==0):
            row = row[:-1]
        output.append(row)
    return output

#TODO:determined_uthreshold
def determined_uthreshold(data, n_sigma=3, contamination=None, total_selected_label=None, **kwags):
    total_selected_label = np.full(data.reshape(-1).shape[0], True) if(isinstance(total_selected_label, type(None))) else total_selected_label
    #如果進容差的資料太少，就還是以原data來調
    if(dict(zip(*np.unique(total_selected_label, return_counts=1))).get(True, 0)<4):
        total_selected_label = np.full(data.shape[0], True)
    data_good = data.reshape(-1)[total_selected_label]
    if(contamination==None):
        std = np.std(data_good, axis=0)
        median = np.median(data_good, axis=0)
        return median+n_sigma*std
    return np.percentile(data_good, (1-contamination)*100, axis=0)

#TODO:determined_uthreshold
def determined_lthreshold(data, n_sigma=3, contamination=None, total_selected_label=None, **kwags):
    total_selected_label = np.full(data.reshape(-1).shape[0], True) if(isinstance(total_selected_label, type(None))) else total_selected_label
    #如果進容差的資料太少，就還是以原data來調
    if(dict(zip(*np.unique(total_selected_label, return_counts=1))).get(True, 0)<4):
        total_selected_label = np.full(data.shape[0], True)
    data_good = data.reshape(-1)[total_selected_label]
    if(contamination==None):
        std = np.std(data_good, axis=0)
        median = np.median(data_good, axis=0)
        return median-n_sigma*std
    return np.percentile(data_good, contamination*100, axis=0)

def determined_bthreshold(data, n_sigma_l=3, contamination_l=None, n_sigma_u=3, contamination_u=None, total_selected_label=None, **kwags):
    uthreshold = determined_uthreshold(data, n_sigma=n_sigma_u, contamination=contamination_u, total_selected_label=total_selected_label, **kwags)
    lthreshold = determined_lthreshold(data, n_sigma=n_sigma_l, contamination=contamination_l, total_selected_label=total_selected_label, **kwags)
    return uthreshold, lthreshold

def find_max_consequential_repeated_number(data, same_value_threshold=0.05, default=None):
    current_number = np.nan
    recent_numbers = [np.nan] * 5
    count = 0
    max_count = 0
    max_repeated_number = default

    for i in range(len(data) - 1):
        if abs(data[i] - data[i + 1]) <= same_value_threshold:
            if abs(data[i] - current_number) > same_value_threshold or count == 0:
                current_number = data[i]
                count = 1
                recent_numbers[0] = data[i]  # 初始化数组
            else:
                count += 1
                if count <= 5:
                    recent_numbers[count - 1] = data[i + 1]
                else:
                    # 将元素向左移动并添加新数据
                    recent_numbers.pop(0)
                    recent_numbers.append(data[i + 1])
        else:
            if count >= 5 and count > max_count:
                max_count = count
                max_repeated_number = np.median(recent_numbers)
            count = 0
            current_number = -1

    if count >= 5 and count > max_count:
        max_repeated_number = np.median(recent_numbers)

    return max_repeated_number if(max_repeated_number!=np.nan) else default

def find_max_consequential_repeated_number1(data, same_value_threshold=0.02, default=None, datacounts_lower_bound=13):
    if(np.array(data).shape[0]<datacounts_lower_bound):
        return default
        
    current_number = np.nan
    count = 0
    max_count = 0
    repeated_number = default
    max_repeated_number = default

    for i in range(len(data) - 1):
        if abs(data[i] - current_number) > same_value_threshold or count == 0:
            current_number = dcp(data[i])
            count = 1
            repeated_number = dcp(data[i])  # 初始化数组
        else:
            repeated_number = (repeated_number * count + data[i])/(count + 1)
            count += 1
            
        if count > max_count:
            max_count = dcp(count)
            max_repeated_number = dcp(repeated_number)

    if max_count <= 5:
        return default

    return max_repeated_number

#eps=5, min_samples=2, metric='manhattan'
def dbscaning(data, do_normalizing=True, return_full_information=False, **kwags):
    ret = {}
    scr = None
    str_array_type = str(type(data))
    if(str_array_type.find('tuple')>-1):
        data = np.array(data)
        ret.update(dbscaning(data, **kwags))
    if(str_array_type.find('list')>-1):
        data = np.array(data)
        ret.update(dbscaning(data, **kwags))
    if(str_array_type.find('pandas')>-1):
        drop_columns = data.columns[data.applymap(lambda s:True if(isnonnumber(s)) else np.isnan(s)).any()]
        data = data.drop(drop_columns, axis=1)
        nmd_data, scr = normalization(data) if(do_normalizing) else (data, None)
        clustering = DBSCAN(**kwags).fit(nmd_data)
        label_mask = clustering.labels_
        label_set = list(set(clustering.labels_))
        if(return_full_information):
            clustering.label_set = label_set
            clustering.scr = scr
            ret.update({'dbcst':clustering, 'scr':scr})
        else:
            ret.update({'label_mask':label_mask, 'label_set':label_set})
    if(str_array_type.find('ndarray')>-1):
        nmd_data, scr = normalization(data) if(do_normalizing) else (data, None)
        clustering = DBSCAN(**kwags).fit(nmd_data)
        label_mask = clustering.labels_
        label_set = list(set(clustering.labels_))
        if(return_full_information):
            clustering.label_set = label_set
            clustering.scr = scr
            ret.update({'dbcst':clustering, 'scr':scr})
        else:
            ret.update({'label_mask':label_mask, 'label_set':label_set})
    return ret

#TODO:feature_clustering
def feature_clustering_auto_eps(data, eps_lbd=0, eps_ubd=10, min_samples=100, max_iter=10, outlier_size_ubd=None, 
                       feature_range_checking_method=(lambda d:True), **kwags):
    addlog = kwags.get('addlog', LOGger.addloger(logfile=''))
    eps_ubd = kwags.get('eps', (eps_ubd + eps_lbd)/2) * 2 - eps_lbd
    outlier_size_ubd = (np.array(data).shape[0] - min_samples) if(outlier_size_ubd==None) else outlier_size_ubd
    success = False
    i=0
    log_counter = {}
    while (i<max_iter):
        i += 1
        is_continue = False
        eps=(eps_lbd+eps_ubd)/2
        label_mask, label_set = tuple(dbscaning(data, eps=eps, min_samples=min_samples).values())
        dis_num = dict(zip(*np.unique(label_mask, return_counts=1))).get(-1, 0)
        if(dis_num>outlier_size_ubd):
            addlog('半徑太小[%.4f]，噪音個數%d>%d，無法成群!!!'%(eps, dis_num, outlier_size_ubd), stamps=[i])
            eps_lbd = dcp(eps)
            log_counter['show_noise'] = 0
            continue
        addlog('噪音個數%d(<=%d)'%(dis_num, outlier_size_ubd), stamps=[i], 
               log_counter=log_counter, log_counter_stamp='show_noise')
        for label in [v for v in label_set if v!=-1]:
            data_label = data[label_mask==label]
            addlog('data shape:(%d, %d)'%(tuple(data_label.shape)), stamps={'iter':i, 'label':label})
            if(not feature_range_checking_method(data_label)):
                addlog('半徑太大[%.4f]!!!'%(eps), stamps={'iter':i, 'label':label})
                eps_ubd = dcp(eps)
                is_continue = True
                break
        if(is_continue):
            continue
        success = True
        break
    if(not success):
        addlog('分群未成功!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
    return label_mask, label_set

#要作完整資料的oscn，adjacent=(data.shape[0],0)，adjacent=(len(data.index),0)
#tm_coln=None表示使用tm_col=range(data.shape[0])
def making_oscn(data, adjacent=(2,0), use_abs=True, is_adjacent_memo=False, 
                tm_coln=0, opr_coln=None, opr_columns=[], 
                tm_col=[], return_full_data=True, return_oscns_data=False, 
                return_na_columns=False, **kwags):
    ret = {}
    str_type_data = str(type(data))
    if(str_type_data.find('ndarray')>-1):
        opr_coln = opr_coln if(opr_coln) else 1
        tm_col = dcp(tm_col) if(np.array(tm_col).shape[0]==np.array(
                data).shape[0]) else range(data.shape[0])
        new_data = data[adjacent[0]:(-adjacent[1]),:] if(
                adjacent[1]) else data[adjacent[0]:,:]
        if(opr_columns==[]):
            oscns = []
            for i in range(new_data.shape[0]):
                data_i = data[i:(i+adjacent[0]+adjacent[1]+1),:]
                tm_col_i =  [v for v in data_i[:,tm_coln]] if(
                        tm_coln!=None) else tm_col[i:(i+adjacent[0]+adjacent[1]+1)]
                #只以第tm_coln欄作用第opr_coln欄
                oscn_value = oscn([v for v in data_i[:,opr_coln]], tm_col_i, use_abs=use_abs)
                oscns.append(oscn_value)
            np_oscns = np.array(oscns)
            new_data = np.insert(new_data, opr_coln+1, np_oscns, axis=1)
        else:
            if(opr_columns=='full'):
                opr_columns = [v for v in range(data.shape[1]) if(
                        v!=tm_coln if(tm_coln!=None) else True)]
            np_oscns, na_columns = [], []
            for opr_column in sorted(opr_columns, reverse=True):#反著過來作就不會改變前面的順序 ─ opr_coln的指定
                try:
                    data[:,opr_column] = data[:,opr_column].astype(float)
                except:
                    printer('[%s]存在非數字元素'%opr_column)
                    na_columns.append(opr_column)
                else:
                    oscn_data = making_oscn(data, adjacent=adjacent, use_abs=use_abs,
                               tm_coln=tm_coln, 
                               opr_coln=opr_column, opr_columns=[], tm_col=[],
                               return_full_data=False, return_oscns_data=True, **kwags)[
                                            'oscn_data']
                    new_data = np.insert(new_data, opr_column+1, oscn_data, axis=1)
                    np_oscns.append(oscn_data)
            np_oscns = np.transpose(np.array(np_oscns))
    if(str_type_data.find('pandas.core')>-1):
        opr_coln = list(data.columns).index(opr_coln) if(opr_coln) else 1
        tm_coln = list(data.columns).index(tm_coln) if(tm_coln) else 0
        opr_columns = [v for v in data.columns if(
                v!=data.columns[tm_coln] if(tm_coln!=None) else True)] if(
                opr_columns=='full') else  opr_columns
        np_opr_columns = [list(data.columns).index(c) for c in opr_columns]
        tm_col = dcp(tm_col) if(np.array(tm_col).shape[0]>0) else []
        new_data = data.iloc[adjacent[0]:(-adjacent[1])] if(
                adjacent[1]) else data.iloc[adjacent[0]:]
        np_data = np.array(data)
        product = making_oscn(np_data, adjacent=adjacent, use_abs=use_abs,
                   tm_coln=tm_coln, opr_coln=opr_coln, 
                   opr_columns=np_opr_columns, tm_col=tm_col,
                   return_full_data=True, return_oscns_data=True,
                   return_na_columns=True, **kwags)
        na_columns = [data.columns[c] for c in product['na_columns']]
        new_columns = ('----'.join(['%s----%s'%(str(c), 
                str(c)+'_oscn%s'%(
                '[%d,%d]'%adjacent if(is_adjacent_memo) else '')) if(
                c in opr_columns and 
                not c in na_columns) else str(
                        c) for c in data.columns])).split('----')
        new_data = asnumeric(pd.DataFrame(product['new_data'], 
                                columns = new_columns, 
                                index=new_data.index))
        new_oscn_columns = [str(data.columns[c])+'_oscn%s'%(
                '[%d,%d]'%adjacent if(is_adjacent_memo) else '')
                for c in reversed(np_opr_columns) if not data.columns[c] in na_columns]
        np_oscns = pd.DataFrame(product['oscn_data'],
                                columns = new_oscn_columns, 
                                index=new_data.index)
    if(return_full_data):
        ret.update({'new_data':new_data})
    if(return_oscns_data):
        ret.update({'oscn_data':np_oscns})
    if(return_na_columns):
        ret.update({'na_columns':na_columns})
    return ret

def making_pulse(data, tm_coln=0, opr_coln='', opr_columns=None, tm_col=[],
                 return_full_data=True, return_pulse_data=False, 
                 return_na_columns=False, **kwags):
    ret = {}
    new_data, pulse_data, na_columns = pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    str_type_data = str(type(data))
    if(str_type_data.find('ndarray')>-1):
        opr_coln = opr_coln if(opr_coln) else 1
        tm_col = dcp(tm_col) if(np.array(tm_col).shape[0]==np.array(
                data).shape[0]) else range(data.shape[0])
        new_data = data[1:,:]
        if(opr_columns==[]):
            tm_col_exc = data[:,tm_coln] if(tm_coln!=None) else tm_col
            numerators = (data[:,opr_coln] - transition(
                    data[:,opr_coln], step=1))[1:]
            denominators = (tm_col_exc - (transition(tm_col_exc, step=1)))[1:]
            if((denominators==0).any()):
                LOGger.addlog('detect zeros!!! pos:%s'%(str(np.where(denominators==0)[0])), 
                              stamps=[making_pulse.__name__, tm_coln], logfile='')
                LOGger.addlog('tm roots:%s'%(str(tm_col_exc)[:200]), 
                              logfile='')
            pulse_data = numerators/denominators
            new_data = np.insert(new_data, opr_coln+1, pulse_data, axis=1)
            np_pulses = pulse_data
        else:
            if(opr_columns=='full'):
                opr_columns = [v for v in range(data.shape[1]) if(
                        v!=tm_coln if(tm_coln!=None) else True)]
            np_pulses, na_columns = [], []
            for opr_column in sorted(opr_columns, reverse=True):#反著過來作就不會改變前面的順序 ─ opr_coln的指定
                try:
                    data[:,opr_column] = data[:,opr_column].astype(float)
                except:
                    printer('[%s]存在非數字元素'%opr_column)
                    na_columns.append(opr_column)
                else:
                    pulse_data = making_pulse(
                            data, tm_coln=tm_coln, opr_coln=opr_coln, 
                            opr_columns=[], tm_col=tm_col,
                            return_full_data=False, return_pulse_data=True, 
                            **kwags)['pulse_data']
                    new_data = np.insert(new_data, opr_column+1, pulse_data, axis=1)
                    np_pulses.append(pulse_data)
            np_pulses = np.transpose(np.array(np_pulses))
    elif(str_type_data.find('pandas.core')>-1):
        opr_coln = list(data.columns).index(opr_coln) if(opr_coln) else 1
        data_columns = dcp(data.columns)
        if(tm_coln in data.columns):
            tm_ordered_columns = [tm_coln]+[v for v in data.columns if v!=tm_coln]
            data = data[tm_ordered_columns]
            tm_coln = 0
        else:
            tm_coln = None
        opr_columns = [v for v in data.columns if(
                v!=data.columns[tm_coln] if(tm_coln!=None) else True)] if(
                opr_columns=='full') else  opr_columns
        np_opr_columns = [list(data.columns).index(c) for c in opr_columns]
        tm_col = dcp(tm_col) if(np.array(tm_col).shape[0]>0) else []
        new_data = data.iloc[1:]
        np_data = np.array(data)
        product = making_pulse(
                    np_data, tm_coln=tm_coln, opr_coln=opr_coln, 
                    opr_columns=np_opr_columns, tm_col=tm_col,
                    return_full_data=True, return_pulse_data=True,
                    return_na_columns=True, **kwags)
        na_columns = [data.columns[c] for c in product['na_columns']]
        new_columns = ('----'.join(['%s----%s'%(str(c), str(c)+'_p') if(
                c in opr_columns and not c in na_columns) else str(
                        c) for c in data.columns])).split('----')
        new_data = asnumeric(pd.DataFrame(product['new_data'], 
                                columns = new_columns, 
                                index=new_data.index))
        new_columns = ('----'.join(['%s----%s'%(str(c), str(c)+'_p') if(
                c in opr_columns and not c in na_columns) else str(
                        c) for c in data_columns])).split('----')
        new_data = new_data[new_columns]
        new_pulse_columns = [str(c)+'_p'
                for c in reversed(opr_columns) if not c in na_columns]
        np_pulses = pd.DataFrame(product['pulse_data'],
                                columns = new_pulse_columns, 
                                index=new_data.index)
    else:
        LOGger.addlog('data type disitnguish!!!!', stamps=[making_pulse.__name__], logfile='')
    if(return_full_data):
        ret.update({'new_data':new_data})
    if(return_pulse_data):
        ret.update({'pulse_data':np_pulses})
    if(return_na_columns):
        ret.update({'na_columns':na_columns})
    return ret

def big_data_attaching(array, data, axis=0, 
                       attach_direction=1, new_index=[], **kwags):
    new_data = data.copy()
    if(axis==0):
        pd_array = (pd.DataFrame(np.array(array), 
                        columns=new_index, 
                        index=array.columns).T if(
                    len(array.shape)==1) else pd.DataFrame(
                        np.array(array), 
                        index=new_index, 
                        columns=array.columns)) if(
                                new_index!=[]) else array
        new_data = pd_array.copy() if(data.empty) else (
                new_data.append(pd_array, sort=False).copy() if(
                attach_direction) else pd_array.append(
                        new_data, sort=False).copy())
    elif(axis==1):
        new_data = new_data.T
        new_data = big_data_attaching(array, new_data, axis=0, 
                      attach_direction=attach_direction, 
                      new_index=new_index**kwags)
        new_data = new_data.T
    return new_data

def partial_vrb_analysis(data, columns='full', 
                         outcome = {'mdn':np.median, 'std':np.std},
                         return_things=['simple'], 
                         attach_direction=1, **kwags):
    ret = {}
    str_type = str(type(data))
    if(str_type.find('pandas.core')>-1):
        columns = data.columns if(columns=='full') else columns
        new_data = data.copy()
        big_datas = {}
        for k in outcome:
            operation = outcome[k]
            ret_k = pd.DataFrame(data[columns].apply(
                operation, axis=0), columns=[k]).T
            if('simple' in return_things):
                ret[k] = ret_k.copy()
            new_data = data.append(ret_k, sort=False).copy() if(
                        attach_direction) else ret_k.append(
                            data, sort=False).copy()
            if('big_datas' in kwags):
                if(not k in kwags['big_datas']):
                    continue
                new_common_index = list(kwags['new_common_index']) if(
                        'new_common_index' in kwags) else []
                big_data = kwags['big_datas'][k].copy()
                big_datas[k] = big_data_attaching(
                        ret_k, big_data, 
                        new_index = [str(v)+'_%s'%k
                         for v in new_common_index], **kwags).copy()
        if('attached' in return_things):
            ret['new_data'] = new_data
        if('big_datas' in kwags):
            ret['big_datas'] = dict(big_datas)
    if(str_type.find('ndarray')>-1):
        pd_data = pd.DataFrame(data)
        ret = partial_vrb_analysis(pd_data, columns=columns, **kwags)
        for k in ret:
            ret[k] = np.array(ret[k][0]) if(
                    ret[k].shape[0]==1) else np.array(ret[k])
    return ret

def data_na_process(data, outer_alog=None, n_show_columns=0, 
                    show_detial=True, return_na_data=False, **kwags):
    na_columns = list(data.columns[data.isna().any()])
    data_na_columns = data[na_columns]
    data_na = data[data_na_columns.T.isna()]
    data_nonna = data[
            np.logical_not(data_na_columns.T.isna().any())]
    log = 'nonna shape:%s'%str(data_nonna.shape)
    outer_alog(log) if(outer_alog) else printer(log)
    log = 'na values!! in:%s'%str(na_columns)
    outer_alog(log) if(outer_alog) else printer(log)
    if(show_detial):
        pd.set_option('display.max_columns', n_show_columns)
        data_show = random_arrays(data_na, **kwags)
        log = '\n%s'%str(data_show)
        outer_alog(log) if(outer_alog) else printer(log)
    pd.set_option('display.max_columns', 0)
    return (data_nonna, data_na) if(return_na_data) else data_nonna

def random_arrays(data, n=3, adjacent=(0,0), **kwags):
    rdm_list = rdm.sample(list(data.index), min(n, data.shape[0]))
    mask = str_multi_contains(data.index, rdm_list)
    if(adjacent!=(0,1)):
        backward = adjacent[0]
        forward = adjacent[1]
        mask |= transition(mask, step=forward)
        mask |= transition(mask, step=-backward)
    data = data[mask]
    return data

def find_local_extremes(data, ws=1, direction=1, outer_lbd=None, outer_ubd=None):
    local_maxima = []
    if(direction>0):
        data = np.array(data)
        num_points = len(data)
        if(num_points <= 2*ws):
            return local_maxima
    
        for i in range(ws, num_points - ws):
            window = data[i - ws:i + ws + 1]
            center_value = data[i]
            if isinstance(astype(outer_lbd), float):
                if(center_value < astype(outer_lbd)):
                    continue
            if isinstance(astype(outer_ubd), float):
                if(center_value > astype(outer_ubd)):
                    continue
            if center_value == max(window):
                local_maxima.append((i, center_value))
    elif(direction<0):
        locs, values = tuple(zip(*find_local_extremes(list(map((lambda x:-x), data)), ws=ws)))
        values = list(map((lambda x:-x), values))
        local_maxima = list(zip(*[locs, values]))
    return local_maxima

def array_mirror(array):
    init_array = np.array(array)
    dim = int(init_array.shape[0])
    P = np.vstack([np.roll(
                [0]*(dim-1)+[1], i*-1) for i in range(dim)])
    return np.matmul(init_array, P)

#TODO:special_matrix_rolling_array
#[1,1,1,0] -> [[1,1,1,0],[0,1,1,1],[1,0,1,1],[1,1,0,1]]
def special_matrix_rolling_array(init_roll, roll_dim=None, axis=0, triangularize = 0, direction=1, 
                                 shift=1, **kwags):
    if(isinstance(kwags.get('for_example', None), dict)):
        for_example = kwags['for_example']
        for_example = {'init_roll':[1,1,1,0]} if(for_example=={}) else for_example
        ret = special_matrix_rolling_array(**for_example)
        print('init_roll:\n%s\n->\nret:\n%s'%(str(for_example['init_roll'])[:200], str(ret)[:200]))
        return
    init_roll = np.array(init_roll)
    roll_dim = roll_dim if(not isinstance(roll_dim, type(None))) else init_roll.shape[0]
    if(direction==1):
        M = np.vstack([np.roll(init_roll, i) for i in range(0,roll_dim,shift)])
        if(triangularize>0):
            M = np.triu(M, -1)
        elif(triangularize<0):
            M = np.tril(M, -1)
    elif(direction==-1):
        M = special_matrix_rolling_array(
                init_roll, roll_dim=roll_dim, axis=axis, 
                triangularize = triangularize, **kwags)
        P = np.vstack([np.roll(
                [0]*(roll_dim-1)+[1], i*-1) for i in range(roll_dim)])
        M = np.matmul(M,P)
    return M

def shift_by_buffer(data, buffer=0, 
                    root_column=0, buffers={}, root=[], 
                    opr_columns='full', **kwags):
    new_data = dcp(data)
    str_type = str(type(data))
    if(str_type.find('ndarray')>-1):
        root_k = dcp(root) if(
                np.array(root).shape[0]==data.shape[0]) else np.array(
                        data[:,root_column])
        try:
            root_k.astype(float)
        except:
            printer('there are nas in root!!')
            return np.empty(0)
        if(not ismonotonic(root_k)):
            printer('root not strictly monotonic!!')
            return np.empty(0)
        if(len(data.shape)==2):
            opr_columns = [v for v in range(data.shape[1]) if(
                    v!=root_column)] if(
                    opr_columns=='full') else [v for v in opr_columns if(
                    v!=root_column)]
            for opr_column in opr_columns:
#                                        {False:0}[opr_column==7]
                buffer_ = float(buffers[opr_column]) if(
                        opr_column in buffers) else float(buffer)
                np_abs_d_buffer = np.abs(root_k - buffer_)
                tn_start_index = np.where(
                       np_abs_d_buffer==np.min(np_abs_d_buffer))[0][0]
                old_array = dcp(data[:,opr_column])
                try:
                    old_array = old_array.astype(float)
                except:
                    continue
                else:
                    new_array = transition(data[:,opr_column], 
                           step=data.shape[0]-tn_start_index-1)
                    new_data[:,opr_column] = dcp(new_array)
        elif(len(data.shape)==1 and 
             np.array(root).shape[0]==data.shape[0]):
            stk_data = np.transpose(np.vstack([np.array(root), data]))
            np_new_data = shift_by_buffer(
                    stk_data, root_column=0,
                    opr_columns=[1], buffer=buffer, 
                    buffers=buffers, root=[], **kwags)
            new_data = np_new_data[:,1]
        else:
            printer('data shape error:(%s,%s)'%data.shape)
            return np.empty(0)
    if(str_type.find('pandas.core')>-1):
        np_root_column = list(data.columns).index(root_column) if(
                np.array(root).shape[0]!=data.shape[0]) else None
        opr_columns = data.columns if(
                opr_columns=='full') else opr_columns
        np_opr_columns = [list(
                data.columns).index(v) for v in opr_columns]
        np_buffers = {list(data.columns).index(k):buffers[
                k] for k in buffers if k in data.columns}
        np_data = np.array(data)
        np_new_data = shift_by_buffer(np_data, 
                        root_column=np_root_column,
                        opr_columns=np_opr_columns,
                        buffers=np_buffers, root=root, **kwags)
        if(np_new_data.shape[0]!=data.shape[0]):
            return pd.DataFrame()
        if(str_type.find('frame')>-1):
            new_data = pd.DataFrame(np_new_data, 
                                index = data.index,
                                columns = data.columns)
        else:
            new_data = pd.Series(np_new_data, 
                                index = data.index,
                                name = data.name)
        new_data = asnumeric(new_data)
    return new_data

def batch_rename(path, format_type='%04d'):
    count = 0
    for fname in os.listdir(path):
        new_fname = format_type%count
        print(os.path.join(path, fname))
        os.rename(os.path.join(path, fname), os.path.join(path, new_fname))
        count = count + 1

def addup_duplicated_index_data(data, add_columns=None, **kwags):
     add_columns = add_columns if(isinstance(add_columns, list)) else data.columns
     other_columns = [v for v in data.columns if not v in add_columns]
     new_df = pd.DataFrame(index=add_columns)
     uni_index = np.unique(data.index)
     for index in uni_index:
         source_part = data[data.index==index].copy()
         np_source_part = np.sum(source_part[add_columns], axis=0)
         new_df[index] = np_source_part
     new_df = new_df.T
     data_undup = data[np.logical_not(data.index.duplicated(keep='first'))|np.logical_not(data.index.duplicated(keep=False))]
     new_df = new_df.join(data_undup[other_columns], sort=False).copy()
     return new_df[np.logical_not(new_df.T.isna().any())]
     
def copy_by_inlier_case(data, total_counts_header, inlier_counts_header, inlier_header='inlier', index_sep='...',
                        do_pop_counts_header=True):
    new_data = pd.DataFrame(columns=list(tuple(data.columns))+[inlier_header])
    for iloc, index in enumerate(data.index):
        total_count = astype(data[total_counts_header][index], d_type=int, default=0)
        inlier_count = astype(data[inlier_counts_header][index], d_type=int, default=None)
        if(inlier_count==None):
            continue
        outlier_count = total_count - inlier_count
        if(inlier_count>0):
            np_data_index = np.array((tuple(data.iloc[iloc]) + (1,))*inlier_count).reshape(inlier_count,-1)
            pd_index_data = pd.DataFrame(
                np_data_index, index=[str(index)+'%sinlier#%d'%(index_sep, i) for i in range(inlier_count)],
                columns=list(tuple(data.columns))+[inlier_header])
            new_data = new_data.append(pd_index_data, sort=False).copy()
        if(outlier_count>0):
            np_data_index = np.array((tuple(data.iloc[iloc]) + (0,))*outlier_count).reshape(outlier_count,-1)
            pd_index_data = pd.DataFrame(
                np_data_index, index=[str(index)+'%soutlier#%d'%(index_sep, i) for i in range(outlier_count)],
                columns=list(tuple(data.columns))+[inlier_header])
            new_data = new_data.append(pd_index_data, sort=False).copy()
    new_data.pop(total_counts_header) if(do_pop_counts_header) else None
    new_data.pop(inlier_counts_header) if(do_pop_counts_header) else None
    return new_data
            
def copy_by_inlier_ratio(data, ground_unit, ratio_header, inlier_header='inlier', index_sep='...',
                         do_pop_ratio_header=True, log_rate=1.0001, **kwags):
    addlog = kwags.get('addlog', LOGger.addloger(logfile=''))
    new_data = pd.DataFrame(columns=list(tuple(data.columns))+[inlier_header])
    for iloc, index in enumerate(data.index):
        ratio = astype(data[ratio_header][index], default=None)
        if(ratio==None):
            continue
        if(ratio==0):
            outlier_count = 1
            inlier_count = 0
        elif(int(ratio*ground_unit)==0):
            outlier_count = dcp(ground_unit)
            inlier_count = 1
        else:
            inlier_count = int(ratio*ground_unit)
            gcd = np.gcd(ground_unit, int(ratio*ground_unit))
            outlier_count = int((ground_unit - inlier_count)/gcd)
            inlier_count = int(inlier_count/gcd)
        if(inlier_count>0):
            np_data_index = np.array((tuple(data.iloc[iloc]) + (1,))*inlier_count).reshape(inlier_count,-1)
            pd_index_data = pd.DataFrame(
                np_data_index, index=[str(index)+'%sinlier#%d'%(index_sep, i) for i in range(inlier_count)],
                columns=list(tuple(data.columns))+[inlier_header])
            new_data = new_data.append(pd_index_data, sort=False).copy()
        if(outlier_count>0):
            np_data_index = np.array((tuple(data.iloc[iloc]) + (0,))*outlier_count).reshape(outlier_count,-1)
            pd_index_data = pd.DataFrame(
                np_data_index, index=[str(index)+'%soutlier#%d'%(index_sep, i) for i in range(outlier_count)],
                columns=list(tuple(data.columns))+[inlier_header])
            new_data = new_data.append(pd_index_data, sort=False).copy()
        addlog('running....', stamps={'iloc':iloc}) if(int(np.log(iloc+1)/np.log(log_rate))%3==1) else None 
    new_data.pop(ratio_header) if(do_pop_ratio_header) else None
    return new_data

def save_as_excel(data, file, wrt=None, rewrite=True, save=False, stack_direction=-1, **kwags):
    sheet_name=kwags.get('sheet_name', kwags.get('sheet', kwags.get('sht',0)))
    wrt = pd.ExcelWriter(file, engine=kwags.get('engine', 'xlsxwriter')) if(wrt==None) else wrt
    if(not rewrite):
        old_data = pd.read_excel(file, sheet_name=sheet_name, index_col=kwags.get('index_col',0))
        data = old_data.append(data, sort=False) if(stack_direction<=0) else data.append(old_data, sort=False)
    data.to_excel(wrt, sheet_name=sheet_name)
    wrt.save() if(save) else None

def is_parallel_rectlike(box, parallel_error_thd=0.005):
    np_box = np.array(box)
    if(np_box.shape!=(4,2)):
        return False
    c1 = np.inner(box[0,:] - box[1,:], box[2,:] - box[3,:])/(np.sqrt(np.sum((box[0,:] - box[1,:])**2))*np.sqrt(np.sum((box[2,:] - box[3,:])**2)))
    c2 = np.inner(box[0,:] - box[2,:], box[1,:] - box[3,:])/(np.sqrt(np.sum((box[0,:] - box[2,:])**2))*np.sqrt(np.sum((box[1,:] - box[3,:])**2)))
    c3 = np.inner(box[0,:] - box[3,:], box[1,:] - box[2,:])/(np.sqrt(np.sum((box[0,:] - box[3,:])**2))*np.sqrt(np.sum((box[1,:] - box[2,:])**2)))
    return not (np.array(sorted(np.abs(np.array([c1,c2,c3]) - 1))[:2]) > parallel_error_thd).any()
    
def is_rectlike(box, perpendicular_thd=0.05, parallel_error_thd=0.005):
    np_box = np.array(box)
    if(np_box.shape!=(4,2)):
        return False
    c1 = np.inner(box[0,:] - box[1,:], box[0,:] - box[2,:])
    c2 = np.inner(box[0,:] - box[1,:], box[0,:] - box[3,:])
    c3 = np.inner(box[0,:] - box[2,:], box[0,:] - box[3,:])
    return (np.abs([c1,c2,c3]) < perpendicular_thd).any() and is_parallel_rectlike(box, parallel_error_thd=parallel_error_thd)
    
def edgelength_of_rectlike(box, perpendicular_thd=0.05, parallel_error_thd=0.005):
    np_box = np.array(box)
    if(not is_rectlike(box, perpendicular_thd=perpendicular_thd, parallel_error_thd=parallel_error_thd)):
        return np.nan, np.nan
    #平面中的近似矩形中，兩點能形成的最短距離就是短邊，兩點能形成的最長距離就是對角線長，透過畢氏定理得長邊
    all_edges = []
    for i in range(np_box.shape[0]):
        for j in range(i):
            np_box_i = np_box[i,:]
            np_box_j = np_box[j,:]
            edge = np.sqrt(np.sum((np_box_i - np_box_j)**2))
            all_edges.append(edge)
    cross_length = np.max(all_edges)
    b_length = np.min(all_edges)
    a_length = np.sqrt(np.max(all_edges)**2 - np.min(all_edges)**2)
    return a_length, b_length, cross_length

def polygon_area(box):
    np_box = np.array(box)
    if(np_box.shape[1]!=2):
        return False
    n = np_box.shape[0]
    area = 0.0
    for i in range(n):
        j = (i + 1) % n
        area += np_box[i, 0] * np_box[j, 1]
        area -= np_box[j, 0] * np_box[i, 1]
    area = abs(area) / 2.0
    return area

def polygon_area_sequential(*box):
    return polygon_area(box)

def statistics_operations_1d(data, header=None, mask=None, stamps=None, ret=None, auto_oper_on_nums=True, fake_nums=[np.nan],
                             operations=None, exception_params=['logfile'], rewrite=False, uniquifying_method=m_uniq_thru_set, **kwags):
    '''
    Parameters
    ----------
    data : TYPE
        DESCRIPTION.
    header : TYPE, optional
        DESCRIPTION. The default is None.
    mask : TYPE, optional
        DESCRIPTION. The default is None.
    stamps : TYPE, optional
        DESCRIPTION. The default is None.
    ret : TYPE, optional
        DESCRIPTION. The default is None. key order: *stamps -> hd -> mask_class
    auto_oper_on_nums : TYPE, optional
        DESCRIPTION. The default is True.
    fake_nums : TYPE, optional
        DESCRIPTION. The default is [np.nan].
    operations : TYPE, optional
        DESCRIPTION. The default is None.
    exception_params : TYPE, optional
        DESCRIPTION. The default is ['logfile'].
    rewrite : TYPE, optional
        DESCRIPTION. The default is False.
    **kwags : TYPE
        DESCRIPTION.

    Returns
    -------
    None.

    '''
    stamps = stamps if(isinstance(stamps, list)) else []
    if(header==None and hasattr(data, 'columns')):
        print('header:\n', *list(data.columns), sep=',')
        for hd in data.columns:
            statistics_operations_1d(data, header=hd, mask=mask, ret=ret, stamps=stamps+[hd], 
                                     auto_oper_on_nums=auto_oper_on_nums, fake_nums=fake_nums,
                                     operations=operations, exception_params=exception_params, rewrite=rewrite, 
                                     uniquifying_method=uniquifying_method, **kwags)
        return
    data_count = np.array(data).shape[0]
    if(isinstance(mask, np.ndarray) or isinstance(mask, pd.core.series.Series)):
        if(mask.shape[0]==data_count):
            classes = uniquifying_method(getattr(mask,'values',mask))
            print('classes:')
            print(*classes, sep=',')
            for aclass in classes:
                data_m = data[mask==aclass].copy()
                statistics_operations_1d(data_m, header=header, stamps=stamps+[parse(aclass)], ret=ret, 
                                         auto_oper_on_nums=auto_oper_on_nums, fake_nums=fake_nums,
                                         operations=operations, exception_params=exception_params, rewrite=rewrite, 
                                         uniquifying_method=uniquifying_method, **kwags)
            return
    main_seq = data[header] if(header!=None) else data
    try:
        if(auto_oper_on_nums):
            main_seq = np.array(main_seq)[np.array(tuple(map((lambda x: not x in fake_nums if(astype(x)!=None) else False), main_seq)))]
        operations = operations if(isinstance(operations, list)) else [LOGger.statistics_properties]
        for oi, operation in enumerate(operations):
            key = LOGger.stamp_process('',stamps+([] if(len(operations)==1) else [getattr(operation, '__name__', 'o%d'%oi)]),'','','','_')
            if(not rewrite):
                if(key in ret):
                    continue
            ret[key] = dcp(operation(main_seq))
    except Exception as e:
        exception_kwags = {k:v for k,v in kwags.items() if k in exception_params}
        LOGger.exception_process(e, stamps=stamps, **exception_kwags)

class myAttributeAgent(LOGger.myAttributeAgent):
    def __init__(self, buffer=None, stamps=None, buffer_ubd=None, first_pop_index=0, exp_fd='.', save_types=None,
                 cleaning_waitng_time=1, rewrite=True, operate_method=None, operate_new_header='operated', **kwags):
        self.operate_method = operate_method
        self.operate_new_header = operate_new_header
        save_types = save_types if(isinstance(save_types, list)) else ['csv']
        super().__init__(buffer=buffer, stamps=stamps, buffer_ubd=buffer_ubd, first_pop_index=first_pop_index, 
                         exp_fd=exp_fd, save_types=save_types, cleaning_waitng_time=cleaning_waitng_time, 
                         rewrite=rewrite, **kwags)
     
    def export(self):
        buffer = super().export()
        return pd.DataFrame.from_dict(buffer).T
    
    def export_and_clear(self):
        pd_buffer = dcp(self.export())
        self.clear()
        return pd_buffer
    
    def export_and_operate(self, method, new_header):
        pd_buffer = dcp(self.export())
        pd_buffer = pd_buffer.applymap(astype)
        pd_buffer[new_header] = method(pd_buffer)
        return pd_buffer
        
    def save(self, exp_fd=None, save_types=None, rewrite=True):
        pd_buffer = self.export()
        if(pd_buffer.empty):
            return True
        save_types = save_types if(isinstance(save_types, list)) else ['csv']
        stamp = LOGger.stamp_process('',['attr'] + self.stamps + [dt.now().strftime('%Y%m%d')],'','','','-',for_file=1)
        try:
            save(pd_buffer, exp_fd=exp_fd, fn='%s-file'%stamp, save_types=save_types, rewrite=rewrite)
        except Exception as e:
            self.stop()
            LOGger.exception_process(e, logfile='', stamps=[self.save.__name__]+self.stamps)
            return False
        return True
        
    def save_and_operate(self, method, new_header, exp_fd=None, save_types=None, rewrite=True):
        try:
            pd_buffer = self.export()
            if(pd_buffer.empty):
                return True
            pd_buffer = pd_buffer.applymap(astype)
            pd_buffer[new_header] = method(pd_buffer)
            save_types = save_types if(isinstance(save_types, list)) else ['csv']
            stamp = LOGger.stamp_process('',['attr'] + self.stamps + [dt.now().strftime('%Y%m%d')],'','','','-',for_file=1)
            save(pd_buffer, exp_fd=exp_fd, fn='%s-file'%stamp, save_types=save_types, rewrite=rewrite)
        except Exception as e:
            self.stop()
            LOGger.exception_process(e, logfile='', stamps=[self.save.__name__]+self.stamps)
            return False
        return True
    
    def cleaning(self):
        start_time = dt.now() if(self.cleaning_waitng_time) else None
        while(not self.stop_flag):
            if(self.cleaning_waitng_time):
                if((dt.now() - start_time).total_seconds() < self.cleaning_waitng_time):
                    continue
                start_time = dt.now()
            (self.save(self.exp_fd, self.save_types, rewrite = self.rewrite) if(
                not self.operate_method) else self.save_and_operate(
                    self.operate_method, self.operate_new_header, self.exp_fd, self.save_types, rewrite = self.rewrite)) if(
                self.save_types) else None
            self.clear()
            
class myThreadAgent(LOGger.myThreadAgent):
    pass

class DataStatistics():
    def __init__(self, data=pd.DataFrame(), infrm=None):
        # super().__init__()
        self.infrm = infrm if(isinstance(infrm, dict)) else {}
        self.data = data
        self.temp_data = pd.DataFrame()
        self.data_keys = ['data', 'temp_data']
        self.act_counter = {}
    
    def new(self, *attr_stgs):
        attr_stgs = attr_stgs if(len(attr_stgs)>0) else ['temp_data']
        for attr_stg in attr_stgs:
            setattr(self, attr_stg, pd.DataFrame())
            if(not attr_stg in self.data_keys): self.data_keys.append(attr_stg)
    
    def select(self, asignment, initial_selected=True, operand=(lambda x,y:x&y), attr_stg='data'):
        data = getattr(self, attr_stg)
        output_selected = np.full(data.shape[0], initial_selected)
        for k,v in asignment.items():
            output_selected = operand(output_selected, data[k]==v)
        return data[output_selected]
    
    def set_infrm(self, value, stamps=None, sep='_'):
        stamps = stamps if(isinstance(stamps, list)) else []
        key = LOGger.stamp_process('',stamps,'','','',sep)
        self.infrm[key] = value
    
    def export_astype_or_remain(self, attr_stg='data', **kwags):
        data = getattr(self, attr_stg)
        method = lambda x:astype_or_remain(x, **kwags)
        return data.applymap(method)
        
    def append(self, major='data', minor='temp_data'):
        data = getattr(self, major)
        temp_data = getattr(self, minor) if(isinstance(minor, str)) else minor
        try:
            data = (temp_data if(data.empty) else data.append(temp_data, sort=False)).copy()
            setattr(self, major, data)
            act_counter_key = LOGger.stamp_process('',['append']+([] if(major=='data') else [major]),'','','','_')
            self.act_counter[act_counter_key] = self.act_counter.get(act_counter_key,0) + 1
        except Exception as e:
            LOGger.exception_process(e, logfile='', stamps=[self.append.__name__])
            print('data\n', data)
            print('temp_data\n', temp_data)
        
    def join(self , major='data', minor='temp_data'):
        try:
            data = getattr(self, major)
            temp_data = getattr(self, minor)
            data = (temp_data if(data.empty) else data.join(temp_data, sort=False)).copy()
            setattr(self, major, data)
            act_counter_key = LOGger.stamp_process('',['join']+([] if(major=='data') else [major]),'','','','_')
            self.act_counter[act_counter_key] = self.act_counter.get(act_counter_key,0) + 1
        except Exception as e:
            LOGger.exception_process(e, logfile='', stamps=[self.join.__name__])
            print('data\n', data)
            print('temp_data\n', temp_data)
        
    def statistics_operations_1d(self, header=None, mask=None, stamps=None, **kwags):
        return statistics_operations_1d(self.data, header=header, mask=mask, stamps=stamps, ret=self.infrm, **kwags)
        
class DataFromFile():
    def __init__(self, file, exp_fd=None, exp_fn=None, stamps=None, infrm=None, notnum_header=None, num_header=None):
        """
            exp_fn: don't put extented filename
        """
        # super().__init__()
        self.file = file
        self.data = import_data(file)
        self.exp_fd = exp_fd if(LOGger.isinstance_not_empty(exp_fd, str)) else os.path.dirname(file)
        self.exp_fn = exp_fn if(LOGger.isinstance_not_empty(exp_fn, str)) else os.path.basename(file).split('.')[0]
        self.stamps = stamps if(isinstance(stamps, list)) else []
        self.notnum_header = notnum_header if(isinstance(notnum_header, list)) else self.data.columns[self.data.applymap(lambda x:astype(x)==None).any()]
        self.num_header = num_header if(isinstance(num_header, list)) else [v for v in self.data.columns if v not in self.notnum_header]
        self.stat_host = DataStatistics()
        self.stat_host.data = self.data[self.num_header]
        
    def __str__(self):
        return os.path.basename(self.file).split('.')[0]
    
    def export(self, file=None, exp_fd=None, fn='test', **kwags):
        if(LOGger.isinstance_not_empty(file)):
            exp_fd = os.path.dirname(file)
            fn = os.path.basename(file)
        else:
            exp_fd = exp_fd if(LOGger.isinstance_not_empty(exp_fd, str)) else self.exp_fd
            fn = fn if(LOGger.isinstance_not_empty(fn, str)) else self.exp_fn
        save(self.data, **kwags)